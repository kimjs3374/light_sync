import io
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from modules.auth_decorators import login_required, menu_required
from modules.db_context import get_db
from modules.pagination import make_pagination
from modules.utils import safe_int
from modules.models import ProductCatalog
from modules.services.g2b_catalog_sync import sync_from_g2b
from modules.services.g2b_procurement_sync import sync_daily as sync_procurement_daily, sync_bulk as sync_procurement_bulk

catalog_bp = Blueprint('catalog', __name__)


# --- ACTION_HANDLERS ---

def handle_sync_catalog(db, form, user_name):
    """나라장터 API 동기화 (관리자 전용)"""
    if session.get('role') != 'admin':
        return {'flash': ('권한이 없습니다.', 'danger')}
    result = sync_from_g2b(db)
    total = result['created'] + result['updated']
    if total == 0 and result.get('errors', 0) == 0:
        return {'flash': ('나라장터 API에서 데이터를 받지 못했습니다. 일일 트래픽 초과 또는 API 점검 중일 수 있습니다.', 'warning')}
    msg = f"동기화 완료: 신규 {result['created']}건, 갱신 {result['updated']}건"
    if result.get('errors'):
        msg += f", 오류 {result['errors']}건"
    return {'flash': (msg, 'success')}


def handle_update_price(db, form, user_name):
    """수기 단가 수정 (관리자 전용)"""
    if session.get('role') != 'admin':
        return {'flash': ('권한이 없습니다.', 'danger')}
    catalog_id = safe_int(form.get('catalog_id'))
    new_price = safe_int(form.get('unit_price'))
    if not catalog_id:
        return {'flash': ('대상을 찾을 수 없습니다.', 'warning')}

    item = db.query(ProductCatalog).get(catalog_id)
    if not item:
        return {'flash': ('대상을 찾을 수 없습니다.', 'warning')}

    item.unit_price = new_price if new_price and new_price > 0 else None
    item.price_source = 'manual'
    return {'flash': (f'{item.krn_prdct_nm} 단가 수정 완료', 'success')}


def _build_krn_prdct_nm(form):
    """폼 필드에서 원본 품명(쉼표구분) 조합"""
    parts = [
        (form.get('item_name') or '').strip(),
        (form.get('manufacturer') or '').strip(),
        (form.get('model_name') or '').strip(),
        (form.get('spec') or '').strip(),
    ]
    return ', '.join(p for p in parts if p)


def handle_add_manual(db, form, user_name):
    """수기 제품 등록 (관리자 전용) - 조달물품/견적제품 모두"""
    if session.get('role') != 'admin':
        return {'flash': ('권한이 없습니다.', 'danger')}

    item_name = (form.get('item_name') or '').strip()
    if not item_name:
        return {'flash': ('품목명을 입력하세요.', 'warning')}

    import datetime
    prdct_idnt_no = (form.get('prdct_idnt_no') or '').strip()
    if not prdct_idnt_no:
        prdct_idnt_no = f'M{datetime.datetime.now().strftime("%y%m%d%H%M%S")}'

    existing = db.query(ProductCatalog).filter_by(prdct_idnt_no=prdct_idnt_no).first()
    if existing:
        return {'flash': (f'물품식별번호 {prdct_idnt_no}가 이미 존재합니다.', 'warning')}

    is_quote = form.get('is_quote') == '1'  # 견적제품 여부
    unit_price = safe_int(form.get('unit_price'))
    krn_nm = _build_krn_prdct_nm(form)

    new_item = ProductCatalog(
        prdct_idnt_no=prdct_idnt_no,
        krn_prdct_nm=krn_nm or item_name,
        item_name=item_name,
        manufacturer=(form.get('manufacturer') or '').strip() or None,
        model_name=(form.get('model_name') or '').strip() or None,
        spec=(form.get('spec') or '').strip() or None,
        unit=(form.get('unit') or '').strip() or None,
        unit_price=unit_price if unit_price and unit_price > 0 else None,
        price_source='quote' if is_quote else 'manual',
        g2b_contract_method='견적' if is_quote else '수기등록',
    )
    db.add(new_item)
    label = '견적제품' if is_quote else '조달물품'
    return {'flash': (f'[{label}] "{item_name}" 등록 완료', 'success')}


def handle_edit_item(db, form, user_name):
    """제품 수정 (관리자 전용)"""
    if session.get('role') != 'admin':
        return {'flash': ('권한이 없습니다.', 'danger')}
    catalog_id = safe_int(form.get('catalog_id'))
    item = db.query(ProductCatalog).get(catalog_id)
    if not item:
        return {'flash': ('대상을 찾을 수 없습니다.', 'warning')}

    item.item_name = (form.get('item_name') or '').strip() or item.item_name
    item.manufacturer = (form.get('manufacturer') or '').strip() or None
    item.model_name = (form.get('model_name') or '').strip() or None
    item.spec = (form.get('spec') or '').strip() or None
    item.unit = (form.get('unit') or '').strip() or None
    unit_price = safe_int(form.get('unit_price'))
    item.unit_price = unit_price if unit_price and unit_price > 0 else None
    item.krn_prdct_nm = _build_krn_prdct_nm(form) or item.krn_prdct_nm

    prdct_no = (form.get('prdct_idnt_no') or '').strip()
    if prdct_no and prdct_no != item.prdct_idnt_no:
        dup = db.query(ProductCatalog).filter_by(prdct_idnt_no=prdct_no).first()
        if dup and dup.id != item.id:
            return {'flash': (f'물품식별번호 {prdct_no}가 이미 존재합니다.', 'warning')}
        item.prdct_idnt_no = prdct_no

    is_quote = form.get('is_quote') == '1'
    if is_quote:
        item.price_source = 'quote'
        item.g2b_contract_method = '견적'
    elif item.price_source not in ('api',):
        item.price_source = 'manual'

    return {'flash': (f'"{item.item_name}" 수정 완료', 'success')}


def handle_delete_item(db, form, user_name):
    """제품 삭제 (관리자 전용)"""
    if session.get('role') != 'admin':
        return {'flash': ('권한이 없습니다.', 'danger')}
    catalog_id = safe_int(form.get('catalog_id'))
    item = db.query(ProductCatalog).get(catalog_id)
    if not item:
        return {'flash': ('대상을 찾을 수 없습니다.', 'warning')}
    name = item.krn_prdct_nm
    db.delete(item)
    return {'flash': (f'"{name}" 삭제 완료', 'success')}


def handle_sync_procurement(db, form, user_name):
    """나라장터 조달내역 동기화 (관리자 전용)"""
    if session.get('role') != 'admin':
        return {'flash': ('권한이 없습니다.', 'danger')}
    mode = form.get('procurement_mode', 'daily')
    if mode == 'bulk':
        result = sync_procurement_bulk(db, start_year=2020)
    else:
        result = sync_procurement_daily(db)
    total = result['created'] + result['updated']
    if total == 0 and result.get('errors', 0) == 0:
        msg = '조달내역 동기화: 새로운 데이터가 없습니다.' if mode == 'daily' else '조달내역 벌크 동기화: 데이터를 받지 못했습니다.'
        return {'flash': (msg, 'info')}
    msg = f"조달내역 동기화 완료: 신규 {result['created']}건, 갱신 {result['updated']}건"
    if result.get('errors'):
        msg += f", 오류 {result['errors']}건"
    return {'flash': (msg, 'success')}


ACTION_HANDLERS = {
    'sync_catalog': handle_sync_catalog,
    'sync_procurement': handle_sync_procurement,
    'update_price': handle_update_price,
    'add_manual': handle_add_manual,
    'edit_item': handle_edit_item,
    'delete_item': handle_delete_item,
}


# --- ROUTES ---

@catalog_bp.route('/product_catalog', methods=['GET'])
@login_required
@menu_required('catalog')
def catalog_list():
    """제품 카탈로그 목록"""
    q = (request.args.get('q') or '').strip().lower()
    price_source = request.args.get('price_source', '')
    method = request.args.get('method', '')
    page = safe_int(request.args.get('page'), 1)
    per_page = safe_int(request.args.get('per_page'), 30)

    with get_db() as db:
        query = db.query(ProductCatalog)

        if q:
            query = query.filter(
                (ProductCatalog.item_name.ilike(f'%{q}%')) |
                (ProductCatalog.model_name.ilike(f'%{q}%')) |
                (ProductCatalog.prdct_idnt_no.ilike(f'%{q}%')) |
                (ProductCatalog.krn_prdct_nm.ilike(f'%{q}%')) |
                (ProductCatalog.spec.ilike(f'%{q}%'))
            )

        if price_source == 'api':
            query = query.filter(ProductCatalog.price_source == 'api')
        elif price_source == 'manual':
            query = query.filter(ProductCatalog.price_source == 'manual')
        elif price_source == 'quote':
            query = query.filter(ProductCatalog.price_source == 'quote')
        elif price_source == 'missing':
            query = query.filter(ProductCatalog.unit_price.is_(None))

        if method:
            query = query.filter(ProductCatalog.g2b_contract_method == method)

        total = query.count()
        pagination = make_pagination(page, per_page, total)
        offset = (pagination['page'] - 1) * per_page

        items = query.order_by(ProductCatalog.krn_prdct_nm).offset(offset).limit(per_page).all()

        # 통계
        total_all = db.query(ProductCatalog).count()
        missing_count = db.query(ProductCatalog).filter(
            ProductCatalog.unit_price.is_(None)
        ).count()
        last_sync = db.query(ProductCatalog.last_synced_at).filter(
            ProductCatalog.last_synced_at.isnot(None)
        ).order_by(ProductCatalog.last_synced_at.desc()).first()

        stats = {
            'total': total_all,
            'filtered': total,
            'missing_price': missing_count,
            'has_price': total_all - missing_count,
            'last_synced': last_sync[0].strftime('%Y-%m-%d %H:%M') if last_sync and last_sync[0] else '-',
        }

    filters = {'q': q, 'price_source': price_source, 'method': method}
    return render_template(
        'catalog_list.html',
        catalog_items=items,
        catalog_stats=stats,
        catalog_pagination=pagination,
        catalog_filters=filters,
        # 하위 호환 (partial 외부에서 사용 가능)
        items=items, stats=stats, pagination=pagination, filters=filters,
        is_admin=(session.get('role') == 'admin'),
    )


@catalog_bp.route('/product_catalog', methods=['POST'])
@login_required
@menu_required('catalog')
def catalog_action():
    """POST 액션 처리 (sync_catalog, update_price)"""
    action = request.form.get('action')
    user_name = session.get('full_name') or '사용자'

    with get_db() as db:
        handler = ACTION_HANDLERS.get(action)
        if handler:
            result = handler(db, request.form, user_name)
            if result.get('flash'):
                flash(*result['flash'])
            db.commit()

    return redirect(url_for('catalog.catalog_list'))


@catalog_bp.route('/product_catalog/excel')
@login_required
@menu_required('catalog')
def catalog_excel():
    """카탈로그 전체 엑셀 다운로드"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with get_db() as db:
        items = db.query(ProductCatalog).order_by(ProductCatalog.krn_prdct_nm).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '제품카탈로그'

        headers = ['No', '물품식별번호', '품목명', '제조사', '모델명', '규격', '단위', '단가(원)', '단가출처', '계약방식', '계약번호', '계약시작일', '계약종료일']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        source_map = {'api': 'API', 'manual': '수기입력', 'quote': '견적'}
        for r, item in enumerate(items, 2):
            row_data = [
                r - 1,
                item.prdct_idnt_no or '',
                item.item_name or item.krn_prdct_nm or '',
                item.manufacturer or '',
                item.model_name or '',
                item.spec or '',
                item.unit or '',
                item.unit_price or '',
                source_map.get(item.price_source, item.price_source or ''),
                item.g2b_contract_method or '',
                item.g2b_cntrct_no or '',
                item.cntrct_bgn_date.strftime('%Y-%m-%d') if item.cntrct_bgn_date else '',
                item.cntrct_end_date.strftime('%Y-%m-%d') if item.cntrct_end_date else '',
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = thin_border
                cell.font = Font(size=9)
            # 단가 숫자 포맷
            price_cell = ws.cell(row=r, column=8)
            if isinstance(price_cell.value, int) and price_cell.value:
                price_cell.number_format = '#,##0'

        widths = [6, 16, 28, 14, 20, 18, 8, 14, 10, 12, 18, 14, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(headers))}{len(items) + 1}'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

    from datetime import date
    filename = f'제품카탈로그_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
