"""
재고관리 Blueprint (Inventory Management).

- 재고 현황 대시보드
- 가용재고 조회
- 재고실사 (생성/상세/저장/확정)
- 재고 변동 이력
- 재고회전율 분석
- 엑셀 다운로드
- 수동 재고 조정 / 안전재고 설정
"""

import datetime
import io
import logging

from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, jsonify, abort, session, send_file,
)
from sqlalchemy import desc, func, distinct, or_
from modules.auth_decorators import login_required
from modules.pagination import make_pagination
from modules.utils import safe_int
from modules.db_context import get_db
from modules.models import (
    Item, StockAudit, StockAuditItem, StockMovement,
    MOVEMENT_TYPE_LABELS, BomHeader, BomItem,
)
from modules.services.inventory_utils import (
    record_stock_movement, confirm_audit, calc_turnover_rate,
)

logger = logging.getLogger(__name__)

inventory_bp = Blueprint('inventory', __name__)


def _fmt_money(val):
    if val is None:
        return '0'
    try:
        return f"{int(val):,}"
    except (ValueError, TypeError):
        return str(val)


def _generate_audit_no(db):
    """실사번호 자동 채번: SA{YYYY}-{NNN}"""
    year = datetime.date.today().year
    prefix = f'SA{year}-'
    last = db.query(StockAudit).filter(
        StockAudit.audit_no.like(f'{prefix}%')
    ).order_by(desc(StockAudit.audit_no)).first()
    if last:
        try:
            last_num = int(last.audit_no.replace(prefix, ''))
            return f'{prefix}{last_num + 1:03d}'
        except ValueError:
            pass
    return f'{prefix}001'


# ===================================================================
# 1. 재고 현황 대시보드
# ===================================================================
@inventory_bp.route('/inventory')
@login_required
def inventory_dashboard():
    with get_db() as db:
        # 전체 품목 (활성)
        items = db.query(Item).filter(Item.is_active == True).all()

        total_items = len(items)
        total_stock_value = 0
        total_available_value = 0
        total_reserved_value = 0
        low_stock_count = 0
        low_stock_items = []

        # 카테고리별 집계
        cat_summary = {}

        for item in items:
            stock = item.stock_qty or 0
            reserved = item.reserved_qty or 0
            available = max(0, stock - reserved)
            price = item.last_unit_price or 0
            stock_val = stock * price
            available_val = available * price
            reserved_val = reserved * price

            total_stock_value += stock_val
            total_available_value += available_val
            total_reserved_value += reserved_val

            # 안전재고 미달
            safety = item.safety_stock or 0
            if safety > 0 and stock < safety:
                low_stock_count += 1
                low_stock_items.append({
                    'id': item.id,
                    'name': item.item_name,
                    'code': item.icube_item_cd or '',
                    'stock_qty': stock,
                    'safety_stock': safety,
                    'shortage': safety - stock,
                    'category': item.category or '',
                })

            # 카테고리 집계
            cat = item.category or '미분류'
            if cat not in cat_summary:
                cat_summary[cat] = {
                    'category': cat,
                    'count': 0, 'total_stock': 0, 'reserved': 0,
                    'available': 0, 'value': 0,
                }
            cat_summary[cat]['count'] += 1
            cat_summary[cat]['total_stock'] += stock
            cat_summary[cat]['reserved'] += reserved
            cat_summary[cat]['available'] += available
            cat_summary[cat]['value'] += stock_val

        # 최근 변동 10건
        recent_movements = db.query(StockMovement).order_by(
            desc(StockMovement.created_at)
        ).limit(10).all()

        # 저재고 경고 상위 10건 (부족분 내림차순)
        low_stock_items.sort(key=lambda x: x['shortage'], reverse=True)
        low_stock_items = low_stock_items[:10]

        cat_list = sorted(cat_summary.values(), key=lambda x: x['value'], reverse=True)

        return render_template(
            'inventory_dashboard.html',
            total_items=total_items,
            total_stock_value=total_stock_value,
            total_available_value=total_available_value,
            total_reserved_value=total_reserved_value,
            low_stock_count=low_stock_count,
            low_stock_items=low_stock_items,
            cat_summary=cat_list,
            recent_movements=recent_movements,
            fmt_money=_fmt_money,
            movement_labels=MOVEMENT_TYPE_LABELS,
        )


# ===================================================================
# 1-1. BOM 기준 가용재고 조회
# ===================================================================
@inventory_bp.route('/inventory/bom-stock')
@login_required
def bom_stock():
    """제품 모델명 → BOM 소요부품별 가용재고 조회"""
    q = (request.args.get('q') or '').strip()
    bom_id = safe_int(request.args.get('bom_id'), 0)

    with get_db() as db:
        # BOM 목록 (자동완성용)
        bom_list = db.query(BomHeader).filter(BomHeader.is_active == True)\
            .order_by(BomHeader.product_name).all()

        selected_bom = None
        bom_items_data = []
        producible_qty = None  # 생산 가능 수량

        if bom_id:
            selected_bom = db.query(BomHeader).get(bom_id)
        elif q:
            selected_bom = db.query(BomHeader).filter(
                or_(BomHeader.product_name.ilike(f'%{q}%'),
                    BomHeader.product_code.ilike(f'%{q}%'))
            ).first()

        if selected_bom:
            items = db.query(BomItem).filter(BomItem.bom_id == selected_bom.id)\
                .order_by(BomItem.id).all()

            min_producible = None
            for bi in items:
                item = bi.item if bi.item_id else None
                if not item and bi.item_code:
                    item = db.query(Item).filter(Item.icube_item_cd == bi.item_code).first()

                stock_qty = (item.stock_qty or 0) if item else 0
                reserved_qty = (item.reserved_qty or 0) if item else 0
                available = stock_qty - reserved_qty
                needed = bi.quantity or 0
                can_make = int(available / needed) if needed > 0 else 999999

                if needed > 0:
                    if min_producible is None or can_make < min_producible:
                        min_producible = can_make

                bom_items_data.append({
                    'bom_item': bi,
                    'item': item,
                    'stock_qty': stock_qty,
                    'reserved_qty': reserved_qty,
                    'available': available,
                    'needed': needed,
                    'enough': available >= needed,
                    'can_make': can_make,
                })

            producible_qty = min_producible if min_producible is not None else 0

        return render_template(
            'inventory_bom_stock.html',
            bom_list=bom_list,
            selected_bom=selected_bom,
            bom_items=bom_items_data,
            producible_qty=producible_qty,
            q=q,
            bom_id=bom_id,
        )


@inventory_bp.route('/api/inventory/bom-search')
@login_required
def api_bom_search():
    """BOM 제품명 자동완성 API"""
    q = (request.args.get('q') or '').strip()
    if len(q) < 1:
        return jsonify([])
    with get_db() as db:
        results = db.query(BomHeader).filter(
            BomHeader.is_active == True,
            or_(BomHeader.product_name.ilike(f'%{q}%'),
                BomHeader.product_code.ilike(f'%{q}%'))
        ).limit(15).all()
        return jsonify([{
            'id': r.id,
            'product_code': r.product_code,
            'product_name': r.product_name,
            'product_category': r.product_category or '',
        } for r in results])


# ===================================================================
# 2. 가용재고 조회
# ===================================================================
@inventory_bp.route('/inventory/items')
@login_required
def inventory_items():
    q = (request.args.get('q') or '').strip()
    category = (request.args.get('category') or '').strip()
    stock_status = (request.args.get('stock_status') or '').strip()
    page = safe_int(request.args.get('page'), 1)
    per_page = 50

    with get_db() as db:
        query = db.query(Item).filter(Item.is_active == True)

        if q:
            like = f'%{q}%'
            query = query.filter(or_(
                Item.icube_item_cd.ilike(like),
                Item.item_name.ilike(like),
                Item.item_spec.ilike(like),
            ))
        if category:
            query = query.filter(Item.category == category)

        # stock_status 필터링은 Python에서 처리 (계산 필드)
        all_items = query.order_by(Item.item_name).all()

        # 필터 적용
        filtered = []
        for item in all_items:
            stock = item.stock_qty or 0
            safety = item.safety_stock or 0
            available = max(0, stock - (item.reserved_qty or 0))

            if stock_status == 'low' and (safety <= 0 or stock >= safety):
                continue
            if stock_status == 'excess' and stock <= safety * 2:
                continue
            if stock_status == 'normal' and safety > 0 and stock < safety:
                continue

            filtered.append(item)

        total = len(filtered)
        pagination = make_pagination(page, per_page, total)
        offset = (pagination['page'] - 1) * per_page
        page_items = filtered[offset:offset + per_page]

        categories = [r[0] for r in db.query(distinct(Item.category)).filter(
            Item.category.isnot(None), Item.category != ''
        ).order_by(Item.category).all()]

        return render_template(
            'inventory_items.html',
            items=page_items,
            pagination=pagination,
            q=q,
            category=category,
            stock_status=stock_status,
            categories=categories,
            total=total,
            fmt_money=_fmt_money,
        )


# ===================================================================
# 3. 재고실사 목록
# ===================================================================
@inventory_bp.route('/inventory/audit')
@login_required
def audit_list():
    page = safe_int(request.args.get('page'), 1)
    per_page = 20

    with get_db() as db:
        query = db.query(StockAudit).order_by(desc(StockAudit.audit_date), desc(StockAudit.id))
        total = query.count()
        pagination = make_pagination(page, per_page, total)
        offset = (pagination['page'] - 1) * per_page
        audits = query.offset(offset).limit(per_page).all()

        return render_template(
            'inventory_audit_list.html',
            audits=audits,
            pagination=pagination,
        )


# ===================================================================
# 4. 재고실사 생성
# ===================================================================
@inventory_bp.route('/inventory/audit/create', methods=['GET', 'POST'])
@login_required
def audit_create():
    with get_db() as db:
        if request.method == 'POST':
            audit_date_str = request.form.get('audit_date', '')
            auditor_name = (request.form.get('auditor_name') or '').strip()
            note = (request.form.get('note') or '').strip()
            target_category = (request.form.get('target_category') or '').strip()

            if not auditor_name:
                auditor_name = session.get('full_name', '사용자')

            try:
                audit_date = datetime.datetime.strptime(audit_date_str, '%Y-%m-%d').date()
            except ValueError:
                audit_date = datetime.date.today()

            audit = StockAudit(
                audit_no=_generate_audit_no(db),
                audit_date=audit_date,
                auditor_id=session.get('user_id'),
                auditor_name=auditor_name,
                status='진행중',
                note=note,
            )
            db.add(audit)
            db.flush()

            # 대상 품목 로딩
            item_query = db.query(Item).filter(Item.is_active == True)
            if target_category:
                item_query = item_query.filter(Item.category == target_category)
            target_items = item_query.order_by(Item.item_name).all()

            for item in target_items:
                audit_item = StockAuditItem(
                    audit_id=audit.id,
                    item_id=item.id,
                    system_qty=item.stock_qty or 0,
                )
                db.add(audit_item)

            audit.total_items = len(target_items)
            db.commit()

            flash(f'실사 {audit.audit_no}이 생성되었습니다. ({len(target_items)}개 품목)', 'success')
            return redirect(url_for('inventory.audit_detail', audit_id=audit.id))

        # GET
        categories = [r[0] for r in db.query(distinct(Item.category)).filter(
            Item.category.isnot(None), Item.category != ''
        ).order_by(Item.category).all()]

        return render_template(
            'inventory_audit_create.html',
            categories=categories,
            today=datetime.date.today().strftime('%Y-%m-%d'),
            user_name=session.get('full_name', ''),
        )


# ===================================================================
# 5. 재고실사 상세 (수량 입력)
# ===================================================================
@inventory_bp.route('/inventory/audit/<int:audit_id>')
@login_required
def audit_detail(audit_id):
    with get_db() as db:
        audit = db.query(StockAudit).get(audit_id)
        if not audit:
            abort(404)

        audit_items = db.query(StockAuditItem).filter(
            StockAuditItem.audit_id == audit_id,
        ).all()

        # item 정보 미리 로딩
        item_ids = [ai.item_id for ai in audit_items]
        items_map = {}
        if item_ids:
            for item in db.query(Item).filter(Item.id.in_(item_ids)).all():
                items_map[item.id] = item

        return render_template(
            'inventory_audit_detail.html',
            audit=audit,
            audit_items=audit_items,
            items_map=items_map,
        )


# ===================================================================
# 6. 재고실사 수량 저장 (임시저장)
# ===================================================================
@inventory_bp.route('/inventory/audit/<int:audit_id>/save', methods=['POST'])
@login_required
def audit_save(audit_id):
    with get_db() as db:
        audit = db.query(StockAudit).get(audit_id)
        if not audit or audit.status != '진행중':
            flash('저장할 수 없는 상태입니다.', 'warning')
            return redirect(url_for('inventory.audit_list'))

        audit_items = db.query(StockAuditItem).filter(
            StockAuditItem.audit_id == audit_id,
        ).all()

        saved_count = 0
        for ai in audit_items:
            actual_key = f'actual_{ai.id}'
            reason_key = f'reason_{ai.id}'
            actual_val = request.form.get(actual_key, '').strip()
            reason_val = (request.form.get(reason_key) or '').strip()

            if actual_val != '':
                try:
                    ai.actual_qty = float(actual_val)
                    ai.diff_qty = ai.actual_qty - (ai.system_qty or 0)
                    saved_count += 1
                except (ValueError, TypeError):
                    pass
            ai.diff_reason = reason_val

        db.commit()
        flash(f'{saved_count}건이 저장되었습니다.', 'success')
        return redirect(url_for('inventory.audit_detail', audit_id=audit_id))


# ===================================================================
# 7. 재고실사 조정 확정
# ===================================================================
@inventory_bp.route('/inventory/audit/<int:audit_id>/confirm', methods=['POST'])
@login_required
def audit_confirm(audit_id):
    with get_db() as db:
        audit = db.query(StockAudit).get(audit_id)
        if not audit or audit.status != '진행중':
            flash('확정할 수 없는 상태입니다.', 'warning')
            return redirect(url_for('inventory.audit_list'))

        confirmed_by = session.get('full_name', '사용자')
        count = confirm_audit(db, audit_id, confirmed_by)
        db.commit()

        if count > 0:
            flash(f'실사 조정 확정 완료: {count}건 재고가 조정되었습니다.', 'success')
        else:
            flash('조정할 차이 항목이 없습니다.', 'info')

        return redirect(url_for('inventory.audit_detail', audit_id=audit_id))


# ===================================================================
# 7-0. 실사 삭제
# ===================================================================
@inventory_bp.route('/inventory/audit/<int:audit_id>/delete', methods=['POST'])
@login_required
def audit_delete(audit_id):
    with get_db() as db:
        audit = db.query(StockAudit).get(audit_id)
        if not audit:
            abort(404)
        if audit.status == '완료':
            flash('확정된 실사는 삭제할 수 없습니다.', 'warning')
            return redirect(url_for('inventory.audit_detail', audit_id=audit_id))

        db.query(StockAuditItem).filter(StockAuditItem.audit_id == audit_id).delete()
        db.delete(audit)
        db.commit()
        flash(f'실사 {audit.audit_no}이 삭제되었습니다.', 'success')
        return redirect(url_for('inventory.audit_list'))


# ===================================================================
# 7-1. 실사 엑셀 템플릿 다운로드
# ===================================================================
@inventory_bp.route('/inventory/audit/<int:audit_id>/template-download')
@login_required
def audit_template_download(audit_id):
    """실사 대상 품목 엑셀 다운로드 — 현장에서 실재고 기입용"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with get_db() as db:
        audit = db.query(StockAudit).get(audit_id)
        if not audit:
            abort(404)

        audit_items = db.query(StockAuditItem).filter(
            StockAuditItem.audit_id == audit_id
        ).all()
        item_ids = [ai.item_id for ai in audit_items]
        items_map = {}
        if item_ids:
            for item in db.query(Item).filter(Item.id.in_(item_ids)).all():
                items_map[item.id] = item

        wb = Workbook()
        ws = wb.active
        ws.title = '재고실사'

        # 헤더 스타일
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        yellow_fill = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')

        # 실사 정보 행
        ws.merge_cells('A1:G1')
        ws['A1'] = f'재고실사 [{audit.audit_no}] — {audit.audit_date}'
        ws['A1'].font = Font(bold=True, size=12)

        # 헤더
        headers = ['품목ID', '품번', '품명', '규격', '카테고리', '시스템재고', '실재고(입력)', '비고']
        col_widths = [8, 15, 30, 20, 12, 12, 14, 20]
        for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 8 else 'H'].width = w

        # 데이터
        for row_idx, ai in enumerate(audit_items, 4):
            item = items_map.get(ai.item_id)
            if not item:
                continue

            row_data = [
                item.id,
                item.icube_item_cd or '',
                item.item_name or '',
                item.item_spec or '',
                item.category or '',
                ai.system_qty or 0,
                ai.actual_qty if ai.actual_qty is not None else '',
                '',
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx == 7:  # 실재고 입력 칸
                    cell.fill = yellow_fill
                    cell.font = Font(bold=True)
                if col_idx == 6:  # 시스템재고
                    cell.alignment = Alignment(horizontal='right')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'실사템플릿_{audit.audit_no}_{audit.audit_date}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ===================================================================
# 7-2. 실사 엑셀 업로드 (실재고 반영)
# ===================================================================
@inventory_bp.route('/inventory/audit/<int:audit_id>/upload', methods=['POST'])
@login_required
def audit_upload(audit_id):
    """실재고 기입된 엑셀 업로드 → 실사 항목에 반영"""
    from openpyxl import load_workbook

    with get_db() as db:
        audit = db.query(StockAudit).get(audit_id)
        if not audit or audit.status != '진행중':
            flash('업로드할 수 없는 상태입니다.', 'warning')
            return redirect(url_for('inventory.audit_list'))

        file = request.files.get('audit_file')
        if not file or not file.filename.endswith(('.xlsx', '.xls')):
            flash('엑셀 파일(.xlsx)을 선택해주세요.', 'danger')
            return redirect(url_for('inventory.audit_detail', audit_id=audit_id))

        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            flash(f'엑셀 파일 읽기 실패: {e}', 'danger')
            return redirect(url_for('inventory.audit_detail', audit_id=audit_id))

        # audit_item을 item_id 기준으로 매핑
        audit_items = db.query(StockAuditItem).filter(
            StockAuditItem.audit_id == audit_id
        ).all()
        ai_map = {ai.item_id: ai for ai in audit_items}

        updated = 0
        errors = []
        for row in ws.iter_rows(min_row=4, values_only=False):
            try:
                item_id_val = row[0].value  # A열: 품목ID
                actual_val = row[6].value    # G열: 실재고
                note_val = row[7].value if len(row) > 7 else ''  # H열: 비고

                if item_id_val is None or actual_val is None:
                    continue
                item_id = int(item_id_val)
                actual_qty = float(actual_val)

                ai = ai_map.get(item_id)
                if ai:
                    ai.actual_qty = actual_qty
                    ai.diff_qty = actual_qty - (ai.system_qty or 0)
                    if note_val:
                        ai.note = str(note_val).strip()
                    updated += 1
            except (ValueError, TypeError, IndexError):
                row_num = row[0].row if row[0] else '?'
                errors.append(f'{row_num}행')

        db.commit()
        msg = f'엑셀 업로드 완료: {updated}건 반영'
        if errors:
            msg += f', {len(errors)}건 오류 ({", ".join(errors[:5])})'
        flash(msg, 'success' if updated > 0 else 'warning')
        return redirect(url_for('inventory.audit_detail', audit_id=audit_id))


# ===================================================================
# 7-3. 실사 차이 보고서
# ===================================================================
@inventory_bp.route('/inventory/audit/<int:audit_id>/report')
@login_required
def audit_report(audit_id):
    """실사 차이 분석 보고서 — 인쇄/엑셀 지원"""
    with get_db() as db:
        audit = db.query(StockAudit).get(audit_id)
        if not audit:
            abort(404)

        audit_items = db.query(StockAuditItem).filter(
            StockAuditItem.audit_id == audit_id
        ).all()

        item_ids = [ai.item_id for ai in audit_items]
        items_map = {}
        if item_ids:
            for item in db.query(Item).filter(Item.id.in_(item_ids)).all():
                items_map[item.id] = item

        # 통계 계산
        total_items = len(audit_items)
        counted_items = sum(1 for ai in audit_items if ai.actual_qty is not None)
        diff_items = [ai for ai in audit_items if ai.actual_qty is not None and ai.diff_qty and abs(ai.diff_qty) > 0.001]
        match_items = counted_items - len(diff_items)

        total_system_value = 0
        total_actual_value = 0
        total_diff_value = 0
        for ai in audit_items:
            item = items_map.get(ai.item_id)
            price = (item.last_unit_price or 0) if item else 0
            total_system_value += (ai.system_qty or 0) * price
            if ai.actual_qty is not None:
                total_actual_value += ai.actual_qty * price
                total_diff_value += (ai.diff_qty or 0) * price

        # 엑셀 다운로드 모드
        if request.args.get('format') == 'excel':
            return _audit_report_excel(audit, audit_items, items_map, {
                'total_items': total_items,
                'counted_items': counted_items,
                'diff_items': len(diff_items),
                'match_items': match_items,
                'total_system_value': total_system_value,
                'total_actual_value': total_actual_value,
                'total_diff_value': total_diff_value,
            })

        return render_template(
            'inventory_audit_report.html',
            audit=audit,
            audit_items=audit_items,
            items_map=items_map,
            diff_items=diff_items,
            total_items=total_items,
            counted_items=counted_items,
            match_items=match_items,
            total_system_value=total_system_value,
            total_actual_value=total_actual_value,
            total_diff_value=total_diff_value,
        )


def _audit_report_excel(audit, audit_items, items_map, stats):
    """실사 차이 보고서 엑셀"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '실사보고서'

    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    red_font = Font(color='DC3545', bold=True)
    blue_font = Font(color='0D6EFD', bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 제목
    ws.merge_cells('A1:I1')
    ws['A1'] = f'재고실사 차이 보고서 [{audit.audit_no}]'
    ws['A1'].font = Font(bold=True, size=14)

    # 요약
    ws['A2'] = f'실사일: {audit.audit_date} | 전체: {stats["total_items"]}건 | 실사완료: {stats["counted_items"]}건 | 차이: {stats["diff_items"]}건 | 일치: {stats["match_items"]}건'
    ws['A3'] = f'시스템금액: {int(stats["total_system_value"]):,}원 | 실재고금액: {int(stats["total_actual_value"]):,}원 | 차이금액: {int(stats["total_diff_value"]):,}원'

    # 헤더
    headers = ['품번', '품명', '규격', '카테고리', '시스템재고', '실재고', '차이', '단가', '차이금액', '비고']
    col_widths = [14, 28, 18, 10, 10, 10, 10, 12, 14, 18]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        letter = chr(64 + col_idx) if col_idx <= 9 else chr(64 + col_idx - 9)  # Handle > I
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'J'].width = w

    for row_idx, ai in enumerate(audit_items, 6):
        item = items_map.get(ai.item_id)
        if not item:
            continue
        price = item.last_unit_price or 0
        diff = ai.diff_qty or 0
        diff_value = diff * price

        row_data = [
            item.icube_item_cd or '', item.item_name or '', item.item_spec or '',
            item.category or '', ai.system_qty or 0,
            ai.actual_qty if ai.actual_qty is not None else '-',
            diff, price, diff_value, ai.note or '',
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 7 and isinstance(val, (int, float)):  # 차이
                if val > 0:
                    cell.font = blue_font
                elif val < 0:
                    cell.font = red_font
            if col_idx in (5, 6, 7, 8, 9):
                cell.alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'실사보고서_{audit.audit_no}_{audit.audit_date}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ===================================================================
# 8. 재고 변동 이력
# ===================================================================
@inventory_bp.route('/inventory/movements')
@login_required
def movement_list():
    q = (request.args.get('q') or '').strip()
    m_type = (request.args.get('type') or '').strip()
    date_from = (request.args.get('date_from') or '').strip()
    date_to = (request.args.get('date_to') or '').strip()
    page = safe_int(request.args.get('page'), 1)
    per_page = 50

    with get_db() as db:
        query = db.query(StockMovement)

        if q:
            # 품목명으로 필터
            item_ids = [r[0] for r in db.query(Item.id).filter(
                or_(Item.item_name.ilike(f'%{q}%'), Item.icube_item_cd.ilike(f'%{q}%'))
            ).all()]
            if item_ids:
                query = query.filter(StockMovement.item_id.in_(item_ids))
            else:
                query = query.filter(StockMovement.id < 0)  # 결과 없음

        if m_type:
            query = query.filter(StockMovement.movement_type == m_type)

        if date_from:
            try:
                dt_from = datetime.datetime.strptime(date_from, '%Y-%m-%d')
                query = query.filter(StockMovement.created_at >= dt_from)
            except ValueError:
                pass
        if date_to:
            try:
                dt_to = datetime.datetime.strptime(date_to, '%Y-%m-%d') + datetime.timedelta(days=1)
                query = query.filter(StockMovement.created_at < dt_to)
            except ValueError:
                pass

        total = query.count()
        pagination = make_pagination(page, per_page, total)
        offset = (pagination['page'] - 1) * per_page
        movements = query.order_by(desc(StockMovement.created_at)).offset(offset).limit(per_page).all()

        # item 이름 미리 로딩
        item_ids = list(set(m.item_id for m in movements))
        items_map = {}
        if item_ids:
            for item in db.query(Item).filter(Item.id.in_(item_ids)).all():
                items_map[item.id] = item

        return render_template(
            'inventory_movements.html',
            movements=movements,
            items_map=items_map,
            pagination=pagination,
            movement_labels=MOVEMENT_TYPE_LABELS,
            filters={'q': q, 'type': m_type, 'date_from': date_from, 'date_to': date_to},
            fmt_money=_fmt_money,
        )


# ===================================================================
# 9. 재고회전율 분석
# ===================================================================
@inventory_bp.route('/inventory/turnover')
@login_required
def turnover_report():
    date_from = (request.args.get('date_from') or '').strip()
    date_to = (request.args.get('date_to') or '').strip()
    category = (request.args.get('category') or '').strip()

    if not date_from:
        date_from = (datetime.date.today() - datetime.timedelta(days=90)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = datetime.date.today().strftime('%Y-%m-%d')

    with get_db() as db:
        try:
            start_dt = datetime.datetime.strptime(date_from, '%Y-%m-%d')
            end_dt = datetime.datetime.strptime(date_to, '%Y-%m-%d') + datetime.timedelta(days=1)
        except ValueError:
            start_dt = datetime.datetime.now() - datetime.timedelta(days=90)
            end_dt = datetime.datetime.now()

        results = calc_turnover_rate(db, start_dt, end_dt, category or None)
        results.sort(key=lambda x: x['turnover'], reverse=True)

        categories = [r[0] for r in db.query(distinct(Item.category)).filter(
            Item.category.isnot(None), Item.category != ''
        ).order_by(Item.category).all()]

        return render_template(
            'inventory_turnover.html',
            results=results,
            categories=categories,
            filters={'date_from': date_from, 'date_to': date_to, 'category': category},
            fmt_money=_fmt_money,
        )


# ===================================================================
# 10. 수동 재고 조정
# ===================================================================
@inventory_bp.route('/inventory/item/<int:item_id>/adjust', methods=['POST'])
@login_required
def item_adjust(item_id):
    with get_db() as db:
        item = db.query(Item).get(item_id)
        if not item:
            abort(404)

        adjust_qty_raw = request.form.get('adjust_qty', '').strip()
        adjust_reason = (request.form.get('adjust_reason') or '').strip()

        if not adjust_qty_raw:
            flash('조정 수량을 입력해주세요.', 'warning')
            return redirect(url_for('inventory.inventory_items'))

        if not adjust_reason:
            flash('조정 사유를 입력해주세요.', 'warning')
            return redirect(url_for('inventory.inventory_items'))

        try:
            adjust_qty = float(adjust_qty_raw)
        except (ValueError, TypeError):
            flash('유효한 수량을 입력해주세요.', 'warning')
            return redirect(url_for('inventory.inventory_items'))

        movement_type = 'IN_ADJUST' if adjust_qty > 0 else 'OUT_ADJUST'
        user_name = session.get('full_name', '사용자')

        record_stock_movement(
            db, item_id,
            movement_type=movement_type,
            quantity=adjust_qty,
            note=f'수동조정: {adjust_reason}',
            created_by=user_name,
        )
        db.commit()
        flash(f'[{item.item_name}] 재고가 {adjust_qty:+g} 조정되었습니다.', 'success')

    return redirect(url_for('inventory.inventory_items'))


# ===================================================================
# 11. 안전재고 설정
# ===================================================================
@inventory_bp.route('/inventory/item/<int:item_id>/safety-stock', methods=['POST'])
@login_required
def item_safety_stock(item_id):
    with get_db() as db:
        item = db.query(Item).get(item_id)
        if not item:
            abort(404)

        safety_raw = request.form.get('safety_stock', '').strip()
        try:
            item.safety_stock = max(0, float(safety_raw))
        except (ValueError, TypeError):
            item.safety_stock = 0

        db.commit()
        flash(f'[{item.item_name}] 안전재고가 {item.safety_stock}으로 설정되었습니다.', 'success')

    return redirect(url_for('inventory.inventory_items'))


# ===================================================================
# 12. 엑셀 다운로드
# ===================================================================
@inventory_bp.route('/inventory/export')
@login_required
def inventory_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with get_db() as db:
        items = db.query(Item).filter(Item.is_active == True).order_by(Item.category, Item.item_name).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '재고현황'

        headers = ['품번', '품명', '규격', '분류', '총재고', '예약', '가용재고', '안전재고', '단가', '재고금액', '상태']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for r, item in enumerate(items, 2):
            stock = item.stock_qty or 0
            reserved = item.reserved_qty or 0
            available = max(0, stock - reserved)
            safety = item.safety_stock or 0
            price = item.last_unit_price or 0
            value = stock * price

            status = '정상'
            if safety > 0 and stock < safety:
                status = '부족'

            row_data = [
                item.icube_item_cd or '',
                item.item_name or '',
                item.item_spec or '',
                item.category or '',
                stock, reserved, available, safety,
                price, value, status,
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = thin_border
                cell.font = Font(size=9)
                if col >= 5:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')

        widths = [12, 30, 20, 10, 10, 10, 10, 10, 12, 15, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        ws.auto_filter.ref = f'A1:K{len(items) + 1}'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'재고현황_{datetime.date.today().strftime("%Y%m%d")}.xlsx',
    )


# ===================================================================
# API: 대시보드 통계 JSON
# ===================================================================
@inventory_bp.route('/api/inventory/summary')
@login_required
def api_inventory_summary():
    with get_db() as db:
        items = db.query(Item).filter(Item.is_active == True).all()
        total_value = sum((i.stock_qty or 0) * (i.last_unit_price or 0) for i in items)
        low_count = sum(1 for i in items if (i.safety_stock or 0) > 0 and (i.stock_qty or 0) < (i.safety_stock or 0))

        return jsonify({
            'total_items': len(items),
            'total_value': total_value,
            'low_stock_count': low_count,
        })


# ===================================================================
# API: 저재고 품목 JSON
# ===================================================================
@inventory_bp.route('/api/inventory/low-stock')
@login_required
def api_low_stock():
    with get_db() as db:
        items = db.query(Item).filter(
            Item.is_active == True,
            Item.safety_stock > 0,
        ).all()

        low = []
        for item in items:
            if (item.stock_qty or 0) < (item.safety_stock or 0):
                low.append({
                    'id': item.id,
                    'name': item.item_name,
                    'code': item.icube_item_cd or '',
                    'stock_qty': item.stock_qty or 0,
                    'safety_stock': item.safety_stock or 0,
                    'shortage': (item.safety_stock or 0) - (item.stock_qty or 0),
                })

        low.sort(key=lambda x: x['shortage'], reverse=True)
        return jsonify(low[:20])
