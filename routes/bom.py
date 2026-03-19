"""
BOM 관리 Blueprint (Phase 4).

- BOM 목록/생성/수정/삭제
- 프로젝트별 소요 자재 자동 계산
"""

import datetime
import io
import json
import logging
import os
import tempfile
from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, jsonify, abort, session, send_file,
)
from sqlalchemy import desc, func
from sqlalchemy.orm import joinedload
from modules.auth_decorators import login_required
from modules.pagination import make_pagination
from modules.utils import safe_int
from modules.db_context import get_db
from modules.models import (
    BomHeader, BomItem, Contract, ContractItem, Project,
    PurchaseOrder, PurchaseOrderItem, MaterialOrder,
    Receiving, ReceivingItem, ReceivingHistory, Item,
    Vendor,
)

logger = logging.getLogger(__name__)


def _parse_option_filter_text(v):
    """옵션조건 텍스트를 JSON 문자열로 변환.

    "렌즈=20도" → '{"렌즈":"20도"}'
    "렌즈=20도,반사판=A" → '{"렌즈":"20도","반사판":"A"}'
    빈칸 → None
    """
    if not v or not str(v).strip():
        return None
    result = {}
    for pair in str(v).strip().split(','):
        pair = pair.strip()
        if '=' not in pair:
            continue
        key, val = pair.split('=', 1)
        key, val = key.strip(), val.strip()
        if key and val:
            result[key] = val
    return json.dumps(result, ensure_ascii=False) if result else None

bom_bp = Blueprint('bom', __name__)


def _get_latest_receiving_prices(db, bom_items):
    """BOM 품목들의 최신 입고 단가를 조회.

    1순위: 자체 입고(receivings + receiving_items) - item_name 매칭
    2순위: iCUBE 입고이력(receiving_history) - items_json 내 item_cd 매칭

    Returns: {item_code: {'unit_price': float, 'date': str, 'source': str}}
    """
    result = {}
    item_codes = {bi.item_code for bi in bom_items if bi.item_code}
    item_names = {bi.item_name for bi in bom_items if bi.item_name}
    if not item_codes and not item_names:
        return result

    # 1) 자체 입고에서 최신 단가 (item_name 매칭)
    for name in item_names:
        row = (
            db.query(ReceivingItem.unit_price, Receiving.rcv_date)
            .join(Receiving, ReceivingItem.receiving_id == Receiving.id)
            .filter(
                ReceivingItem.item_name == name,
                ReceivingItem.unit_price > 0,
            )
            .order_by(desc(Receiving.rcv_date))
            .first()
        )
        if row:
            # item_name → item_code 매핑 찾기
            matched_bi = next((bi for bi in bom_items if bi.item_name == name), None)
            key = matched_bi.item_code if matched_bi and matched_bi.item_code else name
            result[key] = {
                'unit_price': row.unit_price,
                'date': row.rcv_date.strftime('%Y-%m-%d') if row.rcv_date else '',
                'source': '자체입고',
            }

    # 2) iCUBE 입고이력에서 최신 단가 (item_cd 매칭) — 자체 입고에 없는 것만
    remaining_codes = item_codes - set(result.keys())
    if remaining_codes:
        histories = (
            db.query(ReceivingHistory)
            .filter(ReceivingHistory.items_json.isnot(None))
            .order_by(desc(ReceivingHistory.receive_date))
            .limit(500)
            .all()
        )
        for hist in histories:
            if not remaining_codes:
                break
            try:
                items_list = json.loads(hist.items_json) if hist.items_json else []
            except (json.JSONDecodeError, TypeError):
                continue
            for it in items_list:
                code = (it.get('item_cd') or '').strip()
                price = it.get('unit_price') or 0
                if code in remaining_codes and price > 0:
                    result[code] = {
                        'unit_price': price,
                        'date': hist.receive_date.strftime('%Y-%m-%d') if hist.receive_date else '',
                        'source': 'iCUBE',
                    }
                    remaining_codes.discard(code)

    return result


# ===================================================================
# 1. BOM 목록
# ===================================================================
@bom_bp.route('/bom')
@login_required
def bom_list():
    q = (request.args.get('q') or '').strip()
    category = (request.args.get('category') or '').strip()
    page = safe_int(request.args.get('page'), 1)
    per_page = 20

    with get_db() as db:
        query = db.query(BomHeader)

        if q:
            like_q = f"%{q}%"
            query = query.filter(
                (BomHeader.product_name.ilike(like_q)) |
                (BomHeader.product_code.ilike(like_q))
            )

        if category:
            query = query.filter(BomHeader.product_category == category)

        total = query.count()
        pagination = make_pagination(page, per_page, total)
        offset = (pagination['page'] - 1) * per_page

        boms = query.order_by(BomHeader.product_category, BomHeader.product_code) \
                    .offset(offset).limit(per_page).all()

        # 각 BOM의 부품 수 + 총 금액
        for bom in boms:
            bom._item_count = db.query(func.count(BomItem.id)).filter(
                BomItem.bom_id == bom.id
            ).scalar() or 0
            bom._total_amount = db.query(func.sum(BomItem.amount)).filter(
                BomItem.bom_id == bom.id
            ).scalar() or 0

        # 제품군별 카운트
        cat_counts = db.query(
            BomHeader.product_category, func.count(BomHeader.id)
        ).filter(
            BomHeader.product_category.isnot(None)
        ).group_by(BomHeader.product_category).order_by(
            func.count(BomHeader.id).desc()
        ).all()

        stats = {
            'total': db.query(func.count(BomHeader.id)).scalar() or 0,
            'active': db.query(func.count(BomHeader.id)).filter(BomHeader.is_active == True).scalar() or 0,
            'categories': [(c, n) for c, n in cat_counts],
        }

        return render_template(
            'bom_list.html',
            boms=boms,
            pagination=pagination,
            stats=stats,
            filters={'q': q, 'category': category},
        )


# ===================================================================
# 2. BOM 생성
# ===================================================================
@bom_bp.route('/bom/create', methods=['GET', 'POST'])
@login_required
def bom_create():
    with get_db() as db:
        if request.method == 'POST':
            product_code = (request.form.get('product_code') or '').strip()
            product_name = (request.form.get('product_name') or '').strip()
            version = (request.form.get('version') or '1.0').strip()
            note = (request.form.get('note') or '').strip()

            if not product_code or not product_name:
                flash('제품코드와 제품명을 입력해주세요.', 'warning')
                return redirect(url_for('bom.bom_create'))

            # 중복 체크
            existing = db.query(BomHeader).filter_by(product_code=product_code).first()
            if existing:
                flash(f'제품코드 "{product_code}"는 이미 존재합니다.', 'warning')
                return redirect(url_for('bom.bom_create'))

            bom = BomHeader(
                product_code=product_code,
                product_name=product_name,
                version=version,
                is_active=True,
                note=note,
            )
            db.add(bom)
            db.flush()

            # BOM 품목 추가
            item_names = request.form.getlist('item_name[]')
            item_specs = request.form.getlist('item_spec[]')
            item_qtys = request.form.getlist('quantity[]')
            item_units = request.form.getlist('unit[]')
            item_notes = request.form.getlist('item_note[]')

            for i in range(len(item_names)):
                name = (item_names[i] if i < len(item_names) else '').strip()
                if not name:
                    continue

                spec = (item_specs[i] if i < len(item_specs) else '').strip()
                qty = float(item_qtys[i]) if i < len(item_qtys) and item_qtys[i] else 1
                unit = (item_units[i] if i < len(item_units) else '').strip()
                item_note = (item_notes[i] if i < len(item_notes) else '').strip()

                db.add(BomItem(
                    bom_id=bom.id,
                    item_name=name,
                    item_spec=spec,
                    quantity=qty,
                    unit=unit,
                    note=item_note,
                ))

            db.commit()
            flash(f'BOM "{product_name}"이 생성되었습니다.', 'success')
            return redirect(url_for('bom.bom_detail', bom_id=bom.id))

        return render_template('bom_create.html')


# ===================================================================
# 3. BOM 상세
# ===================================================================
@bom_bp.route('/bom/<int:bom_id>')
@login_required
def bom_detail(bom_id):
    with get_db() as db:
        bom = db.query(BomHeader).options(
            joinedload(BomHeader.bom_items),
        ).get(bom_id)
        if not bom:
            abort(404)

        # 최신 입고 단가 조회 (품목코드 기준)
        latest_prices = _get_latest_receiving_prices(db, bom.bom_items)

        # 슈퍼BOM: option_schema 파싱 + 각 아이템에 display 텍스트 추가
        option_schema = None
        if bom.option_schema:
            try:
                option_schema = json.loads(bom.option_schema)
            except (json.JSONDecodeError, TypeError):
                pass

        for bi in bom.bom_items:
            bi._option_display = ''
            if bi.option_filter:
                try:
                    pf = json.loads(bi.option_filter)
                    bi._option_display = ', '.join(f'{k}={v}' for k, v in pf.items())
                except (json.JSONDecodeError, TypeError):
                    pass

        return render_template(
            'bom_detail.html',
            bom=bom,
            latest_prices=latest_prices,
            option_schema=option_schema,
        )


# ===================================================================
# 4. BOM 수정
# ===================================================================
@bom_bp.route('/bom/<int:bom_id>/edit', methods=['POST'])
@login_required
def bom_edit(bom_id):
    with get_db() as db:
        bom = db.query(BomHeader).get(bom_id)
        if not bom:
            abort(404)

        bom.product_name = (request.form.get('product_name') or bom.product_name).strip()
        bom.version = (request.form.get('version') or bom.version).strip()
        bom.note = (request.form.get('note') or '').strip()
        bom.is_active = request.form.get('is_active') == 'on'

        # 기존 품목 업데이트 방식 (item_code, bom_item_id 연결 보존)
        existing_items = db.query(BomItem).filter_by(bom_id=bom.id).order_by(BomItem.id).all()

        item_names = request.form.getlist('item_name[]')
        item_specs = request.form.getlist('item_spec[]')
        item_qtys = request.form.getlist('quantity[]')
        item_units = request.form.getlist('unit[]')
        item_notes = request.form.getlist('item_note[]')
        unit_prices = request.form.getlist('unit_price[]')
        option_filters = request.form.getlist('option_filter[]')

        # 기존 아이템 업데이트 + 신규 추가
        for i in range(len(item_names)):
            name = (item_names[i] if i < len(item_names) else '').strip()
            if not name:
                continue

            spec = (item_specs[i] if i < len(item_specs) else '').strip()
            qty = float(item_qtys[i]) if i < len(item_qtys) and item_qtys[i] else 1
            unit = (item_units[i] if i < len(item_units) else '').strip()
            item_note = (item_notes[i] if i < len(item_notes) else '').strip()

            up_str = (unit_prices[i] if i < len(unit_prices) else '').strip()
            up = float(up_str) if up_str else None
            amount = (up * qty) if up else None

            # 옵션조건
            of_raw = (option_filters[i] if i < len(option_filters) else '').strip()
            of_json = _parse_option_filter_text(of_raw) if of_raw else None

            if i < len(existing_items):
                # 기존 아이템 업데이트
                bi = existing_items[i]
                bi.item_name = name
                bi.item_spec = spec
                bi.quantity = qty
                bi.unit = unit
                bi.note = item_note
                bi.option_filter = of_json
                # 단가 변경 시 직전 단가 기록
                if up is not None and bi.unit_price is not None and up != bi.unit_price:
                    bi.prev_unit_price = bi.unit_price
                bi.unit_price = up
                bi.amount = amount
            else:
                # 신규 추가
                db.add(BomItem(
                    bom_id=bom.id,
                    item_name=name,
                    item_spec=spec,
                    quantity=qty,
                    unit=unit,
                    note=item_note,
                    unit_price=up,
                    amount=amount,
                    option_filter=of_json,
                ))

        # form에서 제거된 아이템 삭제
        if len(item_names) < len(existing_items):
            for j in range(len(item_names), len(existing_items)):
                db.delete(existing_items[j])

        # option_schema 재생성 (저장된 option_filter들에서)
        all_items = db.query(BomItem).filter_by(bom_id=bom.id).all()
        new_schema = {}
        for bi in all_items:
            if bi.option_filter:
                try:
                    pf = json.loads(bi.option_filter)
                    for k, v in pf.items():
                        if k not in new_schema:
                            new_schema[k] = []
                        if v not in new_schema[k]:
                            new_schema[k].append(v)
                except (json.JSONDecodeError, TypeError):
                    pass
        bom.option_schema = json.dumps(new_schema, ensure_ascii=False) if new_schema else None

        db.commit()
        flash('BOM이 수정되었습니다.', 'success')
        return redirect(url_for('bom.bom_detail', bom_id=bom_id))


# ===================================================================
# 5. BOM 삭제
# ===================================================================
@bom_bp.route('/bom/<int:bom_id>/delete', methods=['POST'])
@login_required
def bom_delete(bom_id):
    with get_db() as db:
        bom = db.query(BomHeader).get(bom_id)
        if not bom:
            abort(404)

        db.delete(bom)
        db.commit()
        flash('BOM이 삭제되었습니다.', 'success')
        return redirect(url_for('bom.bom_list'))


# ===================================================================
# 5-1. BOM 일괄 삭제
# ===================================================================
@bom_bp.route('/bom/bulk-delete', methods=['POST'])
@login_required
def bom_bulk_delete():
    ids = request.form.getlist('bom_ids')
    if not ids:
        flash('삭제할 BOM을 선택해주세요.', 'warning')
        return redirect(url_for('bom.bom_list'))

    with get_db() as db:
        deleted = 0
        for bom_id_str in ids:
            bom_id = safe_int(bom_id_str, 0)
            if not bom_id:
                continue
            bom = db.query(BomHeader).get(bom_id)
            if bom:
                db.delete(bom)
                deleted += 1
        db.commit()

    flash(f'BOM {deleted}건이 삭제되었습니다.', 'success')
    return redirect(url_for('bom.bom_list'))


# ===================================================================
# 5-1b. BOM 일괄 비활성화
# ===================================================================
@bom_bp.route('/bom/bulk-deactivate', methods=['POST'])
@login_required
def bom_bulk_deactivate():
    ids = request.form.getlist('bom_ids')
    if not ids:
        flash('비활성화할 BOM을 선택해주세요.', 'warning')
        return redirect(url_for('bom.bom_list'))

    with get_db() as db:
        count = 0
        for bom_id_str in ids:
            bom_id = safe_int(bom_id_str, 0)
            if not bom_id:
                continue
            bom = db.query(BomHeader).get(bom_id)
            if bom and bom.is_active:
                bom.is_active = False
                count += 1
        db.commit()

    flash(f'BOM {count}건이 비활성화되었습니다.', 'info')
    return redirect(url_for('bom.bom_list'))


# ===================================================================
# 5-2. BOM 엑셀 다운로드
# ===================================================================
@bom_bp.route('/bom/export')
@login_required
def bom_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with get_db() as db:
        boms = db.query(BomHeader).options(
            joinedload(BomHeader.bom_items),
        ).order_by(BomHeader.product_category, BomHeader.product_code).all()

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=9)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # 제품군별 시트
        categories = {}
        for bom in boms:
            cat = bom.product_category or '기타'
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(bom)

        for cat, cat_boms in categories.items():
            ws_name = cat[:31]  # 시트명 31자 제한
            ws = wb.create_sheet(title=ws_name)

            headers = ['No', '완제품코드', '제품명', '인증번호', '품목코드', '품목명',
                        '규격', '소요량', '단가', '금액', '납품업체', '비고', '옵션조건']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            row = 2
            for bom in cat_boms:
                for idx, bi in enumerate(bom.bom_items):
                    # 옵션조건 표시용 텍스트
                    of_text = ''
                    if bi.option_filter:
                        try:
                            pf = json.loads(bi.option_filter)
                            of_text = ', '.join(f'{k}={v}' for k, v in pf.items())
                        except (json.JSONDecodeError, TypeError):
                            pass

                    row_data = [
                        idx + 1 if idx == 0 else '',
                        bom.product_code if idx == 0 else '',
                        bom.product_name if idx == 0 else '',
                        bom.certification_no or '' if idx == 0 else '',
                        bi.item_code or '',
                        bi.item_name or '',
                        bi.item_spec or '',
                        bi.quantity or 0,
                        bi.unit_price or 0,
                        bi.amount or 0,
                        bi.supplier or '',
                        bi.note or '',
                        of_text,
                    ]
                    for col, val in enumerate(row_data, 1):
                        cell = ws.cell(row=row, column=col, value=val)
                        cell.border = thin_border
                        cell.font = Font(size=9)
                    row += 1

                # 제품 구분선
                if bom.bom_items:
                    row += 1

            widths = [6, 18, 20, 15, 15, 25, 20, 8, 10, 12, 15, 15, 18]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='BOM_목록.xlsx',
    )


# ===================================================================
# 5-3. BOM 엑셀 임포트
# ===================================================================
@bom_bp.route('/bom/import', methods=['GET', 'POST'])
@login_required
def bom_import():
    if request.method == 'POST':
        action = request.form.get('action', '')

        # --- 1단계: 엑셀 업로드 → 비교 ---
        if action == 'upload':
            file = request.files.get('file')
            if not file or not file.filename.endswith(('.xlsx', '.xls')):
                flash('엑셀 파일(.xlsx)을 선택해주세요.', 'warning')
                return redirect(url_for('bom.bom_import'))

            import openpyxl
            try:
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            except Exception as e:
                flash(f'엑셀 파일 읽기 실패: {e}', 'danger')
                return redirect(url_for('bom.bom_import'))

            # 시트별 BOM 파싱
            parsed_boms = {}  # {product_code: {header_info, items: [...]}}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                # 헤더 행 찾기: "No" 또는 "완제품코드" 또는 "품목코드" 포함 행
                header_idx = None
                col_map = {}
                for ri, row in enumerate(rows):
                    cells = [str(c).strip() if c else '' for c in row]
                    for ci, cv in enumerate(cells):
                        lv = cv.lower()
                        if cv in ('No', 'NO', 'no', 'No.'):
                            col_map['no'] = ci
                        elif cv in ('완제품코드', '제품코드') or lv == 'product_code':
                            col_map['product_code'] = ci
                        elif cv in ('품목코드', '부품코드') or lv == 'item_code':
                            col_map['item_code'] = ci
                        elif cv in ('품목명', '부품명', '품명') or lv == 'item_name':
                            col_map['item_name'] = ci
                        elif cv in ('소요량', '수량') or lv == 'quantity':
                            col_map['quantity'] = ci
                        elif cv in ('단가',) or lv == 'unit_price':
                            col_map['unit_price'] = ci
                        elif cv in ('부품업체', '납품업체', '업체') or lv == 'supplier':
                            col_map['supplier'] = ci
                        elif cv in ('비고', '원수') or lv == 'note':
                            if 'note' not in col_map:
                                col_map['note'] = ci
                        elif cv in ('규격', '사양') or lv == 'spec':
                            col_map['spec'] = ci
                        elif cv in ('옵션조건', '옵션', 'option', 'option_filter') or lv == 'option_filter':
                            col_map['option_filter'] = ci
                    if 'item_name' in col_map or 'item_code' in col_map:
                        header_idx = ri
                        break

                if header_idx is None:
                    continue

                # 데이터 파싱
                for row in rows[header_idx + 1:]:
                    cells = [str(c).strip() if c else '' for c in row]

                    def get_col(key):
                        idx = col_map.get(key)
                        return cells[idx] if idx is not None and idx < len(cells) else ''

                    item_name = get_col('item_name')
                    if not item_name:
                        continue

                    product_code = get_col('product_code')
                    if not product_code:
                        continue

                    # 수량/단가 숫자 변환
                    qty_str = get_col('quantity')
                    try:
                        qty = float(qty_str) if qty_str and not qty_str.startswith('=') else 1
                    except (ValueError, TypeError):
                        qty = 1

                    price_str = get_col('unit_price')
                    try:
                        price = float(price_str) if price_str and not price_str.startswith('=') else 0
                    except (ValueError, TypeError):
                        price = 0

                    if product_code not in parsed_boms:
                        parsed_boms[product_code] = {
                            'product_code': product_code,
                            'sheet': sheet_name,
                            'items': [],
                        }

                    parsed_boms[product_code]['items'].append({
                        'item_code': get_col('item_code'),
                        'item_name': item_name,
                        'item_spec': get_col('spec'),
                        'quantity': qty,
                        'unit_price': price,
                        'supplier': get_col('supplier'),
                        'note': get_col('note'),
                        'option_filter': _parse_option_filter_text(get_col('option_filter')),
                    })

                wb.close()

            if not parsed_boms:
                flash('엑셀에서 BOM 데이터를 찾을 수 없습니다. 헤더(완제품코드/품목명 등)를 확인해주세요.', 'warning')
                return redirect(url_for('bom.bom_import'))

            # DB 기존 BOM과 비교
            with get_db() as db:
                existing_boms = db.query(BomHeader).options(
                    joinedload(BomHeader.bom_items),
                ).all()
                db_by_code = {b.product_code: b for b in existing_boms}

                results = []
                for pc, parsed in parsed_boms.items():
                    db_bom = db_by_code.get(pc)
                    if db_bom:
                        # 비교: 품목 수 차이, 변경된 품목 감지
                        db_items = {(bi.item_name, bi.item_code or ''): bi for bi in db_bom.bom_items}
                        excel_items = parsed['items']

                        item_changes = []
                        new_items = 0
                        changed_items = 0
                        for ei in excel_items:
                            key = (ei['item_name'], ei['item_code'])
                            if key in db_items:
                                dbi = db_items[key]
                                diffs = []
                                if ei['quantity'] != (dbi.quantity or 0):
                                    diffs.append(f"소요량: {dbi.quantity}→{ei['quantity']}")
                                if ei['unit_price'] and ei['unit_price'] != (dbi.unit_price or 0):
                                    diffs.append(f"단가: {dbi.unit_price or 0}→{ei['unit_price']}")
                                if ei['supplier'] and ei['supplier'] != (dbi.supplier or ''):
                                    diffs.append(f"업체: {dbi.supplier or ''}→{ei['supplier']}")
                                if diffs:
                                    changed_items += 1
                                    item_changes.append({'name': ei['item_name'], 'diffs': diffs})
                            else:
                                new_items += 1
                                item_changes.append({'name': ei['item_name'], 'diffs': ['신규 품목']})

                        status = 'changed' if (new_items > 0 or changed_items > 0) else 'same'
                        results.append({
                            'product_code': pc,
                            'sheet': parsed['sheet'],
                            'excel_item_count': len(excel_items),
                            'db_item_count': len(db_bom.bom_items),
                            'db_id': db_bom.id,
                            'db_category': db_bom.product_category or '',
                            'status': status,
                            'new_items': new_items,
                            'changed_items': changed_items,
                            'item_changes': item_changes[:5],  # 표시용 최대 5건
                        })
                    else:
                        results.append({
                            'product_code': pc,
                            'sheet': parsed['sheet'],
                            'excel_item_count': len(parsed['items']),
                            'db_item_count': 0,
                            'db_id': None,
                            'db_category': '',
                            'status': 'new',
                            'new_items': len(parsed['items']),
                            'changed_items': 0,
                            'item_changes': [],
                        })

            # 임시 파일에 저장
            import hashlib
            cache_data = {'parsed': parsed_boms, 'results': results}
            cache_id = hashlib.md5(json.dumps(cache_data, ensure_ascii=False, default=str).encode()).hexdigest()[:12]
            cache_dir = os.path.join(tempfile.gettempdir(), 'lightsync_bom_import')
            os.makedirs(cache_dir, exist_ok=True)
            cache_path = os.path.join(cache_dir, f'{cache_id}.json')
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, default=str)

            new_count = sum(1 for r in results if r['status'] == 'new')
            changed_count = sum(1 for r in results if r['status'] == 'changed')
            same_count = sum(1 for r in results if r['status'] == 'same')

            return render_template('bom_import.html',
                                   results=results,
                                   new_count=new_count,
                                   changed_count=changed_count,
                                   same_count=same_count,
                                   cache_id=cache_id,
                                   step='compare')

        # --- 2단계: 선택 항목 적용 ---
        elif action == 'apply':
            import os
            cache_id = request.form.get('cache_id', '')
            cache_path = os.path.join(tempfile.gettempdir(), 'lightsync_bom_import', f'{cache_id}.json')
            if not cache_id or not os.path.exists(cache_path):
                flash('비교 데이터가 만료되었습니다. 다시 업로드해주세요.', 'warning')
                return redirect(url_for('bom.bom_import'))

            with open(cache_path, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
            os.remove(cache_path)

            parsed_boms = cache_data['parsed']
            results = cache_data['results']
            selected = set(request.form.getlist('selected'))

            with get_db() as db:
                added_boms = 0
                updated_boms = 0

                for i, r in enumerate(results):
                    if str(i) not in selected:
                        continue

                    pc = r['product_code']
                    parsed = parsed_boms.get(pc)
                    if not parsed:
                        continue

                    if r['status'] == 'new':
                        # option_schema 자동 생성
                        option_schema = {}
                        for ei in parsed['items']:
                            of = ei.get('option_filter')
                            if of:
                                try:
                                    pf = json.loads(of)
                                    for k, v in pf.items():
                                        if k not in option_schema:
                                            option_schema[k] = []
                                        if v not in option_schema[k]:
                                            option_schema[k].append(v)
                                except (json.JSONDecodeError, TypeError):
                                    pass

                        # 신규 BOM 생성
                        bom = BomHeader(
                            product_code=pc,
                            product_name=pc,
                            product_category=parsed['sheet'],
                            is_active=True,
                            option_schema=json.dumps(option_schema, ensure_ascii=False) if option_schema else None,
                        )
                        db.add(bom)
                        db.flush()

                        for ei in parsed['items']:
                            db.add(BomItem(
                                bom_id=bom.id,
                                item_code=ei['item_code'] or None,
                                item_name=ei['item_name'],
                                item_spec=ei.get('item_spec') or None,
                                quantity=ei['quantity'],
                                unit_price=ei['unit_price'] or None,
                                amount=(ei['quantity'] * ei['unit_price']) if ei['unit_price'] else None,
                                supplier=ei['supplier'] or None,
                                note=ei.get('note') or None,
                                option_filter=ei.get('option_filter'),
                            ))
                        added_boms += 1

                    elif r['status'] == 'changed' and r['db_id']:
                        # 기존 BOM 갱신: 기존 품목 업데이트 + 신규 추가
                        bom = db.query(BomHeader).get(r['db_id'])
                        if not bom:
                            continue

                        existing_items = {(bi.item_name, bi.item_code or ''): bi for bi in bom.bom_items}

                        for ei in parsed['items']:
                            key = (ei['item_name'], ei['item_code'])
                            if key in existing_items:
                                bi = existing_items[key]
                                if ei['quantity']:
                                    bi.quantity = ei['quantity']
                                if ei['unit_price']:
                                    if bi.unit_price and ei['unit_price'] != bi.unit_price:
                                        bi.prev_unit_price = bi.unit_price
                                    bi.unit_price = ei['unit_price']
                                    bi.amount = ei['quantity'] * ei['unit_price']
                                if ei['supplier']:
                                    bi.supplier = ei['supplier']
                            else:
                                db.add(BomItem(
                                    bom_id=bom.id,
                                    item_code=ei['item_code'] or None,
                                    item_name=ei['item_name'],
                                    item_spec=ei.get('item_spec') or None,
                                    quantity=ei['quantity'],
                                    unit_price=ei['unit_price'] or None,
                                    amount=(ei['quantity'] * ei['unit_price']) if ei['unit_price'] else None,
                                    supplier=ei['supplier'] or None,
                                    note=ei.get('note') or None,
                                    option_filter=ei.get('option_filter'),
                                ))
                        updated_boms += 1

                db.commit()

            msg_parts = []
            if added_boms:
                msg_parts.append(f'신규 {added_boms}건 생성')
            if updated_boms:
                msg_parts.append(f'{updated_boms}건 갱신')
            flash(', '.join(msg_parts) + ' 완료!', 'success')
            return redirect(url_for('bom.bom_list'))

    # GET
    return render_template('bom_import.html', step='upload')


# ===================================================================
# 6. 프로젝트별 소요 자재 자동 계산
# ===================================================================
@bom_bp.route('/bom/material-requirement')
@login_required
def material_requirement():
    """프로젝트별 소요 자재 계산 페이지"""
    contract_id = safe_int(request.args.get('contract_id'), 0)

    with get_db() as db:
        contracts = db.query(Contract).join(Project).filter(
            Project.status != '완료'
        ).order_by(desc(Contract.id)).limit(50).all()

        requirement = []
        selected_contract = None

        if contract_id:
            selected_contract = db.query(Contract).options(
                joinedload(Contract.items),
                joinedload(Contract.project),
            ).get(contract_id)

            if selected_contract:
                # 계약 품목 x BOM = 소요 자재
                for ci in selected_contract.items:
                    # BOM 매칭: 품목 카테고리/모델명으로 검색
                    bom = None
                    if ci.model_name:
                        bom = db.query(BomHeader).filter(
                            BomHeader.is_active == True,
                            (BomHeader.product_name.ilike(f'%{ci.model_name}%')) |
                            (BomHeader.product_code.ilike(f'%{ci.model_name}%'))
                        ).first()
                    if not bom and ci.category:
                        bom = db.query(BomHeader).filter(
                            BomHeader.is_active == True,
                            (BomHeader.product_name.ilike(f'%{ci.category}%')) |
                            (BomHeader.product_code.ilike(f'%{ci.category}%'))
                        ).first()

                    if bom:
                        # 슈퍼BOM: 협의내용(item_spec_json)에서 선택된 옵션으로 필터링
                        selected_options = {}
                        if ci.item_spec_json:
                            try:
                                selected_options = json.loads(ci.item_spec_json) if isinstance(ci.item_spec_json, str) else ci.item_spec_json
                            except (json.JSONDecodeError, TypeError):
                                pass

                        for bi in bom.bom_items:
                            # 옵션 필터 적용: 공통부품(null)은 항상 포함, 옵션부품은 매칭 시만
                            if bi.option_filter and selected_options:
                                try:
                                    item_opts = json.loads(bi.option_filter)
                                    skip = False
                                    for k, v in item_opts.items():
                                        if k in selected_options and selected_options[k] != v:
                                            skip = True
                                            break
                                    if skip:
                                        continue
                                except (json.JSONDecodeError, TypeError):
                                    pass

                            needed = (bi.quantity or 0) * (ci.quantity or 0)

                            # 1차: bom_item_id 기반 PurchaseOrderItem 발주량 (정확)
                            ordered_via_po = db.query(
                                func.coalesce(func.sum(PurchaseOrderItem.quantity), 0)
                            ).join(
                                PurchaseOrder, PurchaseOrderItem.po_id == PurchaseOrder.id
                            ).filter(
                                PurchaseOrderItem.bom_item_id == bi.id,
                                PurchaseOrder.contract_id == contract_id,
                                PurchaseOrder.status != '취소',
                            ).scalar() or 0

                            # 2차: MaterialOrder 기반 (하위 호환, bom_item_id 없는 기존 데이터)
                            ordered_via_mo = db.query(
                                func.coalesce(func.sum(MaterialOrder.quantity), 0)
                            ).filter(
                                MaterialOrder.contract_item_id == ci.id,
                                MaterialOrder.order_status.in_(['발주완료', '입고완료']),
                            ).scalar() or 0

                            # 둘 중 큰 값 사용 (중복 방지)
                            ordered = max(float(ordered_via_po), float(ordered_via_mo))

                            # 입고량: bom_item_id 기반 우선, fallback MaterialOrder
                            received_via_mo = db.query(
                                func.coalesce(func.sum(MaterialOrder.quantity), 0)
                            ).filter(
                                MaterialOrder.contract_item_id == ci.id,
                                MaterialOrder.order_status == '입고완료',
                            ).scalar() or 0

                            shortage = max(0, needed - ordered)

                            requirement.append({
                                'contract_item': ci,
                                'bom_item': bi,
                                'product_qty': ci.quantity or 0,
                                'per_unit': bi.quantity or 0,
                                'needed': needed,
                                'ordered': ordered,
                                'received': float(received_via_mo),
                                'shortage': shortage,
                            })
                    else:
                        # BOM 매칭 없음
                        requirement.append({
                            'contract_item': ci,
                            'bom_item': None,
                            'product_qty': ci.quantity or 0,
                            'per_unit': 0,
                            'needed': 0,
                            'ordered': 0,
                            'received': 0,
                            'shortage': 0,
                            'no_bom': True,
                        })

        return render_template(
            'bom_requirement.html',
            contracts=contracts,
            selected_contract=selected_contract,
            requirement=requirement,
            contract_id=contract_id,
        )


# ===================================================================
# 6-1. 소요자재 부족분 -> 발주서 자동 생성
# ===================================================================
@bom_bp.route('/bom/create-po-from-requirement', methods=['POST'])
@login_required
def create_po_from_requirement():
    """소요자재 부족분에서 거래처별 발주서 자동 생성"""
    from routes.purchase_order import _generate_po_no

    contract_id = safe_int(request.form.get('contract_id'), 0)
    selected_items_json = request.form.get('selected_items', '[]')

    if not contract_id:
        flash('계약 정보가 없습니다.', 'warning')
        return redirect(url_for('bom.material_requirement'))

    try:
        selected_items = json.loads(selected_items_json)
    except (json.JSONDecodeError, TypeError):
        flash('선택 항목 데이터가 올바르지 않습니다.', 'danger')
        return redirect(url_for('bom.material_requirement', contract_id=contract_id))

    if not selected_items:
        flash('발주할 자재를 선택해주세요.', 'warning')
        return redirect(url_for('bom.material_requirement', contract_id=contract_id))

    with get_db() as db:
        contract = db.query(Contract).get(contract_id)
        if not contract:
            flash('계약을 찾을 수 없습니다.', 'danger')
            return redirect(url_for('bom.material_requirement'))

        # supplier 기준 그룹핑
        from collections import defaultdict
        groups = defaultdict(list)
        for item in selected_items:
            supplier = (item.get('supplier') or '미지정').strip()
            groups[supplier].append(item)

        created_pos = []
        for supplier_name, items in groups.items():
            # Vendor 매칭 (ilike)
            vendor = db.query(Vendor).filter(
                Vendor.name.ilike(f'%{supplier_name}%'),
                Vendor.is_active == True,
            ).first()

            # 없으면 자동 생성
            if not vendor:
                vendor = Vendor(name=supplier_name, is_active=True)
                db.add(vendor)
                db.flush()

            # PurchaseOrder 생성
            po = PurchaseOrder(
                po_no=_generate_po_no(db),
                po_date=datetime.date.today(),
                vendor_id=vendor.id,
                contract_id=contract_id,
                project_id=contract.project_id,
                assigned_to=session.get('user_id'),
                status='작성중',
                note=f'BOM 소요자재 자동생성',
                created_by=session.get('user_id'),
            )
            db.add(po)
            db.flush()

            total_amount = 0
            for item in items:
                qty = float(item.get('quantity') or 0)
                price = float(item.get('unit_price') or 0)
                amount = qty * price

                po_item = PurchaseOrderItem(
                    po_id=po.id,
                    item_name=item.get('item_name', ''),
                    item_spec=item.get('item_spec', ''),
                    quantity=qty,
                    unit_price=price,
                    amount=amount,
                    unit=item.get('unit', ''),
                    bom_item_id=safe_int(item.get('bom_item_id'), 0) or None,
                )
                db.add(po_item)
                total_amount += amount

            po.total_amount = total_amount
            po.tax_amount = round(total_amount * 0.1)
            created_pos.append(po)

        db.commit()

        if len(created_pos) == 1:
            flash(f'발주서 {created_pos[0].po_no}가 생성되었습니다.', 'success')
            return redirect(url_for('purchase_order.po_detail', po_id=created_pos[0].id))
        else:
            po_nos = ', '.join(p.po_no for p in created_pos)
            flash(f'발주서 {len(created_pos)}건이 생성되었습니다: {po_nos}', 'success')
            return redirect(url_for('purchase_order.po_list'))


# ===================================================================
# 7. BOM 검색 API (AJAX)
# ===================================================================
@bom_bp.route('/api/bom/search')
@login_required
def api_bom_search():
    q = (request.args.get('q') or '').strip()
    with get_db() as db:
        query = db.query(BomHeader).filter(BomHeader.is_active == True)
        if q:
            like_q = f"%{q}%"
            query = query.filter(
                (BomHeader.product_name.ilike(like_q)) |
                (BomHeader.product_code.ilike(like_q))
            )
        boms = query.order_by(BomHeader.product_name).limit(20).all()
        return jsonify([{
            'id': b.id,
            'product_code': b.product_code,
            'product_name': b.product_name,
            'version': b.version,
        } for b in boms])
