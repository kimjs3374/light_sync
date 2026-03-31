import datetime
import io
import os
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_file
from sqlalchemy import and_
from modules.auth_decorators import login_required, admin_required, menu_required
from modules.db_context import get_db
from modules.utils import safe_int, parse_date
from modules.pagination import make_pagination
from modules.models import (
    Certification, CERT_TYPE_CHOICES, Notification, User,
)

cert_bp = Blueprint("certification", __name__)

UPLOAD_SUBDIR = 'certifications'


def _save_cert_file(file_obj):
    """인증서 파일 저장 → 상대경로 반환"""
    if not file_obj or not file_obj.filename:
        return None
    upload_dir = os.path.join(current_app.static_folder, 'uploads', UPLOAD_SUBDIR)
    os.makedirs(upload_dir, exist_ok=True)
    ts = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    safe_name = file_obj.filename.replace(' ', '_')
    fname = f"{ts}_{safe_name}"
    file_obj.save(os.path.join(upload_dir, fname))
    return f"uploads/{UPLOAD_SUBDIR}/{fname}"


# -------------------------------------------------------------------
# 1. 인증서 목록 + 만료 알림 대시보드
# -------------------------------------------------------------------
@cert_bp.route('/certifications')
@menu_required('certification')
def cert_list():
    page = safe_int(request.args.get('page'), 1)
    per_page = 50
    cert_type_filter = request.args.get('cert_type', '').strip()
    status_filter = request.args.get('status', '').strip()
    search_q = request.args.get('q', '').strip()

    today = datetime.date.today()

    with get_db() as db:
        query = db.query(Certification).filter(Certification.is_active.is_(True))

        if cert_type_filter:
            query = query.filter(Certification.cert_type == cert_type_filter)

        if search_q:
            like = f'%{search_q}%'
            query = query.filter(
                Certification.cert_name.ilike(like)
                | Certification.cert_no.ilike(like)
                | Certification.product_model.ilike(like)
                | Certification.issued_by.ilike(like)
            )

        if status_filter:
            if status_filter == 'expired':
                query = query.filter(Certification.expiry_date < today)
            elif status_filter == 'critical':
                query = query.filter(
                    Certification.expiry_date >= today,
                    Certification.expiry_date <= today + datetime.timedelta(days=7),
                )
            elif status_filter == 'warning':
                query = query.filter(
                    Certification.expiry_date > today + datetime.timedelta(days=7),
                    Certification.expiry_date <= today + datetime.timedelta(days=30),
                )
            elif status_filter == 'ok':
                query = query.filter(Certification.expiry_date > today + datetime.timedelta(days=30))

        query = query.order_by(Certification.expiry_date.asc().nullslast())

        total = query.count()
        pagination = make_pagination(page, per_page, total)
        certs = query.offset((pagination['page'] - 1) * per_page).limit(per_page).all()

        # 만료 통계
        all_active = db.query(Certification).filter(Certification.is_active.is_(True)).all()
        stats = {
            'total': len(all_active),
            'expired': sum(1 for c in all_active if c.expiry_date and c.expiry_date < today),
            'critical': sum(1 for c in all_active if c.expiry_date and 0 <= (c.expiry_date - today).days <= 7),
            'warning': sum(1 for c in all_active if c.expiry_date and 7 < (c.expiry_date - today).days <= 30),
            'ok': sum(1 for c in all_active if c.expiry_date and (c.expiry_date - today).days > 30),
        }

    filters = {'cert_type': cert_type_filter, 'status': status_filter, 'q': search_q}
    return render_template(
        'certification_list.html',
        certs=certs,
        stats=stats,
        pagination=pagination,
        cert_types=CERT_TYPE_CHOICES,
        filters=filters,
        today=today,
    )


# -------------------------------------------------------------------
# 2. 인증서 등록/수정
# -------------------------------------------------------------------
@cert_bp.route('/certifications/create', methods=['GET', 'POST'])
@menu_required('certification', write=True)
def cert_create():
    with get_db() as db:
        if request.method == 'POST':
            file_path = _save_cert_file(request.files.get('cert_file'))
            cert = Certification(
                cert_type=request.form.get('cert_type', '기타'),
                cert_name=request.form.get('cert_name', '').strip(),
                cert_no=request.form.get('cert_no', '').strip() or None,
                issued_by=request.form.get('issued_by', '').strip() or None,
                issued_date=parse_date(request.form.get('issued_date')),
                expiry_date=parse_date(request.form.get('expiry_date')),
                product_model=request.form.get('product_model', '').strip() or None,
                alert_days=safe_int(request.form.get('alert_days'), 30),
                file_path=file_path,
                note=request.form.get('note', '').strip() or None,
            )
            db.add(cert)
            db.commit()
            flash(f'인증서 "{cert.cert_name}" 등록 완료', 'success')
            return redirect(url_for('certification.cert_list'))

    return render_template('certification_form.html', cert=None, cert_types=CERT_TYPE_CHOICES)


@cert_bp.route('/certifications/<int:cert_id>/edit', methods=['GET', 'POST'])
@menu_required('certification', write=True)
def cert_edit(cert_id):
    with get_db() as db:
        cert = db.query(Certification).get(cert_id)
        if not cert:
            flash('인증서를 찾을 수 없습니다.', 'danger')
            return redirect(url_for('certification.cert_list'))

        if request.method == 'POST':
            cert.cert_type = request.form.get('cert_type', cert.cert_type)
            cert.cert_name = request.form.get('cert_name', '').strip()
            cert.cert_no = request.form.get('cert_no', '').strip() or None
            cert.issued_by = request.form.get('issued_by', '').strip() or None
            cert.issued_date = parse_date(request.form.get('issued_date'))
            cert.expiry_date = parse_date(request.form.get('expiry_date'))
            cert.product_model = request.form.get('product_model', '').strip() or None
            cert.alert_days = safe_int(request.form.get('alert_days'), 30)
            cert.note = request.form.get('note', '').strip() or None

            new_file = request.files.get('cert_file')
            if new_file and new_file.filename:
                cert.file_path = _save_cert_file(new_file)

            db.commit()
            flash('인증서 수정 완료', 'success')
            return redirect(url_for('certification.cert_list'))

    return render_template('certification_form.html', cert=cert, cert_types=CERT_TYPE_CHOICES)


@cert_bp.route('/certifications/<int:cert_id>/delete', methods=['POST'])
@menu_required('certification', write=True)
def cert_delete(cert_id):
    with get_db() as db:
        cert = db.query(Certification).get(cert_id)
        if cert:
            cert.is_active = False
            db.commit()
            flash('인증서 비활성화 완료', 'success')
    return redirect(url_for('certification.cert_list'))


# -------------------------------------------------------------------
# 3. 엑셀 템플릿 다운로드
# -------------------------------------------------------------------
@cert_bp.route('/certifications/template')
@menu_required('certification')
def cert_template():
    """인증서 일괄등록용 엑셀 템플릿 다운로드"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── 시트1: 인증서 (회사 인증) ──
    ws1 = wb.active
    ws1.title = '인증서'
    headers1 = ['인증유형', '인증서명', '인증번호', '발급기관', '발급일', '만료일', '대상제품/모델', '알림일수', '비고']
    type_list = ', '.join(code for code, _ in CERT_TYPE_CHOICES)

    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(bottom=thin)

    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    # 샘플 행
    sample1 = ['KS인증', 'KS C 7658(가로등 및 보안등기구)', '', '한국표준협회', '2024-03-27', '2026-08-04', '', '30', '']
    for col_idx, v in enumerate(sample1, 1):
        cell = ws1.cell(row=2, column=col_idx, value=v)
        cell.font = Font(color='999999', size=10)
        cell.border = border

    # 안내행
    ws1.cell(row=4, column=1, value=f'※ 인증유형: {type_list}').font = Font(color='FF0000', size=9)
    ws1.cell(row=5, column=1, value='※ 날짜 형식: YYYY-MM-DD (예: 2026-08-04)').font = Font(color='FF0000', size=9)
    ws1.cell(row=6, column=1, value='※ 알림일수: 만료 며칠 전부터 알림 (기본 30)').font = Font(color='FF0000', size=9)

    widths1 = [14, 35, 15, 15, 12, 12, 20, 10, 20]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = w

    # ── 시트2: 고효율인증 ──
    ws2 = wb.create_sheet('고효율인증')
    headers2 = ['인증번호', '만료일', '형식', '모델명', '비고']

    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    sample2 = ['37685', '2026-08-19', 'LED투광등기구', 'MT-FL-400S', '']
    for col_idx, v in enumerate(sample2, 1):
        cell = ws2.cell(row=2, column=col_idx, value=v)
        cell.font = Font(color='999999', size=10)
        cell.border = border

    ws2.cell(row=4, column=1, value='※ 이 시트의 데이터는 모두 "고효율인증" 유형으로 등록됩니다').font = Font(color='FF0000', size=9)

    widths2 = [15, 14, 25, 20, 20]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='인증서_임포트_템플릿.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# -------------------------------------------------------------------
# 4. 엑셀 임포트
# -------------------------------------------------------------------
def _parse_date_cell(val):
    """엑셀 셀 값 → date 객체"""
    if not val:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


@cert_bp.route('/certifications/import', methods=['POST'])
@menu_required('certification', write=True)
def cert_import():
    """엑셀 파일에서 인증서 일괄 임포트"""
    from openpyxl import load_workbook

    file = request.files.get('import_file')
    if not file or not file.filename:
        flash('파일을 선택해주세요.', 'warning')
        return redirect(url_for('certification.cert_list'))

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('xlsx 파일만 지원합니다.', 'danger')
        return redirect(url_for('certification.cert_list'))

    valid_types = {code for code, _ in CERT_TYPE_CHOICES}
    created = 0
    skipped = 0
    errors = []

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        flash(f'엑셀 파일 읽기 실패: {e}', 'danger')
        return redirect(url_for('certification.cert_list'))

    with get_db() as db:
        # ── 시트1: 인증서 (일반 인증) ──
        if '인증서' in wb.sheetnames:
            ws = wb['인증서']
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row or len(row) < 2:
                    continue
                cert_type = str(row[0] or '').strip()
                cert_name = str(row[1] or '').strip()
                if not cert_name:
                    continue
                if cert_type and cert_type not in valid_types:
                    errors.append(f'시트[인증서] {row_idx}행: 유형 "{cert_type}" 없음')
                    continue

                db.add(Certification(
                    cert_type=cert_type or '기타',
                    cert_name=cert_name,
                    cert_no=str(row[2] or '').strip() or None if len(row) > 2 else None,
                    issued_by=str(row[3] or '').strip() or None if len(row) > 3 else None,
                    issued_date=_parse_date_cell(row[4]) if len(row) > 4 else None,
                    expiry_date=_parse_date_cell(row[5]) if len(row) > 5 else None,
                    product_model=str(row[6] or '').strip() or None if len(row) > 6 else None,
                    alert_days=safe_int(row[7], 30) if len(row) > 7 else 30,
                    note=str(row[8] or '').strip() or None if len(row) > 8 else None,
                ))
                created += 1

        # ── 시트2: 고효율인증 ──
        he_sheet = None
        for name in wb.sheetnames:
            if '고효율' in name:
                he_sheet = wb[name]
                break

        if he_sheet:
            # 원본 엑셀(B열 시작) vs 템플릿(A열 시작) 자동 감지
            first_row = list(he_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            offset = 1 if (first_row[0] is None and len(first_row) > 1 and first_row[1]) else 0

            for row_idx, row in enumerate(he_sheet.iter_rows(min_row=2, values_only=True), 2):
                if not row or len(row) < (2 + offset):
                    continue
                cert_no = str(row[0 + offset] or '').strip()
                if not cert_no or cert_no == '인증번호':
                    continue

                expiry = _parse_date_cell(row[1 + offset]) if len(row) > (1 + offset) else None
                form_type = str(row[2 + offset] or '').strip() if len(row) > (2 + offset) else ''
                model = str(row[3 + offset] or '').strip() if len(row) > (3 + offset) else ''
                note = str(row[4 + offset] or '').strip() if len(row) > (4 + offset) else ''

                cert_name = f'고효율인증 {model}' if model else f'고효율인증 {cert_no}'
                db.add(Certification(
                    cert_type='고효율인증',
                    cert_name=cert_name,
                    cert_no=cert_no or None,
                    expiry_date=expiry,
                    product_model=model or None,
                    alert_days=60,
                    note=f'{form_type}. {note}'.strip('. ') if form_type or note else None,
                ))
                created += 1

        # ── 매그나텍 원본 시트 호환 (인증관련 시트) ──
        orig_sheet = None
        for name in wb.sheetnames:
            if '인증관련' in name:
                orig_sheet = wb[name]
                break

        if orig_sheet:
            TYPE_MAP = {
                'KS C': 'KS인증', '우수제품': '조달우수제품', '혁신제품': '혁신제품',
                '녹색기술': '녹색기술인증', 'ISO': 'ISO', 'MAS': 'MAS계약',
                '조달우수': '조달우수제품', '중소기업': '중소기업확인서',
                '직접생산': '직접생산증명', 'G-PASS': 'G-PASS', '광기술원': '기타',
                '단체표준': '단체표준',
            }
            for row_idx, row in enumerate(orig_sheet.iter_rows(min_row=2, values_only=False), 2):
                vals = {c.column_letter: c.value for c in row if c.value is not None}
                cert_name = str(vals.get('B', '')).strip()
                if not cert_name or cert_name == '인증명':
                    continue
                # 비고/메모성 행 스킵
                if any(kw in cert_name for kw in ('일정확인', '기간연장신청', '특이사항', '이부분')):
                    continue

                # 유형 자동 매핑
                cert_type = '기타'
                for prefix, mapped in TYPE_MAP.items():
                    if prefix in cert_name:
                        cert_type = mapped
                        break

                issued = _parse_date_cell(vals.get('C'))
                expiry = _parse_date_cell(vals.get('D'))
                note_val = str(vals.get('F', '')).strip() or None

                db.add(Certification(
                    cert_type=cert_type,
                    cert_name=cert_name,
                    issued_date=issued,
                    expiry_date=expiry,
                    alert_days=30,
                    note=note_val,
                ))
                created += 1

        db.commit()

    msg = f'임포트 완료: {created}건 등록'
    if skipped:
        msg += f', {skipped}건 건너뜀'
    if errors:
        msg += f', {len(errors)}건 오류'
        for e in errors[:5]:
            flash(e, 'warning')
    flash(msg, 'success')
    return redirect(url_for('certification.cert_list'))


# -------------------------------------------------------------------
# 5. 인증서 만료 알림 체크 (대시보드에서 호출)
# -------------------------------------------------------------------
def check_cert_expiry_alerts(db):
    """만료 임박 인증서 알림 생성 (대시보드 로드 시 호출)
    Returns: list of dicts for dashboard display
    """
    today = datetime.date.today()
    alerts = []
    certs = db.query(Certification).filter(
        Certification.is_active.is_(True),
        Certification.expiry_date.isnot(None),
    ).all()

    for c in certs:
        days = (c.expiry_date - today).days
        if days < 0:
            alerts.append({
                'level': 'danger',
                'message': f'[인증서 만료] {c.cert_name} ({c.cert_no or "-"}) — {abs(days)}일 경과',
                'cert_id': c.id,
            })
        elif days <= c.alert_days:
            level = 'danger' if days <= 7 else 'warning'
            alerts.append({
                'level': level,
                'message': f'[인증서 만료 {days}일 전] {c.cert_name} ({c.cert_no or "-"})',
                'cert_id': c.id,
            })
    return alerts
