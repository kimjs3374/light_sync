"""
routes/lighting_layout.py — 조명배치도 (타워별 투광등 넘버링 + 렌즈각도)
Blueprint: lighting_layout_bp  prefix: /lighting-layout
"""
import io
import json
from flask import (Blueprint, render_template, request, redirect,
                   url_for, session, flash, jsonify, send_file)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from modules.auth_decorators import login_required, menu_required
from modules.db_context import get_db
from modules.models.entities import Project, TowerLayout, TowerLayoutPosition, LensAngleConfig
from modules.history_board import append_history_log, get_user_display_name

lighting_layout_bp = Blueprint('lighting_layout', __name__, url_prefix='/lighting-layout')


# ──────────────────────────────────────────────────────────────
# 현장별 조명배치도 목록
# ──────────────────────────────────────────────────────────────
@lighting_layout_bp.route('/')
@login_required
def layout_list():
    with get_db() as db:
        layouts = (
            db.query(TowerLayout)
            .join(Project, TowerLayout.project_id == Project.id)
            .order_by(TowerLayout.created_at.desc())
            .all()
        )
        # 프로젝트별 그룹핑
        project_map = {}
        for lay in layouts:
            pid = lay.project_id
            if pid not in project_map:
                project_map[pid] = {
                    'project': lay.project,
                    'towers': []
                }
            project_map[pid]['towers'].append(lay)

        return render_template('lighting_layout/layout_list.html',
                               project_map=project_map)


# ──────────────────────────────────────────────────────────────
# 타워 추가 폼 (GET) + 생성 (POST)
# ──────────────────────────────────────────────────────────────
@lighting_layout_bp.route('/add', methods=['GET', 'POST'])
@login_required
@menu_required('lighting_layout')
def add_tower():
    if request.method == 'GET':
        project_id = request.args.get('project_id', type=int)
        with get_db() as db:
            # LED투광등기구가 있는 프로젝트만 조회
            projects = (
                db.query(Project)
                .filter(Project.project_no.notlike('G-%'))
                .order_by(Project.created_at.desc())
                .all()
            )
            return render_template('lighting_layout/add_tower.html',
                                   projects=projects,
                                   preselect_id=project_id)

    # POST: 타워 생성
    with get_db() as db:
        project_id = request.form.get('project_id', type=int)
        tower_name = request.form.get('tower_name', '').strip()
        rows = request.form.get('rows', 2, type=int)
        cols = request.form.get('cols', 3, type=int)
        model_name = request.form.get('model_name', '').strip() or None
        watt = request.form.get('watt', type=int)
        note = request.form.get('note', '').strip() or None

        if not project_id or not tower_name or rows < 1 or cols < 1:
            flash('필수 입력값을 확인해주세요.', 'danger')
            return redirect(url_for('lighting_layout.add_tower', project_id=project_id))

        tower = TowerLayout(
            project_id=project_id,
            tower_name=tower_name,
            rows=rows,
            cols=cols,
            model_name=model_name,
            watt=watt,
            note=note,
            created_by=session.get('full_name', '사용자'),
        )
        db.add(tower)
        db.flush()

        # 위치 자동 생성 (좌상단→우측 순번)
        no = 1
        for r in range(rows):
            for c in range(cols):
                db.add(TowerLayoutPosition(
                    tower_layout_id=tower.id,
                    position_no=no,
                    row_idx=r,
                    col_idx=c,
                ))
                no += 1

        append_history_log(db, project_id=project_id,
                           user_name=get_user_display_name(),
                           content=f"조명배치도 타워 추가: {tower_name} ({rows}×{cols}={rows*cols}등)",
                           scope='design')
        db.commit()
        flash(f'타워 [{tower_name}] 배치도가 생성되었습니다.', 'success')
        return redirect(url_for('lighting_layout.tower_detail', tower_id=tower.id))


# ──────────────────────────────────────────────────────────────
# 타워 상세 (그리드 시각화 + 렌즈각도 편집)
# ──────────────────────────────────────────────────────────────
@lighting_layout_bp.route('/<int:tower_id>')
@login_required
@menu_required('lighting_layout')
def tower_detail(tower_id):
    with get_db() as db:
        tower = db.query(TowerLayout).get(tower_id)
        if not tower:
            flash('타워 배치도를 찾을 수 없습니다.', 'danger')
            return redirect(url_for('lighting_layout.layout_list'))

        # 같은 프로젝트의 다른 타워들
        siblings = (
            db.query(TowerLayout)
            .filter(TowerLayout.project_id == tower.project_id,
                    TowerLayout.id != tower.id)
            .order_by(TowerLayout.tower_name)
            .all()
        )

        # 렌즈각도 옵션 조회 (lens_angle_configs 테이블 우선)
        lens_options = []
        matched_model = False

        if tower.model_name:
            # 1순위: lens_angle_configs에서 모델 키워드 매칭
            configs = db.query(LensAngleConfig).all()
            for cfg in configs:
                if cfg.model_name.upper() in tower.model_name.upper():
                    lens_options = cfg.angle_list
                    matched_model = True
                    break

        # 2순위: 전체 config 합산
        if not lens_options:
            all_configs = db.query(LensAngleConfig).all()
            if all_configs:
                seen = []
                for cfg in all_configs:
                    for a in cfg.angle_list:
                        if a not in seen:
                            seen.append(a)
                lens_options = seen

        # 3순위: 기본값
        if not lens_options:
            lens_options = ['10°', '15°', '20°', '25°', '30°', '40°', '60°', '90°', '120°']

        return render_template('lighting_layout/tower_detail.html',
                               tower=tower,
                               siblings=siblings,
                               lens_options=lens_options,
                               matched_model=matched_model)


# ──────────────────────────────────────────────────────────────
# 렌즈각도 일괄 저장 (AJAX)
# ──────────────────────────────────────────────────────────────
@lighting_layout_bp.route('/<int:tower_id>/save', methods=['POST'])
@login_required
@menu_required('lighting_layout')
def save_positions(tower_id):
    with get_db() as db:
        tower = db.query(TowerLayout).get(tower_id)
        if not tower:
            return jsonify({'ok': False, 'msg': '타워를 찾을 수 없습니다.'}), 404

        data = request.get_json(silent=True)
        if not data or 'positions' not in data:
            return jsonify({'ok': False, 'msg': '데이터가 없습니다.'}), 400

        changed = []
        for pos_data in data['positions']:
            pos = db.query(TowerLayoutPosition).get(pos_data.get('id'))
            if pos and pos.tower_layout_id == tower_id:
                old_angle = pos.lens_angle
                new_angle = (pos_data.get('lens_angle') or '').strip() or None
                new_note = (pos_data.get('note') or '').strip() or None
                if old_angle != new_angle:
                    changed.append(f"#{pos.position_no}: {old_angle or '-'}→{new_angle or '-'}")
                pos.lens_angle = new_angle
                pos.note = new_note

        if changed:
            append_history_log(db, project_id=tower.project_id,
                               user_name=get_user_display_name(),
                               content=f"조명배치도 [{tower.tower_name}] 렌즈각도 변경: {', '.join(changed[:5])}{'...' if len(changed) > 5 else ''}",
                               scope='design')
        db.commit()
        return jsonify({'ok': True, 'msg': f'{len(data["positions"])}개 위치 저장 완료'})


# ──────────────────────────────────────────────────────────────
# 타워 정보 수정 (POST)
# ──────────────────────────────────────────────────────────────
@lighting_layout_bp.route('/<int:tower_id>/edit', methods=['POST'])
@login_required
@menu_required('lighting_layout')
def edit_tower(tower_id):
    with get_db() as db:
        tower = db.query(TowerLayout).get(tower_id)
        if not tower:
            flash('타워를 찾을 수 없습니다.', 'danger')
            return redirect(url_for('lighting_layout.layout_list'))

        tower.tower_name = request.form.get('tower_name', tower.tower_name).strip()
        tower.model_name = request.form.get('model_name', '').strip() or None
        tower.watt = request.form.get('watt', type=int)
        tower.note = request.form.get('note', '').strip() or None

        new_rows = request.form.get('rows', tower.rows, type=int)
        new_cols = request.form.get('cols', tower.cols, type=int)

        # 격자 크기 변경 시 위치 재생성
        if new_rows != tower.rows or new_cols != tower.cols:
            # 기존 각도 데이터 보존 (위치 번호 기준)
            old_angles = {p.position_no: (p.lens_angle, p.note) for p in tower.positions}
            # 기존 삭제
            for p in tower.positions:
                db.delete(p)
            db.flush()
            tower.rows = new_rows
            tower.cols = new_cols
            # 새로 생성
            no = 1
            for r in range(new_rows):
                for c in range(new_cols):
                    old = old_angles.get(no, (None, None))
                    db.add(TowerLayoutPosition(
                        tower_layout_id=tower.id,
                        position_no=no,
                        row_idx=r,
                        col_idx=c,
                        lens_angle=old[0],
                        note=old[1],
                    ))
                    no += 1

        append_history_log(db, project_id=tower.project_id,
                           user_name=get_user_display_name(),
                           content=f"조명배치도 [{tower.tower_name}] 정보 수정",
                           scope='design')
        db.commit()
        flash('타워 정보가 수정되었습니다.', 'success')
        return redirect(url_for('lighting_layout.tower_detail', tower_id=tower.id))


# ──────────────────────────────────────────────────────────────
# 타워 삭제
# ──────────────────────────────────────────────────────────────
@lighting_layout_bp.route('/<int:tower_id>/delete', methods=['POST'])
@login_required
@menu_required('lighting_layout')
def delete_tower(tower_id):
    with get_db() as db:
        tower = db.query(TowerLayout).get(tower_id)
        if not tower:
            flash('타워를 찾을 수 없습니다.', 'danger')
            return redirect(url_for('lighting_layout.layout_list'))

        project_id = tower.project_id
        name = tower.tower_name
        db.delete(tower)
        append_history_log(db, project_id=project_id,
                           user_name=get_user_display_name(),
                           content=f"조명배치도 타워 삭제: {name}",
                           scope='design')
        db.commit()
        flash(f'타워 [{name}] 배치도가 삭제되었습니다.', 'success')
        return redirect(url_for('lighting_layout.layout_list'))


# ──────────────────────────────────────────────────────────────
# 현장 전체 삭제 (해당 프로젝트의 모든 타워 삭제)
# ──────────────────────────────────────────────────────────────
@lighting_layout_bp.route('/site/<int:project_id>/delete', methods=['POST'])
@login_required
@menu_required('lighting_layout')
def delete_site(project_id):
    with get_db() as db:
        towers = db.query(TowerLayout).filter_by(project_id=project_id).all()
        if not towers:
            flash('삭제할 타워가 없습니다.', 'danger')
            return redirect(url_for('lighting_layout.layout_list'))

        count = len(towers)
        for t in towers:
            db.delete(t)
        append_history_log(db, project_id=project_id,
                           user_name=get_user_display_name(),
                           content=f"조명배치도 현장 전체 삭제: 타워 {count}개",
                           scope='design')
        db.commit()
        flash(f'타워 {count}개가 삭제되었습니다.', 'success')
        return redirect(url_for('lighting_layout.layout_list'))


# ──────────────────────────────────────────────────────────────
# 렌즈관리 (모델별 각도 설정)
# ──────────────────────────────────────────────────────────────
@lighting_layout_bp.route('/lens-config')
@login_required
@menu_required('lighting_layout')
def lens_config():
    with get_db() as db:
        configs = db.query(LensAngleConfig).order_by(LensAngleConfig.model_name).all()
        return render_template('lighting_layout/lens_config.html', configs=configs)


@lighting_layout_bp.route('/lens-config/save', methods=['POST'])
@login_required
@menu_required('lighting_layout')
def lens_config_save():
    with get_db() as db:
        data = request.get_json(silent=True)
        if not data or 'items' not in data:
            return jsonify({'ok': False, 'msg': '데이터가 없습니다.'}), 400

        # 기존 전체 삭제 후 재생성
        db.query(LensAngleConfig).delete()
        for item in data['items']:
            model = (item.get('model_name') or '').strip()
            angles = (item.get('angles') or '').strip()
            if model and angles:
                db.add(LensAngleConfig(model_name=model, angles=angles))

        db.commit()
        return jsonify({'ok': True, 'msg': f'{len(data["items"])}개 모델 저장 완료'})


# ──────────────────────────────────────────────────────────────
# 엑셀 템플릿 다운로드
# ──────────────────────────────────────────────────────────────
@lighting_layout_bp.route('/excel-template')
@login_required
def excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '조명배치도'

    # 스타일
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F172A')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')

    # 헤더
    headers = ['현장번호', '현장명', '타워이름', '모델명', '행', '열', '번호', '렌즈각도']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # 텍스트 서식 (날짜 자동변환 방지)
    text_cols = [1, 4, 8]  # 현장번호, 모델명, 렌즈각도
    for col_idx in text_cols:
        for row_idx in range(1, 100):
            ws.cell(row=row_idx, column=col_idx).number_format = '@'

    # 예시 데이터
    examples = [
        ['2026-013', '울산공항 계류장', 'T1', 'STA-600', 2, 5, 1, '20'],
        ['2026-013', '울산공항 계류장', 'T1', 'STA-600', 2, 5, 2, '35'],
        ['2026-013', '울산공항 계류장', 'T1', 'STA-600', 2, 5, 3, '20'],
        ['2026-013', '울산공항 계류장', 'T2', 'STA-600', 2, 3, 1, '55'],
        ['2026-013', '울산공항 계류장', 'T2', 'STA-600', 2, 3, 2, '35'],
    ]
    example_font = Font(color='94A3B8', italic=True)
    for r, row_data in enumerate(examples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = example_font
            cell.border = thin_border
            cell.alignment = center

    # 안내 시트
    ws2 = wb.create_sheet('안내')
    guide = [
        ['컬럼', '설명', '필수'],
        ['현장번호', '설계번호 (예: 2026-013). project_no와 매칭', '필수'],
        ['현장명', '참고용 (매칭에 사용안함)', ''],
        ['타워이름', '타워 구분 (예: T1, T2)', '필수'],
        ['모델명', '투광등 모델명 (렌즈관리 매칭용)', ''],
        ['행', '타워 행 수', '필수'],
        ['열', '타워 열 수', '필수'],
        ['번호', '좌상단→우측 순번 (1~행×열)', '필수'],
        ['렌즈각도', '렌즈각도 값 (예: 20, 35, 광각)', ''],
    ]
    for r, row_data in enumerate(guide, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)

    # 컬럼 너비
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 6
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    response = send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fname = '조명배치도_템플릿.xlsx'
    response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return response


# ──────────────────────────────────────────────────────────────
# 엑셀 임포트
# ──────────────────────────────────────────────────────────────
@lighting_layout_bp.route('/import', methods=['POST'])
@login_required
@menu_required('lighting_layout')
def import_excel():
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('엑셀 파일(.xlsx)을 선택해주세요.', 'danger')
        return redirect(url_for('lighting_layout.layout_list'))

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        with get_db() as db:
            # 현장번호 → project_id 매핑 캐시
            project_cache = {}
            # 타워 캐시: (project_id, tower_name) → TowerLayout
            tower_cache = {}

            created_towers = 0
            updated_positions = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                project_no = str(row[0]).strip()
                tower_name = str(row[2] or '').strip()
                model_name = str(row[3] or '').strip() or None
                rows_val = int(row[4] or 2)
                cols_val = int(row[5] or 3)
                pos_no = int(row[6] or 0)
                lens_angle = str(row[7] or '').strip() or None

                if not project_no or not tower_name or pos_no < 1:
                    continue

                # 현장 찾기
                if project_no not in project_cache:
                    proj = db.query(Project).filter(Project.project_no == project_no).first()
                    project_cache[project_no] = proj.id if proj else None
                pid = project_cache[project_no]
                if not pid:
                    continue

                # 타워 찾거나 생성
                tkey = (pid, tower_name)
                if tkey not in tower_cache:
                    tower = db.query(TowerLayout).filter_by(
                        project_id=pid, tower_name=tower_name).first()
                    if not tower:
                        tower = TowerLayout(
                            project_id=pid, tower_name=tower_name,
                            rows=rows_val, cols=cols_val,
                            model_name=model_name,
                            created_by=session.get('full_name', '사용자'))
                        db.add(tower)
                        db.flush()
                        # 위치 자동 생성
                        no = 1
                        for r in range(rows_val):
                            for c in range(cols_val):
                                db.add(TowerLayoutPosition(
                                    tower_layout_id=tower.id,
                                    position_no=no, row_idx=r, col_idx=c))
                                no += 1
                        db.flush()
                        created_towers += 1
                    tower_cache[tkey] = tower

                tower = tower_cache[tkey]

                # 해당 번호 위치에 렌즈각도 입력
                if lens_angle:
                    pos = db.query(TowerLayoutPosition).filter_by(
                        tower_layout_id=tower.id, position_no=pos_no).first()
                    if pos:
                        pos.lens_angle = lens_angle
                        updated_positions += 1

            db.commit()
            flash(f'임포트 완료: 타워 {created_towers}개 생성, 렌즈각도 {updated_positions}건 입력', 'success')

    except Exception as e:
        flash(f'임포트 실패: {str(e)}', 'danger')

    return redirect(url_for('lighting_layout.layout_list'))
