"""
품목관리 Blueprint.

- 품목 목록/검색/필터 (카테고리별)
- 품목 상세/수정
- 품목 신규 등록
- 카테고리 목록 API
"""

import io
import json
import logging
import tempfile
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from sqlalchemy import or_, func, distinct, desc
from modules.auth_decorators import login_required
from modules.pagination import make_pagination
from modules.utils import safe_int
from modules.db_context import get_db
from modules.models import (
    Item, PurchaseOrderItem, PurchaseOrder, ReceivingItem, Receiving,
    ReceivingHistory, BomItem, BomHeader, Vendor,
)

logger = logging.getLogger(__name__)

item_bp = Blueprint('item', __name__)


# ===================================================================
# 1. 품목 목록
# ===================================================================
@item_bp.route('/item')
@login_required
def item_list():
    q = (request.args.get('q') or '').strip()
    category = (request.args.get('category') or '').strip()
    active_only = request.args.get('active_only', '1')
    page = safe_int(request.args.get('page'), 1)
    per_page = safe_int(request.args.get('per_page'), 50)

    with get_db() as db:
        query = db.query(Item)

        if active_only == '1':
            query = query.filter(Item.is_active == True)

        if q:
            like = f'%{q}%'
            # BomItem.supplier 또는 Vendor.name에서 거래처명 매칭되는 item_id 서브쿼리
            vendor_item_ids = (
                db.query(BomItem.item_id)
                .filter(BomItem.item_id.isnot(None), BomItem.supplier.ilike(like))
                .distinct()
                .subquery()
            )
            query = query.filter(or_(
                Item.icube_item_cd.ilike(like),
                Item.item_name.ilike(like),
                Item.item_spec.ilike(like),
                Item.manufacturer.ilike(like),
                Item.id.in_(vendor_item_ids),
            ))

        if category:
            query = query.filter(Item.category == category)

        total = query.count()
        items = query.order_by(Item.item_name).offset((page - 1) * per_page).limit(per_page).all()
        pagination = make_pagination(page, per_page, total)

        # 카테고리 목록 (필터용)
        categories = [r[0] for r in db.query(distinct(Item.category)).filter(
            Item.category.isnot(None), Item.category != ''
        ).order_by(Item.category).all()]

        # 거래건수: BomItem(item_id or item_code) + iCUBE 입고이력
        trade_counts = {}
        if items:
            item_ids = [i.id for i in items]
            item_codes = {i.id: i.icube_item_cd for i in items if i.icube_item_cd}

            # 1) BOM 사용 건수 (item_id 매칭)
            bom_counts = (
                db.query(BomItem.item_id, func.count(BomItem.id))
                .filter(BomItem.item_id.in_(item_ids))
                .group_by(BomItem.item_id)
                .all()
            )
            for item_id, cnt in bom_counts:
                trade_counts[item_id] = trade_counts.get(item_id, 0) + cnt

            # 2) BOM 사용 건수 (item_code 매칭 — item_id 없는 경우)
            code_to_id = {v: k for k, v in item_codes.items()}
            if code_to_id:
                bom_code_counts = (
                    db.query(BomItem.item_code, func.count(BomItem.id))
                    .filter(
                        BomItem.item_code.in_(list(code_to_id.keys())),
                        BomItem.item_id.is_(None),
                    )
                    .group_by(BomItem.item_code)
                    .all()
                )
                for code, cnt in bom_code_counts:
                    iid = code_to_id.get(code)
                    if iid:
                        trade_counts[iid] = trade_counts.get(iid, 0) + cnt

            # 3) iCUBE 입고이력 건수 (SQL LIKE로 item_cd 매칭)
            from sqlalchemy import text as sa_text
            for iid, code in item_codes.items():
                if iid in trade_counts:
                    continue  # BOM에서 이미 카운트됨
                rcv_cnt = db.execute(sa_text(
                    "SELECT COUNT(*) FROM receiving_history WHERE items_json LIKE :p"
                ), {'p': f'%"{code}"%'}).scalar() or 0
                if rcv_cnt > 0:
                    trade_counts[iid] = trade_counts.get(iid, 0) + rcv_cnt

    return render_template('item_list.html',
                           items=items,
                           pagination=pagination,
                           q=q,
                           category=category,
                           active_only=active_only,
                           categories=categories,
                           total=total,
                           trade_counts=trade_counts)


# ===================================================================
# 2. 품목 상세
# ===================================================================
@item_bp.route('/item/<int:item_id>')
@login_required
def item_detail(item_id):
    with get_db() as db:
        item = db.query(Item).get(item_id)
        if not item:
            flash('품목을 찾을 수 없습니다.', 'warning')
            return redirect(url_for('item.item_list'))

        categories = [r[0] for r in db.query(distinct(Item.category)).filter(
            Item.category.isnot(None), Item.category != ''
        ).order_by(Item.category).all()]

        # FR-05: 단가 추이 (icube_item_cd 없어도 item_name 기반으로 조회)
        price_history = []

        # 1) 자체 입고 단가 이력 - item_name 매칭 (부분 매칭 포함)
        name_filter = or_(
            ReceivingItem.item_name == item.item_name,
            ReceivingItem.item_name.ilike(f'%{item.item_name}%'),
        ) if item.item_name else (ReceivingItem.item_name == item.item_name)
        rcv_prices = (
            db.query(Receiving.rcv_date, ReceivingItem.unit_price, Receiving.vendor_id)
            .join(ReceivingItem, ReceivingItem.receiving_id == Receiving.id)
            .filter(
                name_filter,
                ReceivingItem.unit_price > 0,
            )
            .order_by(Receiving.rcv_date)
            .all()
        )
        for r in rcv_prices:
            vendor = db.query(Vendor).get(r.vendor_id) if r.vendor_id else None
            price_history.append({
                'date': r.rcv_date.strftime('%Y-%m-%d') if r.rcv_date else '',
                'price': r.unit_price,
                'vendor': vendor.name if vendor else '-',
                'source': '자체입고',
            })

        # 2) iCUBE 입고이력에서 단가 — PostgreSQL JSON 검색으로 직접 매칭
        if item.icube_item_cd:
            from sqlalchemy import text as sa_text
            icube_rows = db.execute(sa_text("""
                SELECT receive_date, vendor_name, items_json
                FROM receiving_history
                WHERE items_json LIKE :pattern
                ORDER BY receive_date
            """), {'pattern': f'%"item_cd": "{item.icube_item_cd}"%'}).fetchall()

            # fallback: 공백 없는 패턴도 시도
            if not icube_rows:
                icube_rows = db.execute(sa_text("""
                    SELECT receive_date, vendor_name, items_json
                    FROM receiving_history
                    WHERE items_json LIKE :pattern
                    ORDER BY receive_date
                """), {'pattern': f'%"{item.icube_item_cd}"%'}).fetchall()

            for row in icube_rows:
                try:
                    items_list = json.loads(row.items_json) if row.items_json else []
                except (json.JSONDecodeError, TypeError):
                    continue
                for it in items_list:
                    code = (it.get('item_cd') or '').strip()
                    if code == item.icube_item_cd and (it.get('unit_price') or 0) > 0:
                        price_history.append({
                            'date': row.receive_date.strftime('%Y-%m-%d') if row.receive_date else '',
                            'price': it['unit_price'],
                            'vendor': row.vendor_name or '-',
                            'source': 'iCUBE',
                        })

        price_history.sort(key=lambda x: x['date'], reverse=True)

        # FR-05: 관련 거래처 (이 품목을 납품한 업체들)
        related_vendors = []
        seen_vendors = set()

        # BomItem.supplier에서 - item_code 또는 item_name 매칭
        bom_supplier_filter = []
        if item.icube_item_cd:
            bom_supplier_filter.append(BomItem.item_code == item.icube_item_cd)
        if item.item_name:
            bom_supplier_filter.append(BomItem.item_name == item.item_name)

        if bom_supplier_filter:
            bom_suppliers = (
                db.query(BomItem.supplier)
                .filter(or_(*bom_supplier_filter), BomItem.supplier.isnot(None))
                .distinct()
                .all()
            )
            for (s,) in bom_suppliers:
                if s and s.strip() and s.strip() not in seen_vendors:
                    seen_vendors.add(s.strip())
                    related_vendors.append({'name': s.strip(), 'source': 'BOM'})

        # 단가이력에서 거래처 추출
        for ph in price_history:
            vn = ph.get('vendor', '')
            if vn and vn != '-' and vn not in seen_vendors:
                seen_vendors.add(vn)
                related_vendors.append({'name': vn, 'source': ph['source']})

        # FR-05: 사용되는 BOM 목록 - item_code 또는 item_name 매칭
        used_in_boms = []
        bom_filter = []
        if item.icube_item_cd:
            bom_filter.append(BomItem.item_code == item.icube_item_cd)
        if item.item_name:
            bom_filter.append(BomItem.item_name == item.item_name)

        if bom_filter:
            bom_usages = (
                db.query(BomItem, BomHeader)
                .join(BomHeader, BomItem.bom_id == BomHeader.id)
                .filter(or_(*bom_filter))
                .all()
            )
            seen_bom_ids = set()
            for bi, bh in bom_usages:
                if bh.id not in seen_bom_ids:
                    seen_bom_ids.add(bh.id)
                    used_in_boms.append({
                        'bom_id': bh.id,
                        'product_code': bh.product_code,
                        'product_category': bh.product_category or '',
                        'quantity': bi.quantity,
                    })

    return render_template('item_detail.html',
                           item=item,
                           categories=categories,
                           price_history=price_history,
                           related_vendors=related_vendors,
                           used_in_boms=used_in_boms)


# ===================================================================
# 3. 품목 수정
# ===================================================================
@item_bp.route('/item/<int:item_id>/edit', methods=['POST'])
@login_required
def item_edit(item_id):
    with get_db() as db:
        item = db.query(Item).get(item_id)
        if not item:
            flash('품목을 찾을 수 없습니다.', 'warning')
            return redirect(url_for('item.item_list'))

        item.icube_item_cd = (request.form.get('item_code') or '').strip() or None
        item.item_name = (request.form.get('item_name') or '').strip()
        item.item_spec = (request.form.get('item_spec') or '').strip() or None
        item.unit = (request.form.get('unit') or '').strip() or None
        item.category = (request.form.get('category') or '').strip() or None
        item.manufacturer = (request.form.get('manufacturer') or '').strip() or None
        item.note = (request.form.get('note') or '').strip() or None
        item.is_active = request.form.get('is_active') == '1'

        # 재고 수량 수동 수정 (FR-17)
        stock_qty_raw = request.form.get('stock_qty')
        if stock_qty_raw is not None and stock_qty_raw.strip() != '':
            try:
                new_stock = max(0, float(stock_qty_raw))
                item.stock_qty = new_stock
            except (ValueError, TypeError):
                pass

        db.commit()
        flash('품목이 수정되었습니다.', 'success')

    return redirect(url_for('item.item_detail', item_id=item_id))


# ===================================================================
# 4. 품목 신규 등록
# ===================================================================
@item_bp.route('/item/create', methods=['GET', 'POST'])
@login_required
def item_create():
    if request.method == 'POST':
        with get_db() as db:
            item_code = (request.form.get('item_code') or '').strip() or None
            item_name = (request.form.get('item_name') or '').strip()

            if not item_name:
                flash('품명은 필수입니다.', 'warning')
                return redirect(url_for('item.item_create'))

            # 품번 중복 체크
            if item_code:
                existing = db.query(Item).filter_by(icube_item_cd=item_code).first()
                if existing:
                    flash(f'이미 등록된 품번입니다: {item_code}', 'warning')
                    return redirect(url_for('item.item_create'))

            new_item = Item(
                icube_item_cd=item_code,
                item_name=item_name,
                item_spec=(request.form.get('item_spec') or '').strip() or None,
                unit=(request.form.get('unit') or '').strip() or None,
                category=(request.form.get('category') or '').strip() or None,
                manufacturer=(request.form.get('manufacturer') or '').strip() or None,
                note=(request.form.get('note') or '').strip() or None,
                is_active=True,
            )
            db.add(new_item)
            db.commit()
            flash('품목이 등록되었습니다.', 'success')
            return redirect(url_for('item.item_detail', item_id=new_item.id))

    # GET
    with get_db() as db:
        categories = [r[0] for r in db.query(distinct(Item.category)).filter(
            Item.category.isnot(None), Item.category != ''
        ).order_by(Item.category).all()]

    return render_template('item_create.html', categories=categories)


# ===================================================================
# 5. 품목 삭제 (비활성화)
# ===================================================================
@item_bp.route('/item/<int:item_id>/delete', methods=['POST'])
@login_required
def item_delete(item_id):
    with get_db() as db:
        item = db.query(Item).get(item_id)
        if not item:
            flash('품목을 찾을 수 없습니다.', 'warning')
            return redirect(url_for('item.item_list'))

        item.is_active = False
        db.commit()
        flash(f'품목 [{item.item_name}]이 비활성화되었습니다.', 'info')

    return redirect(url_for('item.item_list'))


# ===================================================================
# API: 카테고리 목록
# ===================================================================
@item_bp.route('/api/item/categories')
@login_required
def api_item_categories():
    with get_db() as db:
        categories = [r[0] for r in db.query(distinct(Item.category)).filter(
            Item.category.isnot(None), Item.category != ''
        ).order_by(Item.category).all()]
    return jsonify(categories)


# ===================================================================
# API: 품목 검색 (재고 매칭용)
# ===================================================================
@item_bp.route('/api/item/search')
@login_required
def api_item_search():
    q = (request.args.get('q') or '').strip()
    if not q or len(q) < 1:
        return jsonify([])
    like = f'%{q}%'
    with get_db() as db:
        items = db.query(Item).filter(
            or_(
                Item.icube_item_cd.ilike(like),
                Item.item_name.ilike(like),
                Item.item_spec.ilike(like),
            )
        ).order_by(Item.item_name).limit(15).all()
        return jsonify([{
            'id': i.id,
            'code': i.icube_item_cd or '',
            'name': i.item_name,
            'spec': i.item_spec or '',
            'stock': i.stock_qty or 0,
        } for i in items])


# ===================================================================
# 재고 매칭 도구
# ===================================================================
@item_bp.route('/item/stock-match', methods=['GET', 'POST'])
@login_required
def stock_match():
    import os
    json_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'reference', 'stock_data.json')

    if request.method == 'POST':
        # 매칭 저장
        import os as _os
        with get_db() as db:
            matched_count = 0
            match_keys = [k for k in request.form.keys() if k.startswith('match_') and request.form[k].strip()]

            # stock_data.json 로드 (매칭 기록 저장용)
            all_stocks = []
            if _os.path.exists(json_path):
                with open(json_path, 'r', encoding='utf-8') as f:
                    all_stocks = json.load(f)

            for key in match_keys:
                item_id_str = request.form[key].strip()
                idx_str = key.replace('match_', '')
                stock_str = request.form.get(f'stock_{idx_str}', '0')
                try:
                    item_id = int(item_id_str)
                    idx = int(idx_str)
                    stock_val = float(stock_str)
                except (ValueError, TypeError):
                    continue
                item = db.get(Item, item_id)
                if item:
                    item.stock_qty = stock_val
                    matched_count += 1
                    # JSON에 매칭 기록 저장
                    if idx < len(all_stocks):
                        all_stocks[idx]['matched_item_id'] = item_id

            db.commit()

            # JSON 업데이트 저장
            if matched_count and all_stocks:
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(all_stocks, f, ensure_ascii=False, indent=2)

            if matched_count:
                flash(f'{matched_count}건 재고가 반영되었습니다.', 'success')
            else:
                flash('매칭된 항목이 없습니다. 품목을 선택해주세요.', 'warning')
        return redirect(url_for('item.stock_match'))

    # GET: 미매칭 목록 로드
    unmatched = []
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            all_stocks = json.load(f)

        for i, s in enumerate(all_stocks):
            if s.get('matched_item_id'):
                continue
            if s['stock'] == 0:
                continue
            unmatched.append({
                'idx': i,
                'sheet': s['sheet'],
                'item_name': s['item_name'],
                'item_code': s.get('item_code', ''),
                'stock': s['stock'],
            })

    return render_template('item_stock_match.html', unmatched=unmatched)


# ===================================================================
# 엑셀 다운로드 (현재 품목 전체)
# ===================================================================
@item_bp.route('/item/export')
@login_required
def item_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with get_db() as db:
        items = db.query(Item).order_by(Item.icube_item_cd).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '품목목록'

        # 헤더
        headers = ['품번', '품명', '규격', '단위', '분류', '제조사', '실재고', '비고', '상태']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # 데이터
        for r, item in enumerate(items, 2):
            row_data = [
                item.icube_item_cd or '',
                item.item_name or '',
                item.item_spec or '',
                item.unit or '',
                item.category or '',
                item.manufacturer or '',
                item.stock_qty or 0,
                item.note or '',
                '사용' if item.is_active else '미사용',
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = thin_border
                cell.font = Font(size=9)

        # 컬럼 너비
        widths = [12, 30, 25, 8, 12, 15, 10, 20, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 필터 설정
        ws.auto_filter.ref = f'A1:I{len(items) + 1}'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='품목목록.xlsx',
    )


# ===================================================================
# 엑셀 업로드 → 비교/마이그레이션
# ===================================================================
@item_bp.route('/item/import', methods=['GET', 'POST'])
@login_required
def item_import():
    if request.method == 'POST':
        action = request.form.get('action', '')

        # --- 1단계: 엑셀 업로드 → 비교 결과 표시 ---
        if action == 'upload':
            file = request.files.get('file')
            if not file or not file.filename.endswith(('.xlsx', '.xls')):
                flash('엑셀 파일(.xlsx)을 선택해주세요.', 'warning')
                return redirect(url_for('item.item_import'))

            import openpyxl
            try:
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
            except Exception as e:
                flash(f'엑셀 파일 읽기 실패: {e}', 'danger')
                return redirect(url_for('item.item_import'))

            # 헤더 자동 감지 (품번/품명/규격 컬럼 위치)
            header_row = None
            col_map = {}
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                cells = [str(c).strip() if c else '' for c in row]
                for ci, cv in enumerate(cells):
                    if cv in ('품번', '품목코드', 'item_cd', 'item_code', 'code'):
                        col_map['code'] = ci
                    elif cv in ('품명', '품목명', 'item_name', 'name'):
                        col_map['name'] = ci
                    elif cv in ('규격', 'spec', 'item_spec', '사양'):
                        col_map['spec'] = ci
                if 'name' in col_map:
                    header_row = row_idx
                    break

            if header_row is None:
                flash('헤더를 찾을 수 없습니다. 품번/품명/규격 컬럼이 필요합니다.', 'warning')
                return redirect(url_for('item.item_import'))

            # 엑셀 데이터 파싱
            excel_items = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                cells = [str(c).strip() if c else '' for c in row]
                code = cells[col_map.get('code', 0)] if 'code' in col_map else ''
                name = cells[col_map.get('name', 1)] if 'name' in col_map else ''
                spec = cells[col_map.get('spec', 2)] if 'spec' in col_map else ''
                if not name:
                    continue
                excel_items.append({'code': code, 'name': name, 'spec': spec})

            wb.close()

            if not excel_items:
                flash('엑셀에 유효한 데이터가 없습니다.', 'warning')
                return redirect(url_for('item.item_import'))

            # DB 기존 품목과 비교
            with get_db() as db:
                existing = db.query(Item).all()
                db_by_code = {}
                db_by_name = {}
                for it in existing:
                    if it.icube_item_cd:
                        db_by_code[it.icube_item_cd] = it
                    db_by_name[it.item_name] = it

                results = []
                for ei in excel_items:
                    match = None
                    match_type = 'new'

                    # 1) 품번으로 매칭
                    if ei['code'] and ei['code'] in db_by_code:
                        match = db_by_code[ei['code']]
                    # 2) 품명으로 매칭
                    elif ei['name'] in db_by_name:
                        match = db_by_name[ei['name']]

                    if match:
                        # 변경 여부 체크
                        changes = []
                        if ei['code'] and ei['code'] != (match.icube_item_cd or ''):
                            changes.append(f"품번: {match.icube_item_cd or '없음'} → {ei['code']}")
                        if ei['name'] != (match.item_name or ''):
                            changes.append(f"품명: {match.item_name} → {ei['name']}")
                        if ei['spec'] != (match.item_spec or ''):
                            changes.append(f"규격: {match.item_spec or '없음'} → {ei['spec']}")

                        match_type = 'changed' if changes else 'same'
                        results.append({
                            'excel': ei,
                            'db_id': match.id,
                            'db_code': match.icube_item_cd or '',
                            'db_name': match.item_name,
                            'db_spec': match.item_spec or '',
                            'status': match_type,
                            'changes': changes,
                        })
                    else:
                        results.append({
                            'excel': ei,
                            'db_id': None,
                            'db_code': '',
                            'db_name': '',
                            'db_spec': '',
                            'status': 'new',
                            'changes': [],
                        })

            # 결과를 임시 파일에 저장 (세션 쿠키 크기 제한 회피)
            import os, hashlib
            cache_id = hashlib.md5(json.dumps(results, ensure_ascii=False).encode()).hexdigest()[:12]
            cache_dir = os.path.join(tempfile.gettempdir(), 'lightsync_import')
            os.makedirs(cache_dir, exist_ok=True)
            cache_path = os.path.join(cache_dir, f'{cache_id}.json')
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False)

            new_count = sum(1 for r in results if r['status'] == 'new')
            changed_count = sum(1 for r in results if r['status'] == 'changed')
            same_count = sum(1 for r in results if r['status'] == 'same')

            return render_template('item_import.html',
                                   results=results,
                                   new_count=new_count,
                                   changed_count=changed_count,
                                   same_count=same_count,
                                   cache_id=cache_id,
                                   step='compare')

        # --- 2단계: 선택 항목 적용 ---
        elif action == 'apply':
            import os
            cache_id = request.form.get('cache_id', '')
            cache_path = os.path.join(tempfile.gettempdir(), 'lightsync_import', f'{cache_id}.json')
            if not cache_id or not os.path.exists(cache_path):
                flash('비교 데이터가 만료되었습니다. 다시 업로드해주세요.', 'warning')
                return redirect(url_for('item.item_import'))
            with open(cache_path, 'r', encoding='utf-8') as f:
                results = json.load(f)
            os.remove(cache_path)
            if not results:
                flash('비교 데이터가 없습니다. 다시 업로드해주세요.', 'warning')
                return redirect(url_for('item.item_import'))

            selected = request.form.getlist('selected')
            selected_set = set(selected)

            with get_db() as db:
                added = 0
                updated = 0
                for i, r in enumerate(results):
                    if str(i) not in selected_set:
                        continue

                    if r['status'] == 'new':
                        new_item = Item(
                            icube_item_cd=r['excel']['code'] or None,
                            item_name=r['excel']['name'],
                            item_spec=r['excel']['spec'] or None,
                            is_active=True,
                        )
                        db.add(new_item)
                        added += 1

                    elif r['status'] == 'changed' and r['db_id']:
                        item = db.query(Item).get(r['db_id'])
                        if item:
                            if r['excel']['code']:
                                item.icube_item_cd = r['excel']['code']
                            item.item_name = r['excel']['name']
                            item.item_spec = r['excel']['spec'] or None
                            updated += 1

                db.commit()

            msg_parts = []
            if added:
                msg_parts.append(f'신규 {added}건 등록')
            if updated:
                msg_parts.append(f'{updated}건 수정')
            flash(', '.join(msg_parts) + ' 완료!', 'success')
            return redirect(url_for('item.item_list'))

    # GET: 업로드 폼
    return render_template('item_import.html', step='upload')
