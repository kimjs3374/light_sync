"""
BOM 서비스 레이어.

routes/bom.py에서 분리된 비즈니스 로직:
- 엑셀 export/import (파싱, 비교, 적용)
- 소요자재 계산
- 발주서 자동 생성
- 입고 단가 조회
"""

import datetime
import hashlib
import io
import json
import logging
import os
import tempfile
from collections import defaultdict

from sqlalchemy import desc, func
from sqlalchemy.orm import joinedload

from modules.models import (
    BomHeader, BomItem, Contract, ContractItem, Project,
    PurchaseOrder, PurchaseOrderItem, MaterialOrder,
    Receiving, ReceivingItem, ReceivingHistory, Item,
    Vendor, BomModelAlias,
)
from modules.utils import safe_int

logger = logging.getLogger(__name__)


# ===================================================================
# 유틸리티
# ===================================================================

# item_spec_json(영문 키) → BOM option_filter(한국어 키) 매핑
SPEC_TO_OPTION_KEY = {
    'lens_angle': '렌즈',
    'color_temp': '색온도',
    'body_type': '형태',
    'reflector': '반사판',
}

# BOM option_filter에서 사용되는 한국어 키 (직접 저장 시 패스스루)
BOM_OPTION_KEYS = {'렌즈', '색온도', '바이저'}


def translate_spec_to_option(spec_dict):
    """item_spec_json → BOM option_filter 키 변환.

    영문 키(lens_angle)는 한국어(렌즈)로, 한국어 키는 패스스루.
    색온도 미지정 시 5700K 기본값 적용.
    """
    if not spec_dict:
        return {}
    result = {}
    for k, v in spec_dict.items():
        if k in SPEC_TO_OPTION_KEY:
            result[SPEC_TO_OPTION_KEY[k]] = v
        elif k in BOM_OPTION_KEYS:
            result[k] = v
    # 색온도 기본값: 미지정이면 5700K
    if '색온도' not in result:
        result['색온도'] = '5700K'
    return result


def match_option_filter(item_option_json, selected_options):
    """BomItem.option_filter vs 선택된 옵션 매칭 (공통 함수).

    - item_option_json=None → 공통부품, 항상 True
    - selected_options 비어있음 → 전체 포함
    - 키가 일치하고 값이 다르면 False
    """
    if not item_option_json:
        return True
    if not selected_options:
        return True
    try:
        item_opts = json.loads(item_option_json) if isinstance(item_option_json, str) else item_option_json
        for k, v in item_opts.items():
            if k in selected_options and selected_options[k] != v:
                return False
        return True
    except (json.JSONDecodeError, TypeError):
        return True


def _normalize_model_name(model_name):
    """G2B 모델명 정규화 — (M), (D), (X), (H), (3000K) 등 괄호 접미사 제거."""
    import re
    return re.sub(r'\([^)]*\)', '', model_name).strip()


def _extract_model_candidates(model_name):
    """G2B 전체 제품명에서 매칭 후보 모델명 추출.

    G2B 형식: "LED투광등기구, 매그나텍, ARENA(M)-600, 600W"
    → 후보: ["ARENA(M)-600", "ARENA-600", "LED투광등기구, 매그나텍, ARENA(M)-600, 600W"]
    """
    import re
    candidates = []
    # 쉼표로 분리된 토큰 중 모델코드 패턴 추출 (영문+숫자+하이픈)
    if ',' in model_name:
        for token in model_name.split(','):
            token = token.strip()
            # 모델코드 패턴: 영문으로 시작, 하이픈+숫자 포함 (예: ARENA-600, BATOO-1000, MT-FL-300A)
            if re.match(r'^[A-Z].*-\d', token, re.IGNORECASE):
                candidates.append(token)
                normalized = _normalize_model_name(token)
                if normalized != token:
                    candidates.append(normalized)
    # 전체 문자열 자체도 후보 (별칭에 전체명이 등록된 경우 대비)
    candidates.append(model_name)
    normalized_full = _normalize_model_name(model_name)
    if normalized_full != model_name:
        candidates.append(normalized_full)
    return candidates


def _try_find_bom(db, name):
    """단일 이름으로 별칭→product_code 2단계 검색."""
    alias = db.query(BomModelAlias).filter(
        BomModelAlias.alias_name == name
    ).first()
    if alias:
        bom = db.query(BomHeader).get(alias.bom_id)
        if bom and bom.is_active:
            return bom
    bom = db.query(BomHeader).filter(
        BomHeader.is_active == True,
        BomHeader.product_code == name,
    ).first()
    if bom:
        return bom
    return None


def find_bom_by_model_name(db, model_name):
    """모델명으로 BOM 찾기 (별칭 우선, G2B 파싱 + 정규화 fallback).

    1순위: 원본 model_name → 별칭/product_code 정확매칭
    2순위: G2B 전체명 파싱 → 모델코드 토큰 추출 → 별칭/product_code 매칭
    3순위: BomHeader.product_name ILIKE (기존 fallback)
    """
    if not model_name:
        return None
    # 1~2순위: 후보 목록 순회 (원본 → G2B 파싱 토큰 → 정규화)
    for candidate in _extract_model_candidates(model_name):
        bom = _try_find_bom(db, candidate)
        if bom:
            return bom
    # 3순위: product_name 부분 매칭 (기존 fallback)
    return db.query(BomHeader).filter(
        BomHeader.is_active == True,
        BomHeader.product_name.ilike(f'%{model_name}%'),
    ).first()


def parse_option_filter_text(v):
    """옵션조건 텍스트를 JSON 문자열로 변환.

    "렌즈=20도" → '{"렌즈":"20도"}'
    "렌즈=20도,반사판=A" → '{"렌즈":"20도","반사판":"A"}'
    빈칸 → None
    """
    if not v or not str(v).strip():
        return None
    result = {}
    for pair in str(v).strip().split(','):
        pair = pair.strip()
        if '=' not in pair:
            continue
        key, val = pair.split('=', 1)
        key, val = key.strip(), val.strip()
        if key and val:
            result[key] = val
    return json.dumps(result, ensure_ascii=False) if result else None


def get_latest_receiving_prices(db, bom_items):
    """BOM 품목들의 최신 입고 단가를 조회.

    1순위: 자체 입고(receivings + receiving_items) - item_name 매칭
    2순위: iCUBE 입고이력(receiving_history) - items_json 내 item_cd 매칭

    Returns: {item_code: {'unit_price': float, 'date': str, 'source': str}}
    """
    result = {}
    item_codes = {bi.item_code for bi in bom_items if bi.item_code}
    item_names = {bi.item_name for bi in bom_items if bi.item_name}
    if not item_codes and not item_names:
        return result

    # 1) 자체 입고에서 최신 단가 (item_name 매칭)
    for name in item_names:
        row = (
            db.query(ReceivingItem.unit_price, Receiving.rcv_date)
            .join(Receiving, ReceivingItem.receiving_id == Receiving.id)
            .filter(
                ReceivingItem.item_name == name,
                ReceivingItem.unit_price > 0,
            )
            .order_by(desc(Receiving.rcv_date))
            .first()
        )
        if row:
            matched_bi = next((bi for bi in bom_items if bi.item_name == name), None)
            key = matched_bi.item_code if matched_bi and matched_bi.item_code else name
            result[key] = {
                'unit_price': row.unit_price,
                'date': row.rcv_date.strftime('%Y-%m-%d') if row.rcv_date else '',
                'source': '자체입고',
            }

    # 2) iCUBE 입고이력에서 최신 단가 (item_cd 매칭) — 자체 입고에 없는 것만
    remaining_codes = item_codes - set(result.keys())
    if remaining_codes:
        histories = (
            db.query(ReceivingHistory)
            .filter(ReceivingHistory.items_json.isnot(None))
            .order_by(desc(ReceivingHistory.receive_date))
            .limit(500)
            .all()
        )
        for hist in histories:
            if not remaining_codes:
                break
            try:
                items_list = json.loads(hist.items_json) if hist.items_json else []
            except (json.JSONDecodeError, TypeError):
                continue
            for it in items_list:
                code = (it.get('item_cd') or '').strip()
                price = it.get('unit_price') or 0
                if code in remaining_codes and price > 0:
                    result[code] = {
                        'unit_price': price,
                        'date': hist.receive_date.strftime('%Y-%m-%d') if hist.receive_date else '',
                        'source': 'iCUBE',
                    }
                    remaining_codes.discard(code)

    return result


def rebuild_option_schema(db, bom_id):
    """BOM의 모든 아이템에서 option_schema를 재생성."""
    all_items = db.query(BomItem).filter_by(bom_id=bom_id).all()
    new_schema = {}
    for bi in all_items:
        if bi.option_filter:
            try:
                pf = json.loads(bi.option_filter)
                for k, v in pf.items():
                    if k not in new_schema:
                        new_schema[k] = []
                    if v not in new_schema[k]:
                        new_schema[k].append(v)
            except (json.JSONDecodeError, TypeError):
                pass
    return json.dumps(new_schema, ensure_ascii=False) if new_schema else None


# ===================================================================
# BOM 수정 로직
# ===================================================================

def update_bom_items(db, bom, form_data):
    """BOM 품목 업데이트 (기존 업데이트 + 신규 추가 + 삭제).

    form_data: dict with keys item_name[], item_spec[], quantity[],
               unit[], item_note[], unit_price[], option_filter[]
    """
    existing_items = db.query(BomItem).filter_by(bom_id=bom.id).order_by(BomItem.id).all()

    item_names = form_data.get('item_name[]', [])
    item_specs = form_data.get('item_spec[]', [])
    item_qtys = form_data.get('quantity[]', [])
    item_units = form_data.get('unit[]', [])
    item_notes = form_data.get('item_note[]', [])
    unit_prices = form_data.get('unit_price[]', [])
    option_filters = form_data.get('option_filter[]', [])

    for i in range(len(item_names)):
        name = (item_names[i] if i < len(item_names) else '').strip()
        if not name:
            continue

        spec = (item_specs[i] if i < len(item_specs) else '').strip()
        qty = float(item_qtys[i]) if i < len(item_qtys) and item_qtys[i] else 1
        unit = (item_units[i] if i < len(item_units) else '').strip()
        item_note = (item_notes[i] if i < len(item_notes) else '').strip()

        up_str = (unit_prices[i] if i < len(unit_prices) else '').strip().replace(',', '')
        up = float(up_str) if up_str else None
        amount = (up * qty) if up else None

        of_raw = (option_filters[i] if i < len(option_filters) else '').strip()
        of_json = parse_option_filter_text(of_raw) if of_raw else None

        if i < len(existing_items):
            bi = existing_items[i]
            bi.item_name = name
            bi.item_spec = spec
            bi.quantity = qty
            bi.unit = unit
            bi.note = item_note
            bi.option_filter = of_json
            if up is not None and bi.unit_price is not None and up != bi.unit_price:
                bi.prev_unit_price = bi.unit_price
            bi.unit_price = up
            bi.amount = amount
        else:
            db.add(BomItem(
                bom_id=bom.id,
                item_name=name,
                item_spec=spec,
                quantity=qty,
                unit=unit,
                note=item_note,
                unit_price=up,
                amount=amount,
                option_filter=of_json,
            ))

    # form에서 제거된 아이템 삭제
    if len(item_names) < len(existing_items):
        for j in range(len(item_names), len(existing_items)):
            db.delete(existing_items[j])

    # option_schema 재생성
    bom.option_schema = rebuild_option_schema(db, bom.id)


# ===================================================================
# 엑셀 Export
# ===================================================================

def export_bom_excel(db):
    """BOM 목록을 엑셀 파일(BytesIO)로 생성하여 반환."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    boms = db.query(BomHeader).options(
        joinedload(BomHeader.bom_items),
    ).order_by(BomHeader.product_category, BomHeader.product_code).all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=9)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 제품군별 시트
    categories = {}
    for bom in boms:
        cat = bom.product_category or '기타'
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(bom)

    for cat, cat_boms in categories.items():
        ws_name = cat[:31]
        ws = wb.create_sheet(title=ws_name)

        headers = ['No', '완제품코드', '제품명', '인증번호', '품목코드', '품목명',
                    '규격', '소요량', '단가', '금액', '납품업체', '비고', '옵션조건']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        row = 2
        for bom in cat_boms:
            for idx, bi in enumerate(bom.bom_items):
                of_text = ''
                if bi.option_filter:
                    try:
                        pf = json.loads(bi.option_filter)
                        of_text = ', '.join(f'{k}={v}' for k, v in pf.items())
                    except (json.JSONDecodeError, TypeError):
                        pass

                row_data = [
                    idx + 1 if idx == 0 else '',
                    bom.product_code if idx == 0 else '',
                    bom.product_name if idx == 0 else '',
                    bom.certification_no or '' if idx == 0 else '',
                    bi.item_code or '',
                    bi.item_name or '',
                    bi.item_spec or '',
                    bi.quantity or 0,
                    bi.unit_price or 0,
                    bi.amount or 0,
                    bi.supplier or '',
                    bi.note or '',
                    of_text,
                ]
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    cell.font = Font(size=9)
                row += 1

            if bom.bom_items:
                row += 1

        widths = [6, 18, 20, 15, 15, 25, 20, 8, 10, 12, 15, 15, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ===================================================================
# 엑셀 Import — 1단계: 파싱 + 비교
# ===================================================================

def parse_bom_excel(file_obj):
    """엑셀 파일을 파싱하여 BOM 데이터 딕셔너리 반환.

    Returns: parsed_boms dict {product_code: {product_code, sheet, items: [...]}}
             또는 에러 시 None
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_obj, read_only=False, data_only=True)
    except Exception as e:
        logger.error(f'BOM 엑셀 로드 실패: {e}')
        return None, str(e)

    parsed_boms = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # 헤더 행 찾기
        header_idx = None
        col_map = {}
        for ri, row in enumerate(rows):
            cells = [str(c).strip() if c else '' for c in row]
            for ci, cv in enumerate(cells):
                lv = cv.lower()
                if cv in ('No', 'NO', 'no', 'No.'):
                    col_map['no'] = ci
                elif cv in ('완제품코드', '제품코드') or lv == 'product_code':
                    col_map['product_code'] = ci
                elif cv in ('품목코드', '부품코드') or lv == 'item_code':
                    col_map['item_code'] = ci
                elif cv in ('품목명', '부품명', '품명') or lv == 'item_name':
                    col_map['item_name'] = ci
                elif cv in ('소요량', '수량') or lv == 'quantity':
                    col_map['quantity'] = ci
                elif cv in ('단가',) or lv == 'unit_price':
                    col_map['unit_price'] = ci
                elif cv in ('부품업체', '납품업체', '업체') or lv == 'supplier':
                    col_map['supplier'] = ci
                elif cv in ('비고', '원수') or lv == 'note':
                    if 'note' not in col_map:
                        col_map['note'] = ci
                elif cv in ('규격', '사양') or lv == 'spec':
                    col_map['spec'] = ci
                elif cv in ('옵션조건', '옵션', 'option', 'option_filter') or lv == 'option_filter':
                    col_map['option_filter'] = ci
            if 'item_name' in col_map or 'item_code' in col_map:
                header_idx = ri
                break

        if header_idx is None:
            continue

        # 데이터 파싱
        for row in rows[header_idx + 1:]:
            cells = [str(c).strip() if c else '' for c in row]

            def get_col(key, _cells=cells):
                idx = col_map.get(key)
                return _cells[idx] if idx is not None and idx < len(_cells) else ''

            item_name = get_col('item_name')
            if not item_name:
                continue

            product_code = get_col('product_code')
            if not product_code:
                continue

            qty_str = get_col('quantity')
            try:
                qty = float(qty_str) if qty_str and not qty_str.startswith('=') else 1
            except (ValueError, TypeError):
                qty = 1

            price_str = get_col('unit_price')
            try:
                price = float(price_str) if price_str and not price_str.startswith('=') else 0
            except (ValueError, TypeError):
                price = 0

            if product_code not in parsed_boms:
                parsed_boms[product_code] = {
                    'product_code': product_code,
                    'sheet': sheet_name,
                    'items': [],
                }

            parsed_boms[product_code]['items'].append({
                'item_code': get_col('item_code'),
                'item_name': item_name,
                'item_spec': get_col('spec'),
                'quantity': qty,
                'unit_price': price,
                'supplier': get_col('supplier'),
                'note': get_col('note'),
                'option_filter': parse_option_filter_text(get_col('option_filter')),
            })

    wb.close()
    return parsed_boms, None


def compare_bom_with_db(db, parsed_boms):
    """파싱된 BOM과 DB 기존 BOM을 비교하여 결과 리스트 반환."""
    existing_boms = db.query(BomHeader).options(
        joinedload(BomHeader.bom_items),
    ).all()
    db_by_code = {b.product_code: b for b in existing_boms}

    results = []
    for pc, parsed in parsed_boms.items():
        db_bom = db_by_code.get(pc)
        if db_bom:
            db_items = {(bi.item_name, bi.item_code or ''): bi for bi in db_bom.bom_items}
            excel_items = parsed['items']

            item_changes = []
            new_items = 0
            changed_items = 0
            for ei in excel_items:
                key = (ei['item_name'], ei['item_code'])
                if key in db_items:
                    dbi = db_items[key]
                    diffs = []
                    if ei['quantity'] != (dbi.quantity or 0):
                        diffs.append(f"소요량: {dbi.quantity}→{ei['quantity']}")
                    if ei['unit_price'] and ei['unit_price'] != (dbi.unit_price or 0):
                        diffs.append(f"단가: {dbi.unit_price or 0}→{ei['unit_price']}")
                    if ei['supplier'] and ei['supplier'] != (dbi.supplier or ''):
                        diffs.append(f"업체: {dbi.supplier or ''}→{ei['supplier']}")
                    if diffs:
                        changed_items += 1
                        item_changes.append({'name': ei['item_name'], 'diffs': diffs})
                else:
                    new_items += 1
                    item_changes.append({'name': ei['item_name'], 'diffs': ['신규 품목']})

            status = 'changed' if (new_items > 0 or changed_items > 0) else 'same'
            results.append({
                'product_code': pc,
                'sheet': parsed['sheet'],
                'excel_item_count': len(excel_items),
                'db_item_count': len(db_bom.bom_items),
                'db_id': db_bom.id,
                'db_category': db_bom.product_category or '',
                'status': status,
                'new_items': new_items,
                'changed_items': changed_items,
                'item_changes': item_changes[:5],
            })
        else:
            results.append({
                'product_code': pc,
                'sheet': parsed['sheet'],
                'excel_item_count': len(parsed['items']),
                'db_item_count': 0,
                'db_id': None,
                'db_category': '',
                'status': 'new',
                'new_items': len(parsed['items']),
                'changed_items': 0,
                'item_changes': [],
            })

    return results


def save_import_cache(parsed_boms, results):
    """파싱 결과를 임시 파일에 저장하고 cache_id 반환."""
    cache_data = {'parsed': parsed_boms, 'results': results}
    cache_id = hashlib.md5(
        json.dumps(cache_data, ensure_ascii=False, default=str).encode()
    ).hexdigest()[:12]
    cache_dir = os.path.join(tempfile.gettempdir(), 'lightsync_bom_import')
    os.makedirs(cache_dir, exist_ok=True)
    cache_path = os.path.join(cache_dir, f'{cache_id}.json')
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(cache_data, f, ensure_ascii=False, default=str)
    return cache_id


def load_import_cache(cache_id):
    """캐시 파일을 로드하고 삭제. 없으면 None 반환."""
    cache_path = os.path.join(tempfile.gettempdir(), 'lightsync_bom_import', f'{cache_id}.json')
    if not cache_id or not os.path.exists(cache_path):
        return None
    with open(cache_path, 'r', encoding='utf-8') as f:
        cache_data = json.load(f)
    os.remove(cache_path)
    return cache_data


# ===================================================================
# 엑셀 Import — 2단계: 적용
# ===================================================================

def apply_bom_import(db, parsed_boms, results, selected_indices):
    """선택된 BOM을 DB에 적용 (신규 생성 / 기존 갱신).

    Returns: (added_count, updated_count)
    """
    added_boms = 0
    updated_boms = 0

    for i, r in enumerate(results):
        if str(i) not in selected_indices:
            continue

        pc = r['product_code']
        parsed = parsed_boms.get(pc)
        if not parsed:
            continue

        if r['status'] == 'new':
            # option_schema 자동 생성
            option_schema = {}
            for ei in parsed['items']:
                of = ei.get('option_filter')
                if of:
                    try:
                        pf = json.loads(of)
                        for k, v in pf.items():
                            if k not in option_schema:
                                option_schema[k] = []
                            if v not in option_schema[k]:
                                option_schema[k].append(v)
                    except (json.JSONDecodeError, TypeError):
                        pass

            bom = BomHeader(
                product_code=pc,
                product_name=pc,
                product_category=parsed['sheet'],
                is_active=True,
                option_schema=json.dumps(option_schema, ensure_ascii=False) if option_schema else None,
            )
            db.add(bom)
            db.flush()

            for ei in parsed['items']:
                db.add(BomItem(
                    bom_id=bom.id,
                    item_code=ei['item_code'] or None,
                    item_name=ei['item_name'],
                    item_spec=ei.get('item_spec') or None,
                    quantity=ei['quantity'],
                    unit_price=ei['unit_price'] or None,
                    amount=(ei['quantity'] * ei['unit_price']) if ei['unit_price'] else None,
                    supplier=ei['supplier'] or None,
                    note=ei.get('note') or None,
                    option_filter=ei.get('option_filter'),
                ))
            added_boms += 1

        elif r['status'] == 'changed' and r['db_id']:
            bom = db.query(BomHeader).get(r['db_id'])
            if not bom:
                continue

            existing_items = {(bi.item_name, bi.item_code or ''): bi for bi in bom.bom_items}

            for ei in parsed['items']:
                key = (ei['item_name'], ei['item_code'])
                if key in existing_items:
                    bi = existing_items[key]
                    if ei['quantity']:
                        bi.quantity = ei['quantity']
                    if ei['unit_price']:
                        if bi.unit_price and ei['unit_price'] != bi.unit_price:
                            bi.prev_unit_price = bi.unit_price
                        bi.unit_price = ei['unit_price']
                        bi.amount = ei['quantity'] * ei['unit_price']
                    if ei['supplier']:
                        bi.supplier = ei['supplier']
                else:
                    db.add(BomItem(
                        bom_id=bom.id,
                        item_code=ei['item_code'] or None,
                        item_name=ei['item_name'],
                        item_spec=ei.get('item_spec') or None,
                        quantity=ei['quantity'],
                        unit_price=ei['unit_price'] or None,
                        amount=(ei['quantity'] * ei['unit_price']) if ei['unit_price'] else None,
                        supplier=ei['supplier'] or None,
                        note=ei.get('note') or None,
                        option_filter=ei.get('option_filter'),
                    ))
            updated_boms += 1

    return added_boms, updated_boms


# ===================================================================
# 소요자재 계산
# ===================================================================

def calculate_material_requirement(db, contract_id):
    """계약 ID 기반으로 소요자재를 계산하여 리스트 반환.

    Returns: (selected_contract, requirement_list)
    """
    selected_contract = db.query(Contract).options(
        joinedload(Contract.items),
        joinedload(Contract.project),
    ).get(contract_id)

    if not selected_contract:
        return None, []

    requirement = []
    for ci in selected_contract.items:
        # BOM 매칭: 별칭 우선 → product_code → ILIKE fallback
        bom = find_bom_by_model_name(db, ci.model_name)
        if not bom and ci.category:
            bom = find_bom_by_model_name(db, ci.category)

        if bom:
            # 슈퍼BOM: 협의내용(item_spec_json) → 옵션 키 변환 → 필터링
            selected_options = {}
            if ci.item_spec_json:
                try:
                    raw = json.loads(ci.item_spec_json) if isinstance(ci.item_spec_json, str) else ci.item_spec_json
                    selected_options = translate_spec_to_option(raw)
                except (json.JSONDecodeError, TypeError):
                    pass

            for bi in bom.bom_items:
                # 공통 옵션 필터 적용
                if not match_option_filter(bi.option_filter, selected_options):
                    continue

                needed = (bi.quantity or 0) * (ci.quantity or 0)

                # 1차: bom_item_id 기반 PurchaseOrderItem 발주량
                ordered_via_po = db.query(
                    func.coalesce(func.sum(PurchaseOrderItem.quantity), 0)
                ).join(
                    PurchaseOrder, PurchaseOrderItem.po_id == PurchaseOrder.id
                ).filter(
                    PurchaseOrderItem.bom_item_id == bi.id,
                    PurchaseOrder.contract_id == contract_id,
                    PurchaseOrder.status != '취소',
                ).scalar() or 0

                # 2차: MaterialOrder 기반 (하위 호환)
                ordered_via_mo = db.query(
                    func.coalesce(func.sum(MaterialOrder.quantity), 0)
                ).filter(
                    MaterialOrder.contract_item_id == ci.id,
                    MaterialOrder.order_status.in_(['발주완료', '입고완료']),
                ).scalar() or 0

                ordered = max(float(ordered_via_po), float(ordered_via_mo))

                received_via_mo = db.query(
                    func.coalesce(func.sum(MaterialOrder.quantity), 0)
                ).filter(
                    MaterialOrder.contract_item_id == ci.id,
                    MaterialOrder.order_status == '입고완료',
                ).scalar() or 0

                shortage = max(0, needed - ordered)

                requirement.append({
                    'contract_item': ci,
                    'bom_item': bi,
                    'product_qty': ci.quantity or 0,
                    'per_unit': bi.quantity or 0,
                    'needed': needed,
                    'ordered': ordered,
                    'received': float(received_via_mo),
                    'shortage': shortage,
                })
        else:
            requirement.append({
                'contract_item': ci,
                'bom_item': None,
                'product_qty': ci.quantity or 0,
                'per_unit': 0,
                'needed': 0,
                'ordered': 0,
                'received': 0,
                'shortage': 0,
                'no_bom': True,
            })

    return selected_contract, requirement


# ===================================================================
# 소요자재 → 발주서 자동 생성
# ===================================================================

def create_po_from_material_requirement(db, contract_id, selected_items, user_id):
    """소요자재 부족분에서 거래처별 발주서를 자동 생성.

    Returns: list of created PurchaseOrder objects
    """
    from routes.purchase_order import _generate_po_no

    contract = db.query(Contract).get(contract_id)
    if not contract:
        return []

    # supplier 기준 그룹핑
    groups = defaultdict(list)
    for item in selected_items:
        supplier = (item.get('supplier') or '미지정').strip()
        groups[supplier].append(item)

    created_pos = []
    for supplier_name, items in groups.items():
        vendor = db.query(Vendor).filter(
            Vendor.name.ilike(f'%{supplier_name}%'),
            Vendor.is_active == True,
        ).first()

        if not vendor:
            vendor = Vendor(name=supplier_name, is_active=True)
            db.add(vendor)
            db.flush()

        po = PurchaseOrder(
            po_no=_generate_po_no(db),
            po_date=datetime.date.today(),
            vendor_id=vendor.id,
            contract_id=contract_id,
            project_id=contract.project_id,
            assigned_to=user_id,
            status='작성중',
            note='BOM 소요자재 자동생성',
            created_by=user_id,
        )
        db.add(po)
        db.flush()

        total_amount = 0
        for item in items:
            qty = float(item.get('quantity') or 0)
            price = float(item.get('unit_price') or 0)
            amount = qty * price

            po_item = PurchaseOrderItem(
                po_id=po.id,
                item_name=item.get('item_name', ''),
                item_spec=item.get('item_spec', ''),
                quantity=qty,
                unit_price=price,
                amount=amount,
                unit=item.get('unit', ''),
                bom_item_id=safe_int(item.get('bom_item_id'), 0) or None,
            )
            db.add(po_item)
            total_amount += amount

        po.total_amount = total_amount
        po.tax_amount = round(total_amount * 0.1)
        created_pos.append(po)

    return created_pos
