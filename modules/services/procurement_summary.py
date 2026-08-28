"""납품집계 서비스 — G2B 조달내역 월별 피벗 집계 + 엑셀 생성"""
import datetime
import re
from io import BytesIO

from sqlalchemy import extract, func, or_, false

from modules.models import G2bProcurement, TaxInvoice


CHART_COLORS = [
    '#3b82f6', '#ef4444', '#10b981', '#f59e0b', '#8b5cf6',
    '#ec4899', '#06b6d4', '#f97316', '#6366f1', '#84cc16',
]


_HANGUL = re.compile(r'[가-힣]')


def split_model_name(raw, category=None):
    """G2B 물품규격명을 (품목, 모델명) 두 조각으로 자른다.

    원본은 `품목, 제조사, 모델코드, 부가스펙…` 형태의 한 줄짜리 긴 문자열이라
    그대로 두면 표에서 읽히지 않는다.
        'LED보안등기구, 매그나텍, MT-SL-030, 30W' -> ('LED보안등기구', 'MT-SL-030')

    모델코드는 콤마 토큰 중 **한글이 없고 영문자를 포함한 첫 토큰**으로 잡는다.
    제조사(매그나텍/천일)는 한글이라 자동으로 걸러지고, 뒤에 붙는 '30W'·'10m'
    같은 부가스펙은 코드보다 뒤에 오므로 먼저 잡히지 않는다.

    모델코드가 아예 없는 수요기관 규격품(185종 중 12종)은 '수요기관규격'으로
    통일해 표기한다. 원본 문자열은 화면 tooltip 으로 남긴다.
    """
    toks = [t.strip() for t in (raw or '').split(',')]
    toks = [t for t in toks if t]
    if not toks:
        return (category or '-'), '-'

    item, rest = toks[0], toks[1:]

    for t in rest:
        if not _HANGUL.search(t) and re.search(r'[A-Za-z]', t) and 2 <= len(t) <= 40:
            return item, t

    if not rest:
        return item, '-'
    tail = rest[-1]
    if '수요기관' in tail.replace(' ', ''):
        return item, '수요기관규격'
    return item, tail


def summary_base_filters():
    """납품집계 공용 기준 — 웹/모바일/MCP 3경로가 이 하나를 공유한다.

    prdct_amt > 0
        금액 0/NULL 행은 집계 대상이 아니다. 건수만 세는 경로가 이 필터를
        빼면 같은 화면인데 건수가 달라진다.
    최종변경차수만
        유니크키가 (계약납품요구번호, 품목순번, 변경차수)라서 변경계약이
        원차수와 나란히 남으면 금액이 이중 계상된다. 지금 데이터는 최종차수만
        들어와 있어 결과가 같지만, 동기화가 바뀌어도 집계가 조용히 틀리지
        않도록 쿼리 쪽에도 방어선을 둔다. NULL 은 과거 데이터에 있을 수 있어
        남기고 'N' 만 제외한다.
    """
    return [
        G2bProcurement.prdct_amt > 0,
        or_(
            G2bProcurement.fnl_cntrct_dlvr_req_chg_ord_yn.is_(None),
            G2bProcurement.fnl_cntrct_dlvr_req_chg_ord_yn != 'N',
        ),
    ]


def get_filter_options(db):
    """필터 옵션: 년도 + 대분류(세부품명) 목록"""
    base = summary_base_filters()

    years = db.query(
        extract('year', G2bProcurement.cntrct_dlvr_req_date).label('yr')
    ).filter(
        *base,
    ).distinct().order_by(
        extract('year', G2bProcurement.cntrct_dlvr_req_date).desc()
    ).all()

    categories = db.query(
        G2bProcurement.dtil_prdct_clsfc_no_nm
    ).filter(
        *base,
        G2bProcurement.dtil_prdct_clsfc_no_nm.isnot(None),
    ).distinct().order_by(
        G2bProcurement.dtil_prdct_clsfc_no_nm
    ).all()

    return {
        'years': [int(r[0]) for r in years if r[0]],
        'categories': [r[0] for r in categories if r[0]],
    }


def get_summary_pivot(db, years=None, category=None, q=None, group_by='category'):
    """납품집계 피벗

    group_by:
        'category' = 대분류(dtil_prdct_clsfc_no_nm) 기준
        'model'    = 모델(prdct_idnt_no_nm) 기준
    """
    filters = summary_base_filters()
    if years:
        filters.append(extract('year', G2bProcurement.cntrct_dlvr_req_date).in_(years))
    if category:
        filters.append(G2bProcurement.dtil_prdct_clsfc_no_nm == category)
    if q:
        filters.append(G2bProcurement.prdct_idnt_no_nm.ilike(f'%{q}%'))

    if group_by == 'model':
        group_col = G2bProcurement.prdct_idnt_no_nm
    else:
        group_col = G2bProcurement.dtil_prdct_clsfc_no_nm

    rows = db.query(
        group_col.label('name'),
        extract('year', G2bProcurement.cntrct_dlvr_req_date).label('year'),
        extract('month', G2bProcurement.cntrct_dlvr_req_date).label('month'),
        func.sum(G2bProcurement.prdct_qty).label('total_qty'),
        func.sum(G2bProcurement.prdct_amt).label('total_amt'),
    ).filter(*filters).group_by(
        group_col,
        extract('year', G2bProcurement.cntrct_dlvr_req_date),
        extract('month', G2bProcurement.cntrct_dlvr_req_date),
    ).all()

    multi_year = years and len(years) > 1
    return _build_pivot(rows, multi_year=multi_year)


def _invoice_done_subq(db):
    """계약별 완료일자 = 매칭된 매출 세금계산서의 최종 발행일.

    부분청구로 계산서가 여러 장인 계약이 112건 있어 MAX 를 쓴다.
    (마지막 청구가 곧 완료 시점)

    합산이 아니라 MAX 라서 승인번호 하이픈 유무로 인한 중복행이 있어도
    날짜는 왜곡되지 않는다. 금액을 더할 때는 반드시
    tax_invoice_agg.deduped_invoice_subq() 를 쓸 것.
    """
    return db.query(
        TaxInvoice.g2b_contract_no.label('g2b_no'),
        func.max(TaxInvoice.issue_date).label('done_date'),
    ).filter(
        TaxInvoice.g2b_contract_no.isnot(None),
        TaxInvoice.direction == '매출',
        TaxInvoice.match_status.in_(['자동매칭', '수동매칭']),
    ).group_by(TaxInvoice.g2b_contract_no).subquery()


def _resolve_model_raws(db, model, category=None):
    """model 파라미터를 실제 규격명(raw) 목록으로 푼다.

    두 종류가 들어온다.
      · 피벗 드릴다운 → 원본 규격명 전체 ('LED보안등기구, 매그나텍, MT-SL-030, 30W')
      · 세부조회 필터 → 화면에 보이는 짧은 코드 ('MT-SL-030')
    둘 다 받아야 해서 원본 우선으로 맞춰보고, 없으면 코드로 푼다.

    '수요기관규격'처럼 원본 표기가 여러 갈래인 코드는 그 갈래를 전부 모은다.
    """
    if not model:
        return None

    cand = db.query(
        G2bProcurement.dtil_prdct_clsfc_no_nm,
        G2bProcurement.prdct_idnt_no_nm,
    ).filter(*summary_base_filters())
    if category:
        cand = cand.filter(G2bProcurement.dtil_prdct_clsfc_no_nm == category)

    raws = []
    for cat, raw in cand.distinct().all():
        if raw == model:
            return [raw]
        if split_model_name(raw, cat)[1] == model:
            raws.append(raw)
    return raws


def get_detail_filter_options(db, category=None):
    """세부조회 필터 옵션 — 연도 + 품목(대분류) + 모델코드

    모델 목록은 품목이 선택돼 있으면 그 품목 것만 준다.
    (전체 181종을 한 드롭다운에 쏟으면 고를 수가 없다)
    """
    opts = get_filter_options(db)

    q = db.query(
        G2bProcurement.dtil_prdct_clsfc_no_nm,
        G2bProcurement.prdct_idnt_no_nm,
    ).filter(*summary_base_filters())
    if category:
        q = q.filter(G2bProcurement.dtil_prdct_clsfc_no_nm == category)

    seen = {}
    for cat, raw in q.distinct().all():
        item, code = split_model_name(raw, cat)
        seen.setdefault(code, item)

    opts['models'] = [
        {'code': c, 'item': seen[c]} for c in sorted(seen, key=lambda x: (x == '-', x))
    ]
    return opts


def get_summary_detail(db, years=None, category=None, q=None,
                       month=None, model=None):
    """납품집계 세부조회 — 품목 단위 원장

    피벗 셀 뒤에 실제로 어떤 품목이 깔려 있는지 보기 위한 목록.
    필터는 피벗과 같은 summary_base_filters() 를 공유하므로
    세부 합계가 피벗 셀 값과 항상 일치한다.

    계약일자 = intl_cntrct_dlvr_req_date (G2B 최초계약납품요구일자)
               변경계약이면 납품요구일과 달라지는 151건이 있다.
    완료일자 = 매칭된 매출 세금계산서 최종 발행일 (미발행이면 None)
    """
    inv = _invoice_done_subq(db)

    filters = summary_base_filters()
    if years:
        filters.append(extract('year', G2bProcurement.cntrct_dlvr_req_date).in_(years))
    if category:
        filters.append(G2bProcurement.dtil_prdct_clsfc_no_nm == category)
    if q:
        filters.append(G2bProcurement.prdct_idnt_no_nm.ilike(f'%{q}%'))
    if model:
        raws = _resolve_model_raws(db, model, category)
        # 못 찾으면 0건. 조용히 전체를 보여주면 안 된다.
        # (문자열 센티널은 쓸 수 없다 — PostgreSQL 은 NUL 문자를 거부한다)
        filters.append(G2bProcurement.prdct_idnt_no_nm.in_(raws) if raws else false())
    if month:
        filters.append(extract('month', G2bProcurement.cntrct_dlvr_req_date) == month)

    rows = db.query(
        G2bProcurement.intl_cntrct_dlvr_req_date.label('contract_date'),
        inv.c.done_date.label('done_date'),
        G2bProcurement.prdct_idnt_no_nm.label('model_name'),
        G2bProcurement.prdct_qty.label('qty'),
        G2bProcurement.prdct_amt.label('amt'),
        G2bProcurement.cntrct_dlvr_req_nm.label('contract_name'),
        G2bProcurement.cntrct_dlvr_req_no.label('req_no'),
        G2bProcurement.dtil_prdct_clsfc_no_nm.label('category'),
        G2bProcurement.dminstt_nm.label('demand_org'),
    ).outerjoin(
        inv, inv.c.g2b_no == G2bProcurement.cntrct_dlvr_req_no
    ).filter(*filters).order_by(
        G2bProcurement.intl_cntrct_dlvr_req_date.desc(),
        G2bProcurement.cntrct_dlvr_req_no.desc(),
        G2bProcurement.prdct_sno,
    ).all()

    items = []
    for r in rows:
        item_name, model_code = split_model_name(r.model_name, r.category)
        items.append({
            'contract_date': r.contract_date,
            'done_date': r.done_date,
            'item_name': item_name,
            'model_code': model_code,
            'model_raw': r.model_name or '-',
            'qty': int(r.qty or 0),
            'amt': int(r.amt or 0),
            'contract_name': r.contract_name or '-',
            'req_no': r.req_no,
            'category': r.category or '(미분류)',
            'demand_org': r.demand_org or '-',
        })

    return {
        'items': items,
        'total_qty': sum(i['qty'] for i in items),
        'total_amt': sum(i['amt'] for i in items),
        'done_count': sum(1 for i in items if i['done_date']),
    }


def _empty_months():
    return {i: {'qty': 0, 'amt': 0} for i in range(1, 13)}


def _add_months(target, source):
    for m in range(1, 13):
        target[m]['qty'] += source[m]['qty']
        target[m]['amt'] += source[m]['amt']


def _build_pivot(rows, multi_year=False):
    """쿼리 결과 → 피벗 dict

    multi_year=True 시 각 항목에 년도별 sub_rows 포함:
    models[].sub_rows = [{'year': 2024, 'months': {...}, 'total_qty', 'total_amt'}, ...]
    """
    # name → year → month 집계
    raw = {}
    for r in rows:
        name = r.name or '(미분류)'
        yr = int(r.year)
        m = int(r.month)
        qty = int(r.total_qty or 0)
        amt = int(r.total_amt or 0)

        if name not in raw:
            raw[name] = {}
        if yr not in raw[name]:
            raw[name][yr] = {'months': _empty_months(), 'total_qty': 0, 'total_amt': 0}

        entry = raw[name][yr]
        entry['months'][m]['qty'] += qty
        entry['months'][m]['amt'] += amt
        entry['total_qty'] += qty
        entry['total_amt'] += amt

    # 합산 + sub_rows 생성
    model_map = {}
    for name, year_data in raw.items():
        agg = {'name': name, 'months': _empty_months(), 'total_qty': 0, 'total_amt': 0, 'sub_rows': []}
        for yr in sorted(year_data.keys()):
            yd = year_data[yr]
            _add_months(agg['months'], yd['months'])
            agg['total_qty'] += yd['total_qty']
            agg['total_amt'] += yd['total_amt']
            if multi_year:
                agg['sub_rows'].append({
                    'year': yr,
                    'months': yd['months'],
                    'total_qty': yd['total_qty'],
                    'total_amt': yd['total_amt'],
                })
        model_map[name] = agg

    models = sorted(model_map.values(), key=lambda x: -x['total_amt'])

    grand = {'months': _empty_months(), 'total_qty': 0, 'total_amt': 0}
    for mdl in models:
        _add_months(grand['months'], mdl['months'])
        grand['total_qty'] += mdl['total_qty']
        grand['total_amt'] += mdl['total_amt']

    return {'models': models, 'grand_total': grand}


def build_chart_data(pivot):
    """Chart.js stacked bar (금액 기준, 상위 10개)"""
    labels = [f'{m}월' for m in range(1, 13)]
    datasets = []
    for idx, mdl in enumerate(pivot['models'][:10]):
        color = CHART_COLORS[idx % len(CHART_COLORS)]
        datasets.append({
            'label': mdl['name'][:30],
            'data': [mdl['months'][m]['amt'] for m in range(1, 13)],
            'backgroundColor': color + 'cc',
            'borderColor': color,
            'borderWidth': 1,
        })
    return {'labels': labels, 'datasets': datasets}


def generate_detail_excel(detail, years, title='납품집계_세부', hide_financial=False):
    """세부조회 → xlsx (품목 원장)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    thin = Side(style='thin', color='999999')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    total_fill = PatternFill(start_color='e2e8f0', end_color='e2e8f0', fill_type='solid')
    total_font = Font(bold=True, size=10)

    year_str = ','.join(str(y) for y in years) if years else '전체'
    ws['A1'] = f'{title} ({year_str})'
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['계약일자', '완료일자', '품목', '모델명', '수량']
    if not hide_financial:
        headers.append('금액')
    headers += ['계약명', '수요기관', '납품요구번호']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row_idx = 4
    for it in detail['items']:
        vals = [
            it['contract_date'], it['done_date'],
            it['item_name'], it['model_code'], it['qty'],
        ]
        if not hide_financial:
            vals.append(it['amt'])
        vals += [it['contract_name'], it['demand_org'], it['req_no']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = border
            if isinstance(v, int):
                cell.number_format = '#,##0'
            elif hasattr(v, 'year'):
                cell.number_format = 'yyyy-mm-dd'
        row_idx += 1

    ws.cell(row=row_idx, column=1, value='합계').font = total_font
    ws.cell(row=row_idx, column=1).fill = total_fill
    ws.cell(row=row_idx, column=1).border = border
    qty_cell = ws.cell(row=row_idx, column=5, value=detail['total_qty'])
    qty_cell.font, qty_cell.fill, qty_cell.border = total_font, total_fill, border
    qty_cell.number_format = '#,##0'
    if not hide_financial:
        amt_cell = ws.cell(row=row_idx, column=6, value=detail['total_amt'])
        amt_cell.font, amt_cell.fill, amt_cell.border = total_font, total_fill, border
        amt_cell.number_format = '#,##0'

    widths = [12, 12, 20, 20, 9] + ([15] if not hide_financial else []) + [46, 26, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_excel(pivot, years, title='납품집계'):
    """피벗 → xlsx"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    thin = Side(style='thin', color='999999')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    total_fill = PatternFill(start_color='e2e8f0', end_color='e2e8f0', fill_type='solid')
    total_font = Font(bold=True, size=10)
    num_fmt = '#,##0'

    year_str = ','.join(str(y) for y in years) if years else '전체'
    ws.merge_cells('A1:Z1')
    ws['A1'] = f'{title} ({year_str})'
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['구분']
    for m in range(1, 13):
        headers.append(f'{m}월(수량)')
        headers.append(f'{m}월(금액)')
    headers += ['합계(수량)', '합계(금액)']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row_idx = 4
    for mdl in pivot['models']:
        ws.cell(row=row_idx, column=1, value=mdl['name']).border = border
        col = 2
        for m in range(1, 13):
            for val in [mdl['months'][m]['qty'], mdl['months'][m]['amt']]:
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.number_format = num_fmt
                cell.border = border
                col += 1
        for val in [mdl['total_qty'], mdl['total_amt']]:
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.number_format = num_fmt
            cell.border = border
            col += 1
        row_idx += 1

    gt = pivot['grand_total']
    ws.cell(row=row_idx, column=1, value='합계').font = total_font
    ws.cell(row=row_idx, column=1).fill = total_fill
    ws.cell(row=row_idx, column=1).border = border
    col = 2
    for m in range(1, 13):
        for val in [gt['months'][m]['qty'], gt['months'][m]['amt']]:
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.number_format = num_fmt
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            col += 1
    for val in [gt['total_qty'], gt['total_amt']]:
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.number_format = num_fmt
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        col += 1

    ws.column_dimensions['A'].width = 40
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 13

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
