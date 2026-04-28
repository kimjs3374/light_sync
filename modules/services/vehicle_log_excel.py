"""업무용차량 운행기록부 별지서식 엑셀 생성

세법 별지 서식 (2016.4.1. 제정) 양식 그대로:
- 헤더: 사업연도 / 법인명 / 사업자등록번호
- 1. 기본정보: 차종(번호), 업무사용비율(100% 고정)
- 2. 업무용 사용비율 계산: 사용일자, 사용자(부서/성명), 운행내역(주행전/후/거리/주유금액/출발지/도착지/사용목적)
- 합계 행: 주행거리, 주유금액
"""
from io import BytesIO
import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COMPANY_NAME = "㈜매그나텍"
COMPANY_BIZ_NO = "408-81-68519"
USAGE_RATIO_LABEL = "업무사용비율: 100%"


def export_vehicle_log_excel(logs, vehicle: str, year: int) -> BytesIO:
    """운행일지 목록(logs) → 별지서식 엑셀 BytesIO 반환

    logs: list[VehicleLog] (use_date 오름차순 정렬됨)
    vehicle: 차종(번호)
    year: 사업연도 (4자리)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년 운행기록부"

    thin = Side(style='thin', color='000000')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True, size=10)
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    light_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')

    # 컬럼 너비 (총 11컬럼: A=사용일자 B=부서 C=성명 D=주행전 E=주행후 F=주행거리 G=주유금액 H=출발지 I=도착지 J=사용목적 K=비고)
    widths = [12, 10, 10, 11, 11, 11, 11, 18, 18, 25, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ========== Row 1: 별지서식 표기 ==========
    ws.cell(row=1, column=1, value="【업무용승용차 운행기록부에 관한 별지 서식】<2016.4.1. 제정>")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws.cell(row=1, column=1).font = Font(size=9, italic=True)
    ws.cell(row=1, column=1).alignment = left

    # ========== Row 2~3: 사업연도 / 제목 / 법인명·사업자등록번호 ==========
    # A2~B3: 사업연도 라벨 + 값
    ws.cell(row=2, column=1, value="사업연도").font = bold
    ws.cell(row=2, column=1).alignment = center
    ws.cell(row=2, column=1).border = box
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)

    period = f"{year}.01.01 ~\n{year}.12.31"
    ws.cell(row=2, column=2, value=period).alignment = center
    ws.cell(row=2, column=2).border = box
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)

    # C2~H3: 제목 (병합)
    ws.cell(row=2, column=3, value="업무용차량 운행기록부").font = title_font
    ws.cell(row=2, column=3).alignment = center
    ws.cell(row=2, column=3).border = box
    ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=8)

    # I2: 법인명 라벨, J2~K2: 값
    ws.cell(row=2, column=9, value="법 인 명").font = bold
    ws.cell(row=2, column=9).alignment = center
    ws.cell(row=2, column=9).border = box
    ws.cell(row=2, column=10, value=COMPANY_NAME).alignment = center
    ws.cell(row=2, column=10).border = box
    ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column=11)

    # I3: 사업자등록번호 라벨, J3~K3: 값
    ws.cell(row=3, column=9, value="사업자등록번호").font = bold
    ws.cell(row=3, column=9).alignment = center
    ws.cell(row=3, column=9).border = box
    ws.cell(row=3, column=10, value=COMPANY_BIZ_NO).alignment = center
    ws.cell(row=3, column=10).border = box
    ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column=11)

    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    # ========== Row 4: 1. 기본정보 ==========
    ws.cell(row=4, column=1, value="1. 기본정보").font = header_font
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=11)

    # ========== Row 5~6: 차 종(번호) + 업무사용비율 ==========
    ws.cell(row=5, column=1, value="차 종(번호)").font = bold
    ws.cell(row=5, column=1).alignment = center
    ws.cell(row=5, column=1).border = box
    ws.cell(row=5, column=1).fill = light_gray
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)

    ws.cell(row=5, column=3, value=vehicle).alignment = center
    ws.cell(row=5, column=3).border = box
    ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=8)

    ws.cell(row=5, column=9, value="업무사용비율").font = bold
    ws.cell(row=5, column=9).alignment = center
    ws.cell(row=5, column=9).border = box
    ws.cell(row=5, column=9).fill = light_gray

    ws.cell(row=5, column=10, value="100%").alignment = center
    ws.cell(row=5, column=10).border = box
    ws.merge_cells(start_row=5, start_column=10, end_row=5, end_column=11)

    # ========== Row 6: 2. 업무용 사용비율 계산 ==========
    ws.cell(row=6, column=1, value="2. 업무용 사용비율 계산").font = header_font
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=11)

    # ========== Row 7~8: 헤더 (병합) ==========
    # 사용일자 | 사용자(부서/성명) | 운행내역(주행전/후/거리/주유금액/목적지(출발지/도착지)/사용목적) | 비고
    headers_top = [
        ("사용일자", 1, 1),  # A7~A8
        ("사용자",   2, 3),  # B7~C7 (B8=부서, C8=성명)
        ("운행내역", 4, 11), # D7~K7
    ]
    for label, c1, c2 in headers_top:
        ws.cell(row=7, column=c1, value=label).font = bold
        ws.cell(row=7, column=c1).alignment = center
        if c1 == c2 == 1:
            ws.merge_cells(start_row=7, start_column=1, end_row=8, end_column=1)
        else:
            ws.merge_cells(start_row=7, start_column=c1, end_row=7, end_column=c2)

    # row 8 sub-headers
    sub = [
        (2, "부서"),
        (3, "성명"),
        (4, "주행 전\n계기판 km"),
        (5, "주행 후\n계기판 km"),
        (6, "주행거리\n(km)"),
        (7, "주유금액\n(원)"),
        (8, "출발지"),
        (9, "도착지"),
        (10, "사용목적"),
        (11, "비고"),
    ]
    for col, label in sub:
        c = ws.cell(row=8, column=col, value=label)
        c.font = bold
        c.alignment = center

    # 헤더 셀 스타일 일괄 적용 (테두리 + 배경)
    for r in (7, 8):
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.border = box
            cell.fill = light_gray
    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 32

    # ========== Row 9~: 데이터 ==========
    total_distance = 0
    total_fuel = 0
    row = 9
    for log in logs:
        use_date_str = log.use_date.strftime('%Y-%m-%d') if log.use_date else ''
        values = [
            use_date_str,
            log.user_department or '',
            log.user_name or '',
            log.odometer_start if log.odometer_start is not None else '',
            log.odometer_end,
            log.distance_km,
            log.fuel_amount or 0,
            log.origin or '',
            log.destination or '',
            log.purpose or '',
            '',  # 비고
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = box
            c.font = Font(size=9)
            if col in (1, 2, 3):
                c.alignment = center
            elif col in (4, 5, 6, 7):
                c.alignment = right
                if isinstance(val, (int, float)):
                    c.number_format = '#,##0'
            else:
                c.alignment = left

        total_distance += log.distance_km or 0
        total_fuel += log.fuel_amount or 0
        row += 1

    # 빈 줄을 일정 행수까지 채워서 양식 모양 유지 (최소 20행)
    min_data_rows = 20
    while row < 9 + min_data_rows:
        for col in range(1, 12):
            c = ws.cell(row=row, column=col, value='')
            c.border = box
        row += 1

    # ========== 합계 행 ==========
    ws.cell(row=row, column=1, value="합  계").font = bold
    ws.cell(row=row, column=1).alignment = center
    ws.cell(row=row, column=1).fill = light_gray
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    ws.cell(row=row, column=6, value=total_distance).font = bold
    ws.cell(row=row, column=6).alignment = right
    ws.cell(row=row, column=6).number_format = '#,##0'
    ws.cell(row=row, column=6).fill = light_gray

    ws.cell(row=row, column=7, value=total_fuel).font = bold
    ws.cell(row=row, column=7).alignment = right
    ws.cell(row=row, column=7).number_format = '#,##0'
    ws.cell(row=row, column=7).fill = light_gray

    ws.cell(row=row, column=8, value="").fill = light_gray
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=11)

    for col in range(1, 12):
        ws.cell(row=row, column=col).border = box

    # 페이지 설정 (가로방향, A4)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def make_filename(vehicle: str, year: int) -> str:
    """다운로드 파일명 — 차량명·연도 포함, 영문/숫자 안전"""
    safe_vehicle = vehicle.replace(' ', '_').replace('/', '_')
    return f"운행기록부_{safe_vehicle}_{year}.xlsx"
