"""
재고관리 비즈니스 로직 (Inventory Actions).

routes/inventory.py에서 분리한 DB 쿼리·집계·엑셀 생성 등.
inventory_utils.py(record_stock_movement, confirm_audit, calc_turnover_rate)와 별도.
"""

import datetime
import io
import logging

from sqlalchemy import desc, func, distinct, or_

from modules.models import (
    Item, StockAudit, StockAuditItem, StockMovement,
    MOVEMENT_TYPE_LABELS, BomHeader, BomItem,
)

logger = logging.getLogger(__name__)


# ── helpers ──────────────────────────────────────────────────────
def fmt_money(val):
    """금액 포맷 (천단위 콤마)."""
    if val is None:
        return '0'
    try:
        return f"{int(val):,}"
    except (ValueError, TypeError):
        return str(val)


def generate_audit_no(db):
    """실사번호 자동 채번: SA{YYYY}-{NNN}"""
    year = datetime.date.today().year
    prefix = f'SA{year}-'
    last = db.query(StockAudit).filter(
        StockAudit.audit_no.like(f'{prefix}%')
    ).order_by(desc(StockAudit.audit_no)).first()
    if last:
        try:
            last_num = int(last.audit_no.replace(prefix, ''))
            return f'{prefix}{last_num + 1:03d}'
        except ValueError:
            pass
    return f'{prefix}001'


# ── 1. 대시보드 집계 ─────────────────────────────────────────────
def get_dashboard_data(db):
    """대시보드에 필요한 집계 데이터를 dict로 반환."""
    items = db.query(Item).filter(Item.is_active == True).all()

    total_items = len(items)
    total_stock_value = 0
    total_available_value = 0
    total_reserved_value = 0
    low_stock_count = 0
    low_stock_items = []
    cat_summary = {}

    for item in items:
        stock = item.stock_qty or 0
        reserved = item.reserved_qty or 0
        available = max(0, stock - reserved)
        price = item.last_unit_price or 0
        stock_val = stock * price
        available_val = available * price
        reserved_val = reserved * price

        total_stock_value += stock_val
        total_available_value += available_val
        total_reserved_value += reserved_val

        safety = item.safety_stock or 0
        if safety > 0 and stock < safety:
            low_stock_count += 1
            low_stock_items.append({
                'id': item.id,
                'name': item.item_name,
                'code': item.icube_item_cd or '',
                'stock_qty': stock,
                'safety_stock': safety,
                'shortage': safety - stock,
                'category': item.category or '',
            })

        cat = item.category or '미분류'
        if cat not in cat_summary:
            cat_summary[cat] = {
                'category': cat,
                'count': 0, 'total_stock': 0, 'reserved': 0,
                'available': 0, 'value': 0,
            }
        cat_summary[cat]['count'] += 1
        cat_summary[cat]['total_stock'] += stock
        cat_summary[cat]['reserved'] += reserved
        cat_summary[cat]['available'] += available
        cat_summary[cat]['value'] += stock_val

    recent_movements = db.query(StockMovement).order_by(
        desc(StockMovement.created_at)
    ).limit(10).all()

    low_stock_items.sort(key=lambda x: x['shortage'], reverse=True)
    low_stock_items = low_stock_items[:10]

    cat_list = sorted(cat_summary.values(), key=lambda x: x['value'], reverse=True)

    return {
        'total_items': total_items,
        'total_stock_value': total_stock_value,
        'total_available_value': total_available_value,
        'total_reserved_value': total_reserved_value,
        'low_stock_count': low_stock_count,
        'low_stock_items': low_stock_items,
        'cat_summary': cat_list,
        'recent_movements': recent_movements,
    }


# ── 1-1. BOM 기준 가용재고 ───────────────────────────────────────
def get_bom_stock_data(db, q, bom_id):
    """BOM 선택 → 소요부품별 가용재고 + 생산가능수량."""
    from modules.utils import safe_int

    bom_list = db.query(BomHeader).filter(BomHeader.is_active == True)\
        .order_by(BomHeader.product_name).all()

    selected_bom = None
    bom_items_data = []
    producible_qty = None

    if bom_id:
        selected_bom = db.query(BomHeader).get(bom_id)
    elif q:
        selected_bom = db.query(BomHeader).filter(
            or_(BomHeader.product_name.ilike(f'%{q}%'),
                BomHeader.product_code.ilike(f'%{q}%'))
        ).first()

    if selected_bom:
        items = db.query(BomItem).filter(BomItem.bom_id == selected_bom.id)\
            .order_by(BomItem.id).all()

        min_producible = None
        for bi in items:
            item = bi.item if bi.item_id else None
            if not item and bi.item_code:
                item = db.query(Item).filter(Item.icube_item_cd == bi.item_code).first()

            stock_qty = (item.stock_qty or 0) if item else 0
            reserved_qty = (item.reserved_qty or 0) if item else 0
            available = stock_qty - reserved_qty
            needed = bi.quantity or 0
            can_make = int(available / needed) if needed > 0 else 999999

            if needed > 0:
                if min_producible is None or can_make < min_producible:
                    min_producible = can_make

            bom_items_data.append({
                'bom_item': bi,
                'item': item,
                'stock_qty': stock_qty,
                'reserved_qty': reserved_qty,
                'available': available,
                'needed': needed,
                'enough': available >= needed,
                'can_make': can_make,
            })

        producible_qty = min_producible if min_producible is not None else 0

    return {
        'bom_list': bom_list,
        'selected_bom': selected_bom,
        'bom_items': bom_items_data,
        'producible_qty': producible_qty,
    }


def search_bom_headers(db, q):
    """BOM 제품명 자동완성 결과 리스트."""
    results = db.query(BomHeader).filter(
        BomHeader.is_active == True,
        or_(BomHeader.product_name.ilike(f'%{q}%'),
            BomHeader.product_code.ilike(f'%{q}%'))
    ).limit(15).all()
    return [{
        'id': r.id,
        'product_code': r.product_code,
        'product_name': r.product_name,
        'product_category': r.product_category or '',
    } for r in results]


# ── 2. 가용재고 목록 필터링 ──────────────────────────────────────
def get_filtered_items(db, q, category, stock_status):
    """품목 목록을 필터링하여 (filtered_items, categories) 반환."""
    query = db.query(Item).filter(Item.is_active == True)

    if q:
        like = f'%{q}%'
        query = query.filter(or_(
            Item.icube_item_cd.ilike(like),
            Item.item_name.ilike(like),
            Item.item_spec.ilike(like),
        ))
    if category:
        query = query.filter(Item.category == category)

    all_items = query.order_by(Item.item_name).all()

    filtered = []
    for item in all_items:
        stock = item.stock_qty or 0
        safety = item.safety_stock or 0
        if stock_status == 'low' and (safety <= 0 or stock >= safety):
            continue
        if stock_status == 'excess' and stock <= safety * 2:
            continue
        if stock_status == 'normal' and safety > 0 and stock < safety:
            continue
        filtered.append(item)

    categories = [r[0] for r in db.query(distinct(Item.category)).filter(
        Item.category.isnot(None), Item.category != ''
    ).order_by(Item.category).all()]

    return filtered, categories


# ── 4. 실사 생성 ────────────────────────────────────────────────
def create_audit(db, audit_date_str, auditor_name, note, target_category, auditor_id):
    """실사를 생성하고 대상 품목을 로딩. (audit, count) 반환."""
    try:
        audit_date = datetime.datetime.strptime(audit_date_str, '%Y-%m-%d').date()
    except ValueError:
        audit_date = datetime.date.today()

    audit = StockAudit(
        audit_no=generate_audit_no(db),
        audit_date=audit_date,
        auditor_id=auditor_id,
        auditor_name=auditor_name,
        status='진행중',
        note=note,
    )
    db.add(audit)
    db.flush()

    item_query = db.query(Item).filter(Item.is_active == True)
    if target_category:
        item_query = item_query.filter(Item.category == target_category)
    target_items = item_query.order_by(Item.item_name).all()

    for item in target_items:
        audit_item = StockAuditItem(
            audit_id=audit.id,
            item_id=item.id,
            system_qty=item.stock_qty or 0,
        )
        db.add(audit_item)

    audit.total_items = len(target_items)
    db.commit()

    return audit, len(target_items)


def get_audit_categories(db):
    """실사 생성 화면용 카테고리 목록."""
    return [r[0] for r in db.query(distinct(Item.category)).filter(
        Item.category.isnot(None), Item.category != ''
    ).order_by(Item.category).all()]


# ── 5. 실사 상세 ────────────────────────────────────────────────
def get_audit_detail(db, audit_id):
    """실사 상세 + 품목 맵 반환. (audit, audit_items, items_map)"""
    audit = db.query(StockAudit).get(audit_id)
    if not audit:
        return None, [], {}

    audit_items = db.query(StockAuditItem).filter(
        StockAuditItem.audit_id == audit_id,
    ).all()

    item_ids = [ai.item_id for ai in audit_items]
    items_map = {}
    if item_ids:
        for item in db.query(Item).filter(Item.id.in_(item_ids)).all():
            items_map[item.id] = item

    return audit, audit_items, items_map


# ── 6. 실사 수량 저장 ───────────────────────────────────────────
def save_audit_quantities(db, audit_id, form_data):
    """실사 항목에 실재고 수량을 저장. 저장 건수 반환."""
    audit_items = db.query(StockAuditItem).filter(
        StockAuditItem.audit_id == audit_id,
    ).all()

    saved_count = 0
    for ai in audit_items:
        actual_key = f'actual_{ai.id}'
        reason_key = f'reason_{ai.id}'
        actual_val = form_data.get(actual_key, '').strip()
        reason_val = (form_data.get(reason_key) or '').strip()

        if actual_val != '':
            try:
                ai.actual_qty = float(actual_val)
                ai.diff_qty = ai.actual_qty - (ai.system_qty or 0)
                saved_count += 1
            except (ValueError, TypeError):
                pass
        ai.diff_reason = reason_val

    db.commit()
    return saved_count


# ── 7-1. 실사 엑셀 템플릿 ───────────────────────────────────────
def build_audit_template_excel(audit, audit_items, items_map):
    """실사 템플릿 엑셀을 BytesIO로 반환."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '재고실사'

    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    yellow_fill = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')

    ws.merge_cells('A1:G1')
    ws['A1'] = f'재고실사 [{audit.audit_no}] — {audit.audit_date}'
    ws['A1'].font = Font(bold=True, size=12)

    headers = ['품목ID', '품번', '품명', '규격', '카테고리', '시스템재고', '실재고(입력)', '비고']
    col_widths = [8, 15, 30, 20, 12, 12, 14, 20]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 8 else 'H'].width = w

    for row_idx, ai in enumerate(audit_items, 4):
        item = items_map.get(ai.item_id)
        if not item:
            continue

        row_data = [
            item.id,
            item.icube_item_cd or '',
            item.item_name or '',
            item.item_spec or '',
            item.category or '',
            ai.system_qty or 0,
            ai.actual_qty if ai.actual_qty is not None else '',
            '',
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 7:
                cell.fill = yellow_fill
                cell.font = Font(bold=True)
            if col_idx == 6:
                cell.alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 7-2. 실사 엑셀 업로드 파싱 ──────────────────────────────────
def parse_audit_upload(db, audit_id, file_obj):
    """업로드된 엑셀 파일을 파싱하여 실사 항목에 반영. (updated, errors) 반환."""
    from openpyxl import load_workbook

    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    audit_items = db.query(StockAuditItem).filter(
        StockAuditItem.audit_id == audit_id
    ).all()
    ai_map = {ai.item_id: ai for ai in audit_items}

    updated = 0
    errors = []
    for row in ws.iter_rows(min_row=4, values_only=False):
        try:
            item_id_val = row[0].value
            actual_val = row[6].value
            note_val = row[7].value if len(row) > 7 else ''

            if item_id_val is None or actual_val is None:
                continue
            item_id = int(item_id_val)
            actual_qty = float(actual_val)

            ai = ai_map.get(item_id)
            if ai:
                ai.actual_qty = actual_qty
                ai.diff_qty = actual_qty - (ai.system_qty or 0)
                if note_val:
                    ai.note = str(note_val).strip()
                updated += 1
        except (ValueError, TypeError, IndexError):
            row_num = row[0].row if row[0] else '?'
            errors.append(f'{row_num}행')

    db.commit()
    return updated, errors


# ── 7-3. 실사 차이 보고서 통계 ───────────────────────────────────
def calc_audit_report_stats(audit_items, items_map):
    """실사 보고서 통계를 dict로 반환."""
    total_items = len(audit_items)
    counted_items = sum(1 for ai in audit_items if ai.actual_qty is not None)
    diff_items = [ai for ai in audit_items
                  if ai.actual_qty is not None and ai.diff_qty and abs(ai.diff_qty) > 0.001]
    match_items = counted_items - len(diff_items)

    total_system_value = 0
    total_actual_value = 0
    total_diff_value = 0
    for ai in audit_items:
        item = items_map.get(ai.item_id)
        price = (item.last_unit_price or 0) if item else 0
        total_system_value += (ai.system_qty or 0) * price
        if ai.actual_qty is not None:
            total_actual_value += ai.actual_qty * price
            total_diff_value += (ai.diff_qty or 0) * price

    return {
        'total_items': total_items,
        'counted_items': counted_items,
        'diff_items_list': diff_items,
        'diff_items': len(diff_items),
        'match_items': match_items,
        'total_system_value': total_system_value,
        'total_actual_value': total_actual_value,
        'total_diff_value': total_diff_value,
    }


def build_audit_report_excel(audit, audit_items, items_map, stats):
    """실사 차이 보고서 엑셀을 BytesIO로 반환."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '실사보고서'

    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    red_font = Font(color='DC3545', bold=True)
    blue_font = Font(color='0D6EFD', bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:I1')
    ws['A1'] = f'재고실사 차이 보고서 [{audit.audit_no}]'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A2'] = (f'실사일: {audit.audit_date} | 전체: {stats["total_items"]}건 | '
                f'실사완료: {stats["counted_items"]}건 | 차이: {stats["diff_items"]}건 | '
                f'일치: {stats["match_items"]}건')
    ws['A3'] = (f'시스템금액: {int(stats["total_system_value"]):,}원 | '
                f'실재고금액: {int(stats["total_actual_value"]):,}원 | '
                f'차이금액: {int(stats["total_diff_value"]):,}원')

    headers = ['품번', '품명', '규격', '카테고리', '시스템재고', '실재고', '차이', '단가', '차이금액', '비고']
    col_widths = [14, 28, 18, 10, 10, 10, 10, 12, 14, 18]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'J'].width = w

    for row_idx, ai in enumerate(audit_items, 6):
        item = items_map.get(ai.item_id)
        if not item:
            continue
        price = item.last_unit_price or 0
        diff = ai.diff_qty or 0
        diff_value = diff * price

        row_data = [
            item.icube_item_cd or '', item.item_name or '', item.item_spec or '',
            item.category or '', ai.system_qty or 0,
            ai.actual_qty if ai.actual_qty is not None else '-',
            diff, price, diff_value, ai.note or '',
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 7 and isinstance(val, (int, float)):
                if val > 0:
                    cell.font = blue_font
                elif val < 0:
                    cell.font = red_font
            if col_idx in (5, 6, 7, 8, 9):
                cell.alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 12. 재고현황 엑셀 ───────────────────────────────────────────
def build_inventory_export_excel(db):
    """재고현황 엑셀을 BytesIO로 반환."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    items = db.query(Item).filter(Item.is_active == True)\
        .order_by(Item.category, Item.item_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '재고현황'

    headers = ['품번', '품명', '규격', '분류', '총재고', '예약', '가용재고', '안전재고', '단가', '재고금액', '상태']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for r, item in enumerate(items, 2):
        stock = item.stock_qty or 0
        reserved = item.reserved_qty or 0
        available = max(0, stock - reserved)
        safety = item.safety_stock or 0
        price = item.last_unit_price or 0
        value = stock * price

        status = '정상'
        if safety > 0 and stock < safety:
            status = '부족'

        row_data = [
            item.icube_item_cd or '',
            item.item_name or '',
            item.item_spec or '',
            item.category or '',
            stock, reserved, available, safety,
            price, value, status,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = thin_border
            cell.font = Font(size=9)
            if col >= 5:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

    widths = [12, 30, 20, 10, 10, 10, 10, 10, 12, 15, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = f'A1:K{len(items) + 1}'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── API 헬퍼 ─────────────────────────────────────────────────────
def get_inventory_summary(db):
    """대시보드 요약 통계 dict."""
    items = db.query(Item).filter(Item.is_active == True).all()
    total_value = sum((i.stock_qty or 0) * (i.last_unit_price or 0) for i in items)
    low_count = sum(1 for i in items
                    if (i.safety_stock or 0) > 0 and (i.stock_qty or 0) < (i.safety_stock or 0))
    return {
        'total_items': len(items),
        'total_value': total_value,
        'low_stock_count': low_count,
    }


def get_low_stock_items(db, limit=20):
    """저재고 품목 리스트."""
    items = db.query(Item).filter(
        Item.is_active == True,
        Item.safety_stock > 0,
    ).all()

    low = []
    for item in items:
        if (item.stock_qty or 0) < (item.safety_stock or 0):
            low.append({
                'id': item.id,
                'name': item.item_name,
                'code': item.icube_item_cd or '',
                'stock_qty': item.stock_qty or 0,
                'safety_stock': item.safety_stock or 0,
                'shortage': (item.safety_stock or 0) - (item.stock_qty or 0),
            })

    low.sort(key=lambda x: x['shortage'], reverse=True)
    return low[:limit]


# ── 8. 재고 변동 이력 쿼리 ───────────────────────────────────────
def query_movements(db, q, m_type, date_from, date_to):
    """변동 이력 쿼리를 필터링하여 반환 (pagination 전)."""
    query = db.query(StockMovement)

    if q:
        item_ids = [r[0] for r in db.query(Item.id).filter(
            or_(Item.item_name.ilike(f'%{q}%'), Item.icube_item_cd.ilike(f'%{q}%'))
        ).all()]
        if item_ids:
            query = query.filter(StockMovement.item_id.in_(item_ids))
        else:
            query = query.filter(StockMovement.id < 0)

    if m_type:
        query = query.filter(StockMovement.movement_type == m_type)

    if date_from:
        try:
            import datetime as _dt
            dt_from = _dt.datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(StockMovement.created_at >= dt_from)
        except ValueError:
            pass
    if date_to:
        try:
            import datetime as _dt
            dt_to = _dt.datetime.strptime(date_to, '%Y-%m-%d') + _dt.timedelta(days=1)
            query = query.filter(StockMovement.created_at < dt_to)
        except ValueError:
            pass

    return query


def load_items_map(db, item_ids):
    """item_id 리스트 → {id: Item} 매핑."""
    items_map = {}
    if item_ids:
        for item in db.query(Item).filter(Item.id.in_(item_ids)).all():
            items_map[item.id] = item
    return items_map


def get_category_list(db):
    """활성 품목의 카테고리 목록."""
    return [r[0] for r in db.query(distinct(Item.category)).filter(
        Item.category.isnot(None), Item.category != ''
    ).order_by(Item.category).all()]
