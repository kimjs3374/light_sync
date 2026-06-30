"""출장관리 목록 → 엑셀(xlsx) 내보내기"""
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# (헤더, 컬럼 너비)
COLUMNS = [
    ('번호', 6),
    ('상태', 9),
    ('출장제목', 28),
    ('출장장소', 24),
    ('출발일시', 18),
    ('복귀일시', 18),
    ('출장인원', 30),
    ('이동수단', 14),
    ('출장목적', 36),
    ('비고', 24),
    ('등록자', 12),
    ('등록일', 18),
]


def _member_text(trip):
    """출장인원을 '이름 직급(부서)' 형태로 전체 표기"""
    parts = []
    for m in trip.members:
        seg = m.user_name or ''
        if m.position:
            seg += f' {m.position}'
        if m.department:
            seg += f'({m.department})'
        if seg.strip():
            parts.append(seg.strip())
    return ', '.join(parts)


def _fmt_dt(dt):
    return dt.strftime('%Y-%m-%d %H:%M') if dt else ''


def export_business_trips_excel(trips) -> BytesIO:
    """출장 목록(trips) → 엑셀 BytesIO 반환"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '출장관리'

    thin = Side(style='thin', color='000000')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='E9ECEF')
    header_font = Font(bold=True, size=10)
    body_font = Font(size=9)
    wrap_top = Alignment(vertical='top', wrap_text=True)
    center = Alignment(horizontal='center', vertical='center')

    # 컬럼 너비
    for i, (_, w) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 헤더
    for col, (label, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        c.border = box
        c.alignment = center

    # 데이터
    row = 2
    for idx, t in enumerate(trips, 1):
        values = [
            idx,
            t.effective_status,
            t.title or '',
            t.destination or '',
            _fmt_dt(t.departure_date),
            _fmt_dt(t.return_date),
            _member_text(t),
            t.vehicle or '',
            t.purpose or '',
            t.note or '',
            t.creator.full_name if t.creator else '',
            _fmt_dt(t.created_at),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = box
            c.font = body_font
            if col in (1, 2, 5, 6, 8, 11, 12):
                c.alignment = center
            else:
                c.alignment = wrap_top
        row += 1

    ws.freeze_panes = 'A2'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def make_filename(today):
    """다운로드 파일명: 출장관리_YYYYMMDD.xlsx"""
    return f"출장관리_{today.strftime('%Y%m%d')}.xlsx"
