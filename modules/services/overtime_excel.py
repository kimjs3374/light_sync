"""특근대장(전자결재 overtime 양식) 엑셀 생성

결재완료된 특근대장 문서 1건 → 회계/노무 제출용 엑셀 1장.
- 상단: 제목(대상월) + 결재란(기안/결재자 도장란) + 기본정보
- 본표: 순번/근무일자/요일/구분/시작/종료/휴게/특근시간/업무내용
- 하단: 합계(총 시간·일수) + 비고

파일명: 부서_이름직급_x월_특근대장.xlsx
"""
from io import BytesIO
import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# A~L. 업무내용(I~L)은 데이터 행에서 병합해 한 칸으로 쓰고,
# 상단 결재란은 폭이 고른 칸으로 나누기 위해 컬럼을 잘게 유지한다.
NCOL = 12
CONTENT_COL = 9
COL_WIDTHS = [6, 13, 6, 8, 8, 8, 9, 11, 10, 10, 9, 9]
WEEKDAYS = ['월', '화', '수', '목', '금', '토', '일']

STEP_MARK = {
    'approved': '승인',
    'rejected': '반려',
    'current': '결재중',
    'waiting': '대기',
}


def _parse_date(v):
    """'YYYY-MM-DD' → date. 실패 시 None."""
    if isinstance(v, datetime.date):
        return v
    try:
        return datetime.datetime.strptime(str(v).strip()[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _to_float(v):
    """'4.5' / 4.5 / '' → float. 실패 시 0.0"""
    try:
        return float(str(v).strip())
    except (ValueError, TypeError, AttributeError):
        return 0.0


def _month_of(doc):
    """대상월 (year, month). work_month 우선, 없으면 첫 내역일자 → 기안일."""
    fd = doc.form_data or {}
    raw = (fd.get('work_month') or '').strip()
    if raw:
        try:
            y, m = raw.split('-')[:2]
            return int(y), int(m)
        except (ValueError, IndexError):
            pass
    for e in (fd.get('entries') or []):
        d = _parse_date(e.get('work_date'))
        if d:
            return d.year, d.month
    base = doc.submitted_at or doc.created_at or datetime.datetime.now()
    return base.year, base.month


def _sorted_entries(doc):
    """내역 리스트를 근무일자 오름차순으로 정렬 (일자 없는 행은 뒤로)."""
    entries = (doc.form_data or {}).get('entries') or []
    entries = [e for e in entries if isinstance(e, dict)]

    def key(e):
        d = _parse_date(e.get('work_date'))
        return (0, d) if d else (1, datetime.date.max)

    return sorted(entries, key=key)


def _split_columns(widths, n):
    """컬럼을 n개 셀로 연속 분할 — 셀 폭이 고르도록 너비 기준 분배.

    widths: 컬럼 너비 리스트(1번 컬럼부터), n: 셀 개수
    → [(start_col, end_col), ...]
    """
    total = len(widths)
    n = max(1, min(n, total))
    target = sum(widths) / n
    spans, col = [], 1
    for i in range(n):
        remaining_cells = n - i - 1
        acc = 0
        end = col
        while end <= total - remaining_cells:
            nxt = acc + widths[end - 1]
            # 남은 컬럼을 남은 셀에 최소 1개씩은 남겨야 함
            if acc and abs(nxt - target) > abs(acc - target):
                break
            acc = nxt
            end += 1
        spans.append((col, end - 1))
        col = end
    # 남은 컬럼은 마지막 셀에 흡수
    if spans and spans[-1][1] < total:
        spans[-1] = (spans[-1][0], total)
    return spans


def export_overtime_excel(doc) -> BytesIO:
    """특근대장 ApprovalDocument → 엑셀 BytesIO 반환"""
    year, month = _month_of(doc)
    entries = _sorted_entries(doc)
    fd = doc.form_data or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년 {month}월 특근대장"

    thin = Side(style='thin', color='000000')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True, size=10)
    title_font = Font(bold=True, size=16)
    small = Font(size=9)
    light_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    right = Alignment(horizontal='right', vertical='center', indent=1)

    # 컬럼 (A=순번 B=근무일자 C=요일 D=구분 E=시작 F=종료 G=휴게 H=특근시간 I~L=업무내용)
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def put(row, col, value, *, font=None, align=None, fill=None, border=True,
            span=None, number_format=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font or small
        cell.alignment = align or center
        if fill:
            cell.fill = fill
        if number_format:
            cell.number_format = number_format
        if span:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span)
        if border:
            for c in range(col, (span or col) + 1):
                ws.cell(row=row, column=c).border = box
                if fill:
                    ws.cell(row=row, column=c).fill = fill
        return cell

    # ========== Row 1: 제목 ==========
    put(1, 1, f"{year}년 {month}월 특 근 대 장", font=title_font, span=NCOL, border=False)
    ws.row_dimensions[1].height = 32

    # ========== Row 2~4: 결재란 (기안 + 결재선) ==========
    cells = [('기안', doc.drafter_name, doc.drafter_position or '', '기안',
              doc.submitted_at or doc.created_at)]
    for s in doc.steps:
        cells.append((s.role_label, s.approver_name, s.approver_position or '',
                      STEP_MARK.get(s.status, s.status), s.acted_at))

    for (c1, c2), (hd, name, pos, mark, acted) in zip(_split_columns(COL_WIDTHS, len(cells)), cells):
        put(2, c1, hd, font=bold, fill=light_gray, span=c2)
        put(3, c1, f"{name} {pos}".strip(), font=Font(bold=True, size=11), span=c2)
        stamp = mark if not acted else f"{mark}  {acted.strftime('%Y-%m-%d')}"
        put(4, c1, stamp, span=c2)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 18

    ws.row_dimensions[5].height = 8  # 여백

    # ========== Row 6~7: 기본정보 ==========
    completed = doc.completed_at
    status_txt = doc.status_label + (f"  ({completed.strftime('%Y-%m-%d')})" if completed else '')

    info = [
        (6, "부서 / 성명", f"{doc.drafter_dept or '-'}  {doc.drafter_name} {doc.drafter_position or ''}".strip(),
            "대상월", f"{year}년 {month}월"),
        (7, "문서번호", doc.doc_no or '임시저장',
            "결재상태", status_txt),
    ]
    for row, l1, v1, l2, v2 in info:
        put(row, 1, l1, font=bold, fill=light_gray, span=2)
        put(row, 3, v1, align=left, span=6)
        put(row, 7, l2, font=bold, fill=light_gray, span=8)
        put(row, 9, v2, align=left, span=NCOL)
        ws.row_dimensions[row].height = 20

    ws.row_dimensions[8].height = 8  # 여백

    # ========== Row 9: 본표 헤더 ==========
    HEAD = 9
    headers = ['순번', '근무일자', '요일', '구분', '시작', '종료', '휴게(h)', '특근시간(h)']
    for col, label in enumerate(headers, 1):
        put(HEAD, col, label, font=bold, fill=light_gray)
    put(HEAD, CONTENT_COL, '업무내용', font=bold, fill=light_gray, span=NCOL)
    ws.row_dimensions[HEAD].height = 24

    # ========== Row 10~: 내역 ==========
    row = HEAD + 1
    total_hours = 0.0
    total_break = 0.0
    for idx, e in enumerate(entries, 1):
        d = _parse_date(e.get('work_date'))
        hours = _to_float(e.get('hours'))
        brk = _to_float(e.get('break_hours'))
        total_hours += hours
        total_break += brk

        put(row, 1, idx)
        put(row, 2, d.strftime('%Y-%m-%d') if d else (e.get('work_date') or ''))
        put(row, 3, WEEKDAYS[d.weekday()] if d else '')
        put(row, 4, e.get('ot_type') or '')
        put(row, 5, e.get('start_time') or '')
        put(row, 6, e.get('end_time') or '')
        put(row, 7, brk if e.get('break_hours') not in (None, '') else '',
            align=right, number_format='0.##')
        put(row, 8, hours, align=right, number_format='0.##')
        put(row, CONTENT_COL, e.get('work_content') or '', align=left, span=NCOL)
        ws.row_dimensions[row].height = 20
        row += 1

    # 내역이 없어도 표 모양은 유지
    if not entries:
        put(row, 1, '내역 없음', span=NCOL)
        row += 1

    # ========== 합계 ==========
    put(row, 1, f"합  계   (총 {len(entries)}일)", font=bold, fill=light_gray, align=right, span=6)
    put(row, 7, total_break if total_break else '', font=bold, fill=light_gray,
        align=right, number_format='0.##')
    put(row, 8, total_hours, font=bold, fill=light_gray, align=right, number_format='0.##')
    put(row, CONTENT_COL, '', fill=light_gray, span=NCOL)
    ws.row_dimensions[row].height = 22
    row += 1

    # ========== 비고 (양식 비고 + 본문) ==========
    memo_parts = [str(fd.get('reason') or '').strip(), str(doc.content or '').strip()]
    memo = '\n'.join(p for p in memo_parts if p).replace('\r\n', '\n').replace('\r', '\n')
    if memo:
        put(row, 1, '비고', font=bold, fill=light_gray, span=2)
        put(row, 3, memo, align=left, span=NCOL)
        # 병합 셀은 자동 행높이가 안 되므로 줄바꿈 + 줄넘김을 감안해 직접 계산
        per_line = int(sum(COL_WIDTHS[2:]))
        lines = sum(max(1, -(-len(ln) * 2 // per_line)) for ln in memo.split('\n'))
        ws.row_dimensions[row].height = max(20, 14 * lines + 6)
        row += 1

    # ========== 출력일 ==========
    row += 1
    put(row, 1, f"출력일: {datetime.date.today().strftime('%Y-%m-%d')}",
        font=Font(size=8), align=right, span=NCOL, border=False)

    # 페이지 설정 (가로방향 A4, 폭 맞춤, 헤더행 반복)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f'{HEAD}:{HEAD}'
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.freeze_panes = ws.cell(row=HEAD + 1, column=1)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def make_filename(doc) -> str:
    """다운로드 파일명 — 부서_이름직급_x월_특근대장.xlsx"""
    _, month = _month_of(doc)
    dept = (doc.drafter_dept or '').strip() or '부서미지정'
    who = f"{(doc.drafter_name or '').strip()}{(doc.drafter_position or '').strip()}"
    parts = [dept, who or '기안자', f"{month}월", "특근대장"]
    safe = ['_'.join(p.split()).replace('/', '_') for p in parts]
    return '_'.join(safe) + '.xlsx'
