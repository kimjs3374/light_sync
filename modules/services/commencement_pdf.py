"""
착수계 PDF 생성 모듈.

엑셀 템플릿 방식 — openpyxl로 환경설정 시트 채우기 → LibreOffice headless PDF 변환.
레이아웃 100% 원본 동일.
"""

import io
import os
import logging
import shutil
import subprocess
import tempfile
from datetime import date, datetime

logger = logging.getLogger(__name__)

# 엑셀 템플릿 경로
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'static', 'templates', 'commencement_template.xlsx')

# Supabase 공통 서류 경로
COMMON_DOCS_PREFIX = 'documents/common/'
DRAWINGS_PREFIX = 'documents/drawings/'

# 공통 서류 목록 (Supabase storage 파일명 — 영문)
COMMON_DOC_FILES = {
    'biz_registration': 'biz_registration.pdf',
    'factory_cert': 'factory_cert.pdf',
    'direct_production_light': 'direct_production_light.pdf',
    'direct_production_pole': 'direct_production_pole.pdf',
    'direct_production_tower': 'direct_production_tower.pdf',
}

# LibreOffice 경로 (OS별)
SOFFICE_PATHS = [
    r'C:\Program Files\LibreOffice\program\soffice.exe',      # Windows
    '/usr/bin/soffice',                                         # Linux
    '/usr/bin/libreoffice',                                     # Linux alt
    '/Applications/LibreOffice.app/Contents/MacOS/soffice',     # macOS
]


# ──────────────────────────────────────
# 한글 금액 변환
# ──────────────────────────────────────

def number_to_korean(num):
    """숫자를 한글 금액으로 변환 (예: 8832000 -> '팔백팔십삼만이천')."""
    if num == 0:
        return '영'
    num = int(num)
    digits = ['', '일', '이', '삼', '사', '오', '육', '칠', '팔', '구']
    small_units = ['', '십', '백', '천']
    big_units = ['', '만', '억', '조', '경']

    if num < 0:
        return '마이너스 ' + number_to_korean(-num)

    result = ''
    group_idx = 0
    while num > 0:
        group = num % 10000
        num //= 10000
        if group > 0:
            group_str = ''
            for i in range(4):
                d = group % 10
                group //= 10
                if d > 0:
                    if d == 1 and i > 0:
                        group_str = small_units[i] + group_str
                    else:
                        group_str = digits[d] + small_units[i] + group_str
            result = group_str + big_units[group_idx] + result
        group_idx += 1

    return result


def _find_soffice():
    """LibreOffice 실행 파일 경로를 찾는다."""
    for path in SOFFICE_PATHS:
        if os.path.exists(path):
            return path
    # PATH에서 검색
    import shutil as _shutil
    found = _shutil.which('soffice') or _shutil.which('libreoffice')
    return found


def generate_commencement_pdf(package, procurements, agent_user=None,
                               stamp_path=None, logo_path=None, official_stamp_path=None,
                               attachments=None):
    """
    착수계 PDF를 생성한다 (엑셀 템플릿 방식).

    Args:
        package: DocumentPackage 객체
        procurements: list of G2bProcurement 객체
        agent_user: User 객체 (현장대리인)
        stamp_path: 미사용 (엑셀에 포함)
        logo_path: 미사용 (엑셀에 포함)
        official_stamp_path: 미사용 (엑셀에 포함)
        attachments: list of DocumentAttachment 객체 (첨부파일)

    Returns:
        io.BytesIO: PDF 바이트 스트림
    """
    import openpyxl

    soffice = _find_soffice()
    if not soffice:
        raise RuntimeError("LibreOffice가 설치되어 있지 않습니다. sudo apt install libreoffice-calc")

    template = os.path.abspath(TEMPLATE_PATH)
    if not os.path.exists(template):
        raise FileNotFoundError(f"착수계 엑셀 템플릿이 없습니다: {template}")

    # 임시 파일에 엑셀 복사 후 데이터 채우기
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'commencement.xlsx')
        shutil.copy(template, xlsx_path)

        # 환경설정 시트 채우기
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb[wb.sheetnames[0]]  # 환경설정 (첫번째 시트)

        # 데이터 추출
        business_name = package.business_name or ''
        contract_no = package.contract_no or ''
        req_no = package.procurement_req_no or ''
        demand_org = package.demand_org or ''
        org_type = package.org_type or '청'
        commencement_date = package.commencement_date or date.today()
        fee = package.fee or 0

        # 계약일, 납품기한은 조달내역에서
        contract_date = package.contract_date
        delivery_due = None
        if procurements:
            if not contract_date:
                contract_date = procurements[0].pdf_contract_date or procurements[0].cntrct_dlvr_req_date
            delivery_due = procurements[0].dlvr_tmlmt_date

        # 품목 정보 (최대 3개) — 모델코드만 추출
        items = []
        for p in procurements:
            model = ''
            if p.prdct_idnt_no_nm:
                # "철제가로등주, 매그나텍, MTPF-201-5, 5m, 1등용, 원2단형" → MTPF-201-5
                parts = [x.strip() for x in p.prdct_idnt_no_nm.split(',')]
                for part in parts:
                    if part.startswith(('MT', 'ARENA', 'BATOO')):
                        model = part
                        break
                if not model and len(parts) >= 3:
                    model = parts[0]  # 첫번째 값 사용
            items.append({
                'model': model,
                'price': p.prdct_uprc or 0,
                'qty': p.prdct_qty or 0,
            })

        # 현장대리인 정보 — VLOOKUP 키 (환경설정 시트의 현장대리인계!K17:K19 참조용)
        agent_name = ''
        if agent_user:
            agent_name = agent_user.full_name or agent_user.username or ''

        # 공문번호 순번 추출 (관리 제 26-032601호 → 1)
        doc_no_seq = 1
        if package.commencement_doc_no:
            import re
            m = re.search(r'(\d+)호', package.commencement_doc_no)
            if m:
                seq_str = m.group(1)
                doc_no_seq = int(seq_str[-2:]) if len(seq_str) >= 2 else int(seq_str)

        # === 환경설정 시트 채우기 (DATA 셀만, FORMULA 셀 건드리지 않음) ===
        ws['B1'] = business_name                              # 사업명
        ws['B2'] = contract_no                                # 계약번호
        ws['D2'] = req_no                                     # 납품요구번호
        ws['B3'] = contract_date                              # 계약일자
        ws['D3'] = delivery_due                               # 납품기한
        ws['B4'] = commencement_date                          # 제출일자 (착수일)
        # D4는 FORMULA (공문번호 자동생성) — 건드리지 않음
        ws['B5'] = len(items)                                 # 물품종류 수
        ws['D5'] = doc_no_seq                                 # 공문번호 순번
        ws['B6'] = demand_org                                 # 발주처
        ws['D6'] = package.demand_org_no or ''                # 수요기관번호
        ws['B7'] = org_type                                   # 관청구분 (청/기관)
        ws['D7'] = '유'                                       # 감리유무
        ws['J1'] = org_type                                   # 관청구분 반복 (공문 참조용)
        ws['K1'] = '유'                                       # 감리유무 반복

        # 물품 1
        if len(items) >= 1:
            ws['B8'] = items[0]['model']                      # 모델명1
            ws['D8'] = items[0]['price']                      # 단가1
            ws['B9'] = items[0]['qty']                        # 수량1
            ws['D9'] = fee                                    # 수수료

        # 물품 2
        if len(items) >= 2:
            ws['G8'] = items[1]['model']
            ws['I8'] = items[1]['price']
            ws['G9'] = items[1]['qty']
        else:
            ws['G8'] = None
            ws['I8'] = 0
            ws['G9'] = 0

        # 물품 3
        if len(items) >= 3:
            ws['L8'] = items[2]['model']
            ws['N8'] = items[2]['price']
            ws['L9'] = items[2]['qty']
        else:
            ws['L8'] = None
            ws['N8'] = 0
            ws['L9'] = 0

        # 현장대리인 (VLOOKUP 키)
        ws['B10'] = agent_name

        # D10 합계금액 직접 계산 (수식 유지하되 값도 확인)
        total_supply = sum(it['price'] * it['qty'] for it in items)

        # [DBNum4] 포맷이 LibreOffice에서 안 먹히므로
        # 착수계 시트의 C10(한글금액)을 직접 텍스트로 덮어쓰기
        ws_chak = wb[wb.sheetnames[2]]  # 착수계
        korean_amt = f'일금 {number_to_korean(total_supply)} 원정'
        ws_chak['C10'] = korean_amt
        ws_chak['C10'].number_format = '@'  # 텍스트 포맷
        # G10(숫자금액)도 직접 설정
        ws_chak['G10'] = total_supply

        # 필요한 시트만 표시, 나머지 숨기기
        # 시트 인덱스: 0=환경설정, 1=공문, 2=착수계, 3=서류, 4~6=직생, 7=현장대리인계,
        #            8=재직증명서, 9=물품계약서, 10~12=공정표, 13=물품납품내역, 14=data
        # 품목 종류에 따라 공정표 시트 선택 (등기구/등주/타워)
        keep_indices = {1, 2, 7, 8, 13}  # 공문, 착수계, 현장대리인계, 재직증명서, 납품내역서

        # 공정표: 기본 등기구(10), 물품에 따라 등주(11) 또는 타워(12) 선택
        first_model = items[0]['model'] if items else ''
        if '타워' in first_model or 'MTT' in first_model:
            keep_indices.add(12)  # 공정표_타워
        elif '등주' in first_model or 'MTP' in first_model.upper():
            keep_indices.add(11)  # 공정표_등주
        else:
            keep_indices.add(10)  # 공정표_조명 (기본)

        wb.save(xlsx_path)
        wb.close()

        # Step 1: LibreOffice로 수식 계산 + 값으로 변환된 중간 파일 생성
        # macro로 수식→값 변환 후 저장하는 대신,
        # 먼저 전체 PDF 생성 후 필요한 페이지만 추출하는 방식 사용
        env = os.environ.copy()
        env['PATH'] = '/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin:' + env.get('PATH', '')
        result = subprocess.run([
            soffice, '--headless', '--calc', '--convert-to', 'pdf',
            '--outdir', tmpdir, xlsx_path
        ], capture_output=True, text=True, timeout=120, env=env)

        if result.returncode != 0:
            logger.error("LibreOffice 변환 실패: %s", result.stderr)
            raise RuntimeError(f"PDF 변환 실패: {result.stderr}")

        pdf_path = xlsx_path.replace('.xlsx', '.pdf')
        if not os.path.exists(pdf_path):
            raise RuntimeError("PDF 파일이 생성되지 않았습니다.")

        # ═══════════════════════════════════════════
        # 엑셀 PDF에서 필요한 페이지 추출 + 첨부파일 합치기
        # 붙임 순서:
        #   공문 → 착수계 → 사업자등록증 → 납세증명서 → 공장등록증명서
        #   → 직접생산확인증명서 → 현장대리인계 → 재직증명서
        #   → 물품계약서(납품요구서) → 예정공정표 → 납품내역서 → 제작도면
        # ═══════════════════════════════════════════
        import pypdf
        reader = pypdf.PdfReader(pdf_path)
        writer = pypdf.PdfWriter()

        # 엑셀 PDF 페이지 매핑 (고정 인덱스)
        P_LETTER = 4       # 공문
        P_BODY = 5         # 착수계
        P_AGENT = 19       # 현장대리인계
        P_CERT = 20        # 재직증명서
        P_SCHED_LIGHT = 22 # 공정표_조명
        P_SCHED_POLE = 23  # 공정표_등주
        P_SCHED_TOWER = 24 # 공정표_타워
        P_DELIVERY = 25    # 물품납품내역서

        def _add_excel_page(page_idx):
            if page_idx < len(reader.pages):
                writer.add_page(reader.pages[page_idx])

        def _add_pdf_bytes(pdf_bytes):
            """PDF 바이트를 페이지로 추가."""
            if not pdf_bytes:
                return False
            try:
                r = pypdf.PdfReader(io.BytesIO(pdf_bytes))
                for page in r.pages:
                    writer.add_page(page)
                return True
            except Exception as e:
                logger.warning("PDF 바이트 읽기 실패: %s", e)
                return False

        def _add_pdf_file(filepath):
            """로컬 PDF 파일 추가."""
            if not filepath or not os.path.exists(filepath):
                return False
            try:
                r = pypdf.PdfReader(filepath)
                for page in r.pages:
                    writer.add_page(page)
                return True
            except Exception as e:
                logger.warning("PDF 파일 읽기 실패 (%s): %s", filepath, e)
                return False

        def _add_supabase_pdf(storage_path):
            """Supabase storage에서 PDF 다운로드 후 추가."""
            from modules.storage_adapter import download_bytes as dl_bytes
            pdf_bytes = dl_bytes(storage_path)
            return _add_pdf_bytes(pdf_bytes)

        def _add_common_doc(doc_key):
            """공통 서류 (Supabase)."""
            filename = COMMON_DOC_FILES.get(doc_key)
            if filename:
                return _add_supabase_pdf(COMMON_DOCS_PREFIX + filename)
            return False

        def _add_attachment_pdf(file_type):
            """현장별 첨부파일 (Supabase 또는 로컬)."""
            if not attachments:
                return
            for att in attachments:
                if att.file_type == file_type and att.storage_path:
                    if att.storage_path.startswith('documents/'):
                        _add_supabase_pdf(att.storage_path)
                    else:
                        base = os.path.join(os.path.dirname(__file__), '..', '..')
                        _add_pdf_file(os.path.join(base, att.storage_path))

        def _add_drawing_by_model(model_code):
            """모델별 제작도면 (Supabase)."""
            if model_code:
                return _add_supabase_pdf(DRAWINGS_PREFIX + model_code + '.pdf')
            return False

        # 품목 모델 판별 (공정표/직생확인서 선택용)
        first_model_upper = (items[0]['model'] if items else '').upper()

        # 1. 공문
        _add_excel_page(P_LETTER)

        # 2. 착수계
        _add_excel_page(P_BODY)

        # 3. 사업자등록증 (공통 — Supabase)
        _add_common_doc('biz_registration')

        # 4~5. 납세증명서 (현장별 — 매번 새로 발급)
        _add_attachment_pdf('tax_cert_national')
        _add_attachment_pdf('tax_cert_local')

        # 6. 공장등록증명서 (공통 — Supabase)
        _add_common_doc('factory_cert')

        # 7. 직접생산확인증명서 (품목에 따라 자동 선택)
        if 'MTT' in first_model_upper:
            _add_common_doc('direct_production_tower')
        elif 'MTPF' in first_model_upper or 'MTPS' in first_model_upper:
            _add_common_doc('direct_production_pole')
        else:
            _add_common_doc('direct_production_light')

        # 8. 현장대리인계
        if agent_user:
            _add_excel_page(P_AGENT)

        # 9. 재직증명서
        if agent_user:
            _add_excel_page(P_CERT)

        # 10. 물품계약서 (업로드한 납품요구서 PDF)
        if package.req_pdf_path:
            req_path = package.req_pdf_path
            if req_path.startswith('documents/'):
                # Supabase Storage
                _add_supabase_pdf(req_path)
            else:
                # 레거시 로컬 경로
                if not os.path.isabs(req_path):
                    req_path = os.path.join(os.path.dirname(__file__), '..', '..', req_path)
                if os.path.exists(req_path):
                    try:
                        req_reader = pypdf.PdfReader(req_path)
                        for page in req_reader.pages:
                            writer.add_page(page)
                    except Exception as e:
                        logger.warning("납품요구서 PDF 읽기 실패: %s", e)

        # 11. 예정공정표
        if 'MTT' in first_model_upper:
            _add_excel_page(P_SCHED_TOWER)
        elif 'MTPF' in first_model_upper or 'MTPS' in first_model_upper:
            _add_excel_page(P_SCHED_POLE)
        else:
            _add_excel_page(P_SCHED_LIGHT)

        # 12. 납품내역서
        _add_excel_page(P_DELIVERY)

        # 13. 제작도면 (모델별 자동 연동 — Supabase)
        for it in items:
            if it['model']:
                _add_drawing_by_model(it['model'])

        buf = io.BytesIO()
        writer.write(buf)
        buf.seek(0)

        logger.info("착수계 PDF 생성 완료: %s (%d pages, %d bytes)",
                     business_name, writer._root_object['/Pages']['/Count'], buf.getbuffer().nbytes)
        return buf
