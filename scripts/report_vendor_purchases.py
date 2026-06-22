"""
2023~2026 업체별 구매(입고) 내역 집계 → 엑셀 생성 → 이메일 발송.

데이터 소스: light_sync.receiving_history (iCUBE 입고이력 마이그레이션)
  - purchase_order_history(발주이력)는 미마이그레이션(0건)
  - 신규 ERP purchase_orders는 2026년 일부(51건)만 존재 → 다년도 일관성 위해 제외
  - 입고이력이 2020~2026 전 기간을 담은 유일한 다년도 구매 데이터

집계 단위: 업체명(vendor_name) × 연도
  - 건수: 입고전표 수
  - 수량: items_json 내 모든 품목 qty 합계 (단위 혼재)
  - 금액: 입고금액(total_amount) 합계
"""

import io
import json
import datetime
from collections import defaultdict

from dotenv import load_dotenv
load_dotenv()

from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from modules.db_context import get_db
from modules.services.email_sender import send_email_with_attachments

YEARS = [2023, 2024, 2025, 2026]
RECIPIENT = "ljh383177@mgnt.kr"

# ---------------------------------------------------------------------------
# 1) 데이터 집계
# ---------------------------------------------------------------------------
def aggregate():
    # vendor -> year -> {'cnt','qty','amt'}
    data = defaultdict(lambda: defaultdict(lambda: {'cnt': 0, 'qty': 0.0, 'amt': 0.0}))
    sql = text("""
        SELECT vendor_name,
               COALESCE(EXTRACT(YEAR FROM receive_date)::int,
                        2000 + substring(icube_rcv_nb, 3, 2)::int) AS yr,
               total_amount,
               items_json
        FROM light_sync.receiving_history
        WHERE COALESCE(EXTRACT(YEAR FROM receive_date)::int,
                       2000 + substring(icube_rcv_nb, 3, 2)::int) BETWEEN 2023 AND 2026
    """)
    with get_db() as db:
        for vendor, yr, amt, items_json in db.execute(sql):
            vendor = (vendor or '(미상)').strip() or '(미상)'
            bucket = data[vendor][yr]
            bucket['cnt'] += 1
            bucket['amt'] += float(amt or 0)
            if items_json:
                try:
                    for it in json.loads(items_json):
                        bucket['qty'] += float(it.get('qty') or 0)
                except (ValueError, TypeError):
                    pass
    return data


def vendor_total(rows):
    return sum(rows[y]['amt'] for y in YEARS if y in rows)


# ---------------------------------------------------------------------------
# 2) 엑셀 생성
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
NUMFMT = '#,##0'


def _style_header(cell, fill=HEADER_FILL):
    cell.fill = fill
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def build_summary_sheet(wb, data):
    ws = wb.active
    ws.title = "업체별_연도요약"

    # 2-row header: 업체명 | (연도: 건수/수량/금액) x N | 합계(건수/수량/금액)
    ws.merge_cells('A1:A2')
    c = ws.cell(row=1, column=1, value="업체명"); _style_header(c)
    ws.cell(row=2, column=1).border = BORDER; ws.cell(row=2, column=1).fill = HEADER_FILL

    col = 2
    groups = [str(y) for y in YEARS] + ["합계"]
    for g in groups:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        gc = ws.cell(row=1, column=col, value=g + "년" if g != "합계" else "합계")
        _style_header(gc, HEADER_FILL if g != "합계" else PatternFill("solid", fgColor="C55A11"))
        for j, sub in enumerate(["건수", "수량", "금액(원)"]):
            sc = ws.cell(row=2, column=col + j, value=sub)
            _style_header(sc, SUB_FILL if g != "합계" else PatternFill("solid", fgColor="ED7D31"))
        col += 3

    # rows sorted by total amount desc
    vendors = sorted(data.keys(), key=lambda v: vendor_total(data[v]), reverse=True)
    grand = defaultdict(lambda: {'cnt': 0, 'qty': 0.0, 'amt': 0.0})
    r = 3
    for vendor in vendors:
        rows = data[vendor]
        ws.cell(row=r, column=1, value=vendor).border = BORDER
        col = 2
        tot = {'cnt': 0, 'qty': 0.0, 'amt': 0.0}
        for y in YEARS:
            b = rows.get(y, {'cnt': 0, 'qty': 0.0, 'amt': 0.0})
            for k in ('cnt', 'qty', 'amt'):
                tot[k] += b[k]; grand[y][k] += b[k]
            for j, k in enumerate(('cnt', 'qty', 'amt')):
                cell = ws.cell(row=r, column=col + j, value=round(b[k]) if b[k] else None)
                cell.number_format = NUMFMT; cell.alignment = RIGHT; cell.border = BORDER
            col += 3
        for j, k in enumerate(('cnt', 'qty', 'amt')):
            cell = ws.cell(row=r, column=col + j, value=round(tot[k]) if tot[k] else 0)
            cell.number_format = NUMFMT; cell.alignment = RIGHT; cell.border = BORDER
            cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL
        r += 1

    # grand total row
    ws.cell(row=r, column=1, value="총계").font = TOTAL_FONT
    ws.cell(row=r, column=1).fill = TOTAL_FILL; ws.cell(row=r, column=1).border = BORDER
    col = 2
    gt = {'cnt': 0, 'qty': 0.0, 'amt': 0.0}
    for y in YEARS:
        for j, k in enumerate(('cnt', 'qty', 'amt')):
            v = grand[y][k]; gt[k] += v
            cell = ws.cell(row=r, column=col + j, value=round(v))
            cell.number_format = NUMFMT; cell.alignment = RIGHT; cell.border = BORDER
            cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL
        col += 3
    for j, k in enumerate(('cnt', 'qty', 'amt')):
        cell = ws.cell(row=r, column=col + j, value=round(gt[k]))
        cell.number_format = NUMFMT; cell.alignment = RIGHT; cell.border = BORDER
        cell.font = TOTAL_FONT; cell.fill = PatternFill("solid", fgColor="F4B183")

    ws.freeze_panes = "B3"
    ws.column_dimensions['A'].width = 28
    for i in range(2, 2 + 3 * (len(YEARS) + 1)):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = 12
    return gt


def build_year_sheet(wb, data, year):
    ws = wb.create_sheet(title=f"{year}년")
    headers = ["순위", "업체명", "건수", "수량", "금액(원)"]
    for i, h in enumerate(headers, 1):
        _style_header(ws.cell(row=1, column=i, value=h))
    rows = []
    for vendor, yrs in data.items():
        if year in yrs:
            b = yrs[year]
            rows.append((vendor, b['cnt'], b['qty'], b['amt']))
    rows.sort(key=lambda x: x[3], reverse=True)
    r = 2
    tc = tq = ta = 0
    for rank, (vendor, cnt, qty, amt) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=rank).alignment = CENTER
        ws.cell(row=r, column=2, value=vendor)
        for j, val in enumerate((cnt, qty, amt), 3):
            cell = ws.cell(row=r, column=j, value=round(val))
            cell.number_format = NUMFMT; cell.alignment = RIGHT
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDER
        tc += cnt; tq += qty; ta += amt
        r += 1
    ws.cell(row=r, column=2, value="합계").font = TOTAL_FONT
    for j, val in enumerate((tc, tq, ta), 3):
        cell = ws.cell(row=r, column=j, value=round(val))
        cell.number_format = NUMFMT; cell.alignment = RIGHT
        cell.font = TOTAL_FONT
    for col in range(1, 6):
        ws.cell(row=r, column=col).fill = TOTAL_FILL
        ws.cell(row=r, column=col).border = BORDER
    ws.freeze_panes = "A2"
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    for col in ('C', 'D', 'E'):
        ws.column_dimensions[col].width = 14


def build_excel(data):
    wb = Workbook()
    gt = build_summary_sheet(wb, data)
    for y in YEARS:
        build_year_sheet(wb, data, y)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), gt


# ---------------------------------------------------------------------------
# 3) 메인
# ---------------------------------------------------------------------------
def main():
    data = aggregate()
    xlsx_bytes, gt = build_excel(data)
    today = datetime.date.today().strftime("%Y%m%d")
    fname = f"업체별_구매내역_2023-2026_{today}.xlsx"

    body = (
        "안녕하세요.\n\n"
        "2023년~2026년 업체별 구매(입고) 내역 집계 자료를 첨부드립니다.\n\n"
        "[집계 개요]\n"
        f" - 대상 기간: 2023 ~ 2026년\n"
        f" - 업체 수: {len(data)}개\n"
        f" - 총 입고건수: {round(gt['cnt']):,}건\n"
        f" - 총 입고금액: {round(gt['amt']):,}원\n\n"
        "[데이터 기준]\n"
        " - 소스: iCUBE 입고이력(receiving_history) 기준 집계\n"
        " - 금액: 입고금액 합계 / 수량: 품목 수량 합계(단위 혼재) / 건수: 입고전표 수\n"
        " - 시트 구성: 업체별_연도요약 + 연도별(2023~2026) 상세\n\n"
        "감사합니다.\n"
    )

    print(f"[집계] 업체 {len(data)}개 / 총 {round(gt['cnt']):,}건 / {round(gt['amt']):,}원")
    print(f"[엑셀] {fname} ({len(xlsx_bytes):,} bytes)")

    result = send_email_with_attachments(
        to_email=RECIPIENT,
        subject="[매그나텍] 2023~2026 업체별 구매(입고) 내역 집계",
        body_text=body,
        attachments=[(fname, xlsx_bytes)],
        from_name="매그나텍 구매팀",
    )
    print(f"[발송] {result}")

    # 로컬 사본(검토용) — static 아님, logs 디렉토리에 저장
    import os
    os.makedirs("logs", exist_ok=True)
    with open(os.path.join("logs", fname), "wb") as f:
        f.write(xlsx_bytes)
    print(f"[사본] logs/{fname}")


if __name__ == "__main__":
    main()
