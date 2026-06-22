"""
연도별(2024·2025·2026) 매출처/매입처 거래처별 집계 → 엑셀 2종 생성 → 이메일 발송.

요청 사양
  - 컬럼: 업체명, 사업자번호, 연도별 거래금액(2024·2025·2026)
  - 금액 기준: 합계금액(부가세 포함)
  - 매출/매입 별도 엑셀 → ljh383177@mgnt.kr 발송

데이터 소스
  - 매출처: light_sync.tax_invoices (당사 발행 전자세금계산서)
      · 공급자=(주)매그나텍 고정, 거래처=공급받는자(buyer)
      · 금액=합계금액(total_amount, VAT 포함). 수정세금계산서 포함(순액)
  - 매입처: light_sync.receiving_history (iCUBE 입고이력) ⋈ vendors(사업자번호)
      · 입고이력이 다년도(2020~2026)를 담은 유일한 매입 데이터
        (purchase_order_history 0건, 신규 purchase_orders 2026년 일부만)
      · 입고금액(공급가액 추정)에 VAT 10% 환산(×1.1) → 합계금액 기준 맞춤
      · vendor_tr_cd → vendors.icube_tr_cd 로 사업자번호 매칭(미매칭은 빈칸)
"""

import io
import re
import datetime
from collections import defaultdict

from dotenv import load_dotenv
load_dotenv()

from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from modules.db_context import get_db
from modules.services.email_sender import send_email_with_attachments

YEARS = [2024, 2025, 2026]
RECIPIENT = "ljh383177@mgnt.kr"
VAT_RATE = 1.1  # 매입 입고금액(공급가액 추정) → 합계금액(부가세 포함) 환산


# ---------------------------------------------------------------------------
# 공통 유틸
# ---------------------------------------------------------------------------
def norm_bizno(raw):
    """사업자번호에서 숫자만 추출. 10자리면 키로 사용, 아니면 None."""
    if not raw:
        return None
    digits = re.sub(r'\D', '', str(raw))
    return digits if len(digits) == 10 else None


def fmt_bizno(digits):
    """10자리 숫자 → XXX-XX-XXXXX."""
    if digits and len(digits) == 10:
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
    return digits or ''


class Aggregator:
    """거래처(사업자번호 우선, 없으면 업체명) × 연도 금액 집계 + 대표업체명 선정."""

    def __init__(self):
        # key -> {'bizno': digits|None, 'years': {yr: amt}, 'names': {name: amt}}
        self.data = defaultdict(lambda: {'bizno': None, 'years': defaultdict(float),
                                         'names': defaultdict(float)})

    def add(self, bizno_raw, name, year, amount):
        digits = norm_bizno(bizno_raw)
        name = (name or '').strip()
        key = digits or ('NM:' + name)
        rec = self.data[key]
        rec['bizno'] = digits
        rec['years'][year] += float(amount or 0)
        if name:
            rec['names'][name] += float(amount or 0)

    def rows(self):
        """[(업체명, 사업자번호표시, {yr:amt}, 합계)] 합계 내림차순."""
        out = []
        for rec in self.data.values():
            name = max(rec['names'].items(), key=lambda x: x[1])[0] if rec['names'] else '(미상)'
            total = sum(rec['years'].get(y, 0) for y in YEARS)
            out.append((name, fmt_bizno(rec['bizno']), dict(rec['years']), total))
        out.sort(key=lambda x: x[3], reverse=True)
        return out


# ---------------------------------------------------------------------------
# 1) 집계
# ---------------------------------------------------------------------------
def aggregate_sales():
    agg = Aggregator()
    sql = text("""
        SELECT buyer_business_no, buyer_name,
               EXTRACT(YEAR FROM issue_date)::int AS yr,
               total_amount
        FROM light_sync.tax_invoices
        WHERE EXTRACT(YEAR FROM issue_date) IN (2024, 2025, 2026)
    """)
    with get_db() as db:
        for bizno, name, yr, amt in db.execute(sql):
            agg.add(bizno, name, yr, amt)
    return agg


def aggregate_purchase():
    agg = Aggregator()
    sql = text("""
        SELECT v.business_no,
               COALESCE(NULLIF(TRIM(v.name), ''), rh.vendor_name) AS name,
               EXTRACT(YEAR FROM rh.receive_date)::int AS yr,
               rh.total_amount
        FROM light_sync.receiving_history rh
        LEFT JOIN light_sync.vendors v ON v.icube_tr_cd = rh.vendor_tr_cd
        WHERE EXTRACT(YEAR FROM rh.receive_date) IN (2024, 2025, 2026)
    """)
    with get_db() as db:
        for bizno, name, yr, amt in db.execute(sql):
            agg.add(bizno, name, yr, float(amt or 0) * VAT_RATE)  # 부가세 환산
    return agg


# ---------------------------------------------------------------------------
# 2) 엑셀 생성
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
NUMFMT = '#,##0'


def build_excel(agg, title_label):
    wb = Workbook()
    ws = wb.active
    ws.title = title_label

    headers = ["순위", "업체명", "사업자번호"] + [f"{y}년 거래금액" for y in YEARS] + ["합계"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER

    rows = agg.rows()
    year_tot = defaultdict(float)
    grand = 0.0
    r = 2
    for rank, (name, bizno, years, total) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=rank).alignment = CENTER
        ws.cell(row=r, column=2, value=name).alignment = LEFT
        ws.cell(row=r, column=3, value=bizno).alignment = CENTER
        for j, y in enumerate(YEARS):
            amt = years.get(y, 0)
            year_tot[y] += amt
            cell = ws.cell(row=r, column=4 + j, value=round(amt) if amt else None)
            cell.number_format = NUMFMT; cell.alignment = RIGHT
        tc = ws.cell(row=r, column=4 + len(YEARS), value=round(total))
        tc.number_format = NUMFMT; tc.alignment = RIGHT; tc.font = TOTAL_FONT; tc.fill = TOTAL_FILL
        grand += total
        for col in range(1, 5 + len(YEARS)):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    # 합계행
    ws.cell(row=r, column=2, value="합계").font = TOTAL_FONT
    ws.cell(row=r, column=2).alignment = LEFT
    for j, y in enumerate(YEARS):
        cell = ws.cell(row=r, column=4 + j, value=round(year_tot[y]))
        cell.number_format = NUMFMT; cell.alignment = RIGHT; cell.font = TOTAL_FONT
    gc = ws.cell(row=r, column=4 + len(YEARS), value=round(grand))
    gc.number_format = NUMFMT; gc.alignment = RIGHT; gc.font = TOTAL_FONT
    for col in range(1, 5 + len(YEARS)):
        ws.cell(row=r, column=col).fill = TOTAL_FILL
        ws.cell(row=r, column=col).border = BORDER

    ws.freeze_panes = "A2"
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 16
    for i in range(4, 5 + len(YEARS)):
        ws.column_dimensions[get_column_letter(i)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(rows), grand, dict(year_tot)


# ---------------------------------------------------------------------------
# 3) 메인
# ---------------------------------------------------------------------------
def _summary_lines(label, cnt, grand, year_tot):
    lines = [f"[{label}] 거래처 {cnt}개 / 합계 {round(grand):,}원"]
    for y in YEARS:
        lines.append(f"    - {y}년: {round(year_tot.get(y, 0)):,}원")
    return "\n".join(lines)


def main(send=True):
    sales = aggregate_sales()
    purchase = aggregate_purchase()

    s_bytes, s_cnt, s_grand, s_yt = build_excel(sales, "매출처_연도별")
    p_bytes, p_cnt, p_grand, p_yt = build_excel(purchase, "매입처_연도별")

    today = datetime.date.today().strftime("%Y%m%d")
    s_name = f"매출처_연도별_거래금액_2024-2026_{today}.xlsx"
    p_name = f"매입처_연도별_거래금액_2024-2026_{today}.xlsx"

    print(_summary_lines("매출처", s_cnt, s_grand, s_yt))
    print(_summary_lines("매입처", p_cnt, p_grand, p_yt))

    body = (
        "안녕하세요.\n\n"
        "2024년·2025년·2026년 연도별 매출처/매입처 거래금액 집계 자료를 첨부드립니다.\n"
        "매출과 매입을 각각 별도 엑셀로 정리했습니다.\n\n"
        f"{_summary_lines('매출처', s_cnt, s_grand, s_yt)}\n"
        f"{_summary_lines('매입처', p_cnt, p_grand, p_yt)}\n\n"
        "[데이터 기준]\n"
        " · 금액 기준: 합계금액(부가세 포함)\n"
        " · 매출처: 당사 발행 전자세금계산서(tax_invoices)의 공급받는자 기준 합계.\n"
        "           수정세금계산서(마이너스) 포함 순액.\n"
        " · 매입처: iCUBE 입고이력(receiving_history) 기준 입고금액에 부가세 10% 환산(×1.1).\n"
        "           사업자번호는 거래처원장(vendors) 매칭값 사용(일부 미매칭 거래처는 빈칸).\n"
        " · 2026년은 진행 중(자료 생성일까지)인 미완결 기간입니다.\n\n"
        "감사합니다.\n"
    )

    import os
    os.makedirs("logs", exist_ok=True)
    for fn, by in ((s_name, s_bytes), (p_name, p_bytes)):
        with open(os.path.join("logs", fn), "wb") as f:
            f.write(by)
        print(f"[사본] logs/{fn} ({len(by):,} bytes)")

    if not send:
        print("[발송 생략] send=False")
        return

    result = send_email_with_attachments(
        to_email=RECIPIENT,
        subject="[매그나텍] 2024~2026 연도별 매출처/매입처 거래금액 집계",
        body_text=body,
        attachments=[(s_name, s_bytes), (p_name, p_bytes)],
        from_name="매그나텍 관리부",
    )
    print(f"[발송] {result}")


if __name__ == "__main__":
    import sys
    main(send="--no-send" not in sys.argv)
