"""
BOM LIST 엑셀 -> DB 마이그레이션 스크립트.

Usage:
    python scripts/import_bom_from_excel.py              # 실행 (dry-run)
    python scripts/import_bom_from_excel.py --commit      # 실제 DB 반영
    python scripts/import_bom_from_excel.py --reset       # 기존 BOM 삭제 후 재임포트
"""

import argparse
import json
import os
import sys

# 프로젝트 루트를 path에 추가
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from sqlalchemy import text

from modules.models.db import engine, SessionLocal
from modules.models.entities import BomHeader, BomItem, Item

# ============================================================
# 시트별 설정
# ============================================================
# (sheet_index, category_name, skip)
SHEET_CONFIG = [
    (0,  "가로등기구"),       # 실내등(SLA)
    (1,  "가로등기구"),       # 실내등(SLC)
    (2,  "보안등기구"),       # 실외등(ML)
    (3,  "터널등기구"),       # 터널등(TL)
    (4,  "투광등기구"),       # 투광등(BATOO)
    (5,  "투광등기구"),       # 투광등(MT-FL)
    (6,  "투광등기구"),       # 투광등(ARENA)
    (7,  None),              # 투광등 표준가 참고 시트 -> 스킵
    (8,  "투광등기구"),       # 투광등(STA)
    (9,  "투광등기구"),       # 투광등(STA-S)
    (10, "조명타워"),         # 보안타워
    (11, None),              # 등대폴대 -> 비표준 구조, 스킵
    (12, "스텐가로등주"),     # 알루미늄폴대
    (13, "철제가로등주"),     # 철재폴대
    (14, "태양광가로등"),     # 태양광보안등
    (15, "태양광가로등"),     # 태양광보안등(분전반)
]

# 표준 시트: No, 인증번호, 완성품코드, 품목코드, 품명, 수량, 납품업체, 소요량, 단가, 금액, 비고
# 변형 시트(보안타워/폴대/태양광): No, 인증번호, 완성품코드, 품명, 규격, 수량, 납품업체, 소요량, 단가, 금액, 비고
# 태양광보안등(14): No, 인증번호, 완성품코드, 품명, 수량, 납품업체, 소요량, 단가, 금액, 비고(인증번호 없는 변형)

# 표준 컬럼 인덱스 (0-based)
STANDARD_COLS = {
    'no': 0, 'cert_no': 1, 'product_code': 2,
    'item_code': 3, 'item_name': 4, 'qty': 5,
    'supplier': 6, 'need_qty': 7, 'unit_price': 8,
    'amount': 9, 'note': 10, 'option_filter': 11,
}

# 변형 컬럼 (보안타워, 폴대류) - 품번 대신 규격
VARIANT_COLS = {
    'no': 0, 'cert_no': 1, 'product_code': 2,
    'item_name': 3, 'item_spec': 4, 'qty': 5,
    'supplier': 6, 'need_qty': 7, 'unit_price': 8,
    'amount': 9, 'note': 10, 'option_filter': 11,
}

# 변형 시트 인덱스 (품번 컬럼이 없는 시트)
VARIANT_SHEETS = {10, 12, 13, 14, 15}


def find_header_row(ws):
    """'No' 셀을 찾아서 헤더 행 번호 반환."""
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=False):
        if row[0].value == 'No':
            return row[0].row
    return None


def safe_str(v):
    if v is None:
        return ''
    return str(v).strip()


def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_option_filter(v):
    """엑셀 옵션조건 셀을 JSON 문자열로 변환.

    형식: "렌즈=20도" → '{"렌즈":"20도"}'
          "렌즈=20도,반사판=A" → '{"렌즈":"20도","반사판":"A"}'
          빈칸/None → None (공통부품)
    """
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    result = {}
    for pair in s.split(','):
        pair = pair.strip()
        if '=' not in pair:
            continue
        key, val = pair.split('=', 1)
        key, val = key.strip(), val.strip()
        if key and val:
            result[key] = val
    if not result:
        return None
    return json.dumps(result, ensure_ascii=False)


def is_standard_sheet(sheet_idx):
    return sheet_idx not in VARIANT_SHEETS


def parse_sheet(ws, sheet_idx, category):
    """시트를 파싱하여 (product_code, cert_no, items_list) 튜플 리스트 반환."""
    header_row = find_header_row(ws)
    if header_row is None:
        print(f"  [SKIP] 헤더 행을 찾을 수 없음")
        return []

    is_standard = is_standard_sheet(sheet_idx)
    cols = STANDARD_COLS if is_standard else VARIANT_COLS

    # 완성품 그룹핑 로직:
    # - 표준 시트: No에 숫자 + 새 완성품코드 = 새 그룹
    # - 변형 시트: 완성품코드가 바뀌면 새 그룹 (No는 항목 순번)
    products_map = {}  # product_code -> (cert_no, [items])
    products_order = []  # 순서 유지

    data_start = header_row + 1
    current_product_code = None

    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=False):
        cells = [c.value for c in row[:12]]

        # 빈 행 스킵
        if all(v is None for v in cells):
            continue

        # TTL 행 스킵
        product_code_val = cells[cols['product_code']] if cols['product_code'] < len(cells) else None
        if safe_str(product_code_val).upper() == 'TTL':
            continue
        if safe_str(cells[0]).upper() == 'TTL':
            continue

        # 완성품코드 확인
        pc = safe_str(product_code_val)

        if is_standard:
            # 표준 시트: No에 숫자 + 완성품코드 있으면 새 그룹
            no_val = cells[cols['no']]
            is_new_group = False
            if no_val is not None:
                try:
                    int(float(str(no_val)))
                    is_new_group = True
                except (ValueError, TypeError):
                    pass

            if is_new_group and pc:
                current_product_code = pc
            elif pc:
                current_product_code = pc  # 같은 코드가 반복되는 행
            # pc 없으면 이전 product_code 유지
        else:
            # 변형 시트: 완성품코드로만 그룹핑
            if pc:
                current_product_code = pc

        if not current_product_code:
            continue

        # 인증번호
        cert_no_val = cells[cols.get('cert_no', 1)]
        cert_no = safe_str(cert_no_val) if cert_no_val else None

        # 품명
        if is_standard:
            item_name = safe_str(cells[cols['item_name']])
            item_code = safe_str(cells[cols['item_code']])
            note_val = safe_str(cells[cols.get('note', 10)])
            item_spec = note_val  # 표준 시트에서는 비고를 spec으로 활용
        else:
            item_name = safe_str(cells[cols['item_name']])
            item_code = ''
            item_spec = safe_str(cells[cols.get('item_spec', 4)])

        if not item_name:
            continue

        supplier = safe_str(cells[cols['supplier']]) if cols['supplier'] < len(cells) else ''
        need_qty = safe_float(cells[cols['need_qty']]) if cols['need_qty'] < len(cells) else 1
        unit_price = safe_float(cells[cols['unit_price']]) if cols['unit_price'] < len(cells) else None
        amount = safe_float(cells[cols['amount']]) if cols['amount'] < len(cells) else None

        # 옵션조건 파싱 (컬럼 11, 없으면 None)
        option_raw = cells[cols['option_filter']] if cols['option_filter'] < len(cells) else None
        option_filter = parse_option_filter(option_raw)

        item_data = {
            'item_code': item_code,
            'item_name': item_name,
            'item_spec': item_spec,
            'quantity': need_qty or 1,
            'unit_price': unit_price,
            'amount': amount,
            'supplier': supplier,
            'option_filter': option_filter,
        }

        if current_product_code not in products_map:
            products_map[current_product_code] = (cert_no, [])
            products_order.append(current_product_code)
        elif cert_no and not products_map[current_product_code][0]:
            # 인증번호 업데이트 (첫 행에 없었을 수 있음)
            products_map[current_product_code] = (cert_no, products_map[current_product_code][1])

        products_map[current_product_code][1].append(item_data)

    # 순서대로 결과 반환
    return [(pc, products_map[pc][0], products_map[pc][1]) for pc in products_order]


def ensure_columns(session):
    """BomHeader/BomItem에 새 컬럼이 없으면 추가."""
    alter_statements = [
        ("bom_headers", "product_category", "VARCHAR(50)"),
        ("bom_headers", "certification_no", "VARCHAR(50)"),
        ("bom_items", "item_code", "VARCHAR(100)"),
        ("bom_items", "item_id", "INTEGER"),
        ("bom_items", "unit_price", "FLOAT"),
        ("bom_items", "amount", "FLOAT"),
        ("bom_items", "supplier", "VARCHAR(200)"),
        ("bom_items", "option_filter", "TEXT"),
        ("bom_headers", "option_schema", "TEXT"),
    ]
    for table, column, col_type in alter_statements:
        try:
            session.execute(text(
                f"ALTER TABLE light_sync.{table} ADD COLUMN IF NOT EXISTS {column} {col_type}"
            ))
        except Exception as e:
            # 이미 존재하면 무시
            if 'already exists' not in str(e).lower() and 'duplicate' not in str(e).lower():
                print(f"  Warning: ALTER TABLE {table}.{column}: {e}")
            session.rollback()
    session.commit()


def main():
    parser = argparse.ArgumentParser(description='BOM 엑셀 -> DB 임포트')
    parser.add_argument('--commit', action='store_true', help='실제 DB 반영 (기본: dry-run)')
    parser.add_argument('--reset', action='store_true', help='기존 BOM 데이터 삭제 후 재임포트')
    args = parser.parse_args()

    excel_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'reference', '제품 및 자재LIST(BOM+인건비).xlsx'
    )

    if not os.path.exists(excel_path):
        print(f"ERROR: 엑셀 파일 없음: {excel_path}")
        sys.exit(1)

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"Sheets: {len(wb.sheetnames)}")

    session = SessionLocal()

    try:
        # 1. 컬럼 추가 (마이그레이션)
        print("\n[1/4] 컬럼 마이그레이션...")
        ensure_columns(session)

        # 2. items 매핑 캐시 생성
        print("[2/4] items 테이블 로드...")
        items = session.query(Item).all()
        item_map = {}  # icube_item_cd -> Item
        for it in items:
            if it.icube_item_cd:
                item_map[it.icube_item_cd.strip()] = it
        print(f"  items 수: {len(item_map)}")

        # 3. 기존 BOM 확인
        existing_count = session.query(BomHeader).count()
        print(f"  기존 BOM 수: {existing_count}")

        if args.reset and existing_count > 0:
            if args.commit:
                print("  -> 기존 BOM 삭제 중...")
                session.query(BomItem).delete()
                session.query(BomHeader).delete()
                session.commit()
                print("  -> 삭제 완료")
            else:
                print("  -> (dry-run) 기존 BOM 삭제 예정")

        # 4. 엑셀 파싱 및 DB 임포트
        print("\n[3/4] 엑셀 파싱 및 임포트...")

        total_headers = 0
        total_items = 0
        matched_items = 0
        updated_manufacturers = 0
        skipped_existing = 0

        for sheet_idx, category in SHEET_CONFIG:
            if category is None:
                print(f"  Sheet {sheet_idx}: (스킵)")
                continue

            ws = wb[wb.sheetnames[sheet_idx]]
            print(f"  Sheet {sheet_idx}: {category} (rows={ws.max_row})")

            products = parse_sheet(ws, sheet_idx, category)
            print(f"    파싱된 완성품: {len(products)}")

            for product_code, cert_no, items_data in products:
                # 중복 체크
                existing = session.query(BomHeader).filter_by(product_code=product_code).first()
                if existing and not args.reset:
                    skipped_existing += 1
                    continue

                # option_schema 자동 생성: 부품들의 option_filter에서 옵션 키/값 수집
                option_schema = {}
                for item_data in items_data:
                    of = item_data.get('option_filter')
                    if of:
                        try:
                            parsed = json.loads(of)
                            for k, v in parsed.items():
                                if k not in option_schema:
                                    option_schema[k] = []
                                if v not in option_schema[k]:
                                    option_schema[k].append(v)
                        except (json.JSONDecodeError, TypeError):
                            pass

                # BomHeader 생성
                bom = BomHeader(
                    product_code=product_code,
                    product_name=product_code,  # 완성품코드를 이름으로 사용
                    product_category=category,
                    certification_no=cert_no,
                    version='1.0',
                    is_active=True,
                    option_schema=json.dumps(option_schema, ensure_ascii=False) if option_schema else None,
                )

                if args.commit:
                    session.add(bom)
                    session.flush()

                total_headers += 1

                for item_data in items_data:
                    ic = item_data['item_code']
                    matched_item = item_map.get(ic) if ic else None

                    bom_item = BomItem(
                        bom_id=bom.id if args.commit else 0,
                        item_code=ic or None,
                        item_id=matched_item.id if matched_item else None,
                        item_name=item_data['item_name'],
                        item_spec=item_data['item_spec'] or None,
                        quantity=item_data['quantity'],
                        unit_price=item_data['unit_price'],
                        amount=item_data['amount'],
                        supplier=item_data['supplier'] or None,
                        option_filter=item_data.get('option_filter'),
                    )

                    if args.commit:
                        session.add(bom_item)

                    total_items += 1
                    if matched_item:
                        matched_items += 1

                    # items.manufacturer 업데이트
                    if matched_item and item_data['supplier'] and not matched_item.manufacturer:
                        matched_item.manufacturer = item_data['supplier']
                        updated_manufacturers += 1

        if args.commit:
            session.commit()
            print("\n  -> DB 커밋 완료!")
        else:
            session.rollback()
            print("\n  -> (dry-run) 커밋하지 않음. --commit 옵션으로 실행하세요.")

        # 5. 결과 리포트
        print(f"\n[4/4] 임포트 결과")
        print(f"  ={'='*50}")
        print(f"  BomHeader 생성: {total_headers}")
        print(f"  BomItem 생성:   {total_items}")
        print(f"  items 매칭:     {matched_items}/{total_items} ({matched_items*100//max(total_items,1)}%)")
        print(f"  manufacturer 업데이트: {updated_manufacturers}")
        print(f"  스킵 (중복):    {skipped_existing}")
        print(f"  ={'='*50}")

    except Exception as e:
        session.rollback()
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        session.close()


if __name__ == '__main__':
    main()
