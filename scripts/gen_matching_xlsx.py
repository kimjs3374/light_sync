"""엑셀 현장별 정보 ↔ DB 활성 계약 매칭 결과를 엑셀로 출력.
등기구 + 등주 두 시트를 각각 매칭합니다.
"""
import sys, io, json, re, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import text
from modules.db_context import get_db
from difflib import SequenceMatcher

wb_old = xlrd.open_workbook('현장별 정보.xls', formatting_info=True)

with get_db() as db:
    active = db.execute(text("""
        SELECT ci.id, ci.model_name, ci.category, ci.item_spec_json, ci.status_sales,
               p.id as pid, p.temp_name, p.short_name, p.project_no,
               c.id as cid, c.contract_name, c.payment_status, c.delivery_due_date
        FROM light_sync.contract_items ci
        JOIN light_sync.contracts c ON c.id = ci.contract_id
        JOIN light_sync.projects p ON p.id = c.project_id
        WHERE p.is_contracted = true
          AND c.payment_status NOT IN ('입금완료', '변경완료', '취소')
        ORDER BY p.id DESC
    """)).fetchall()


def normalize(s):
    return re.sub(r'[\s\-_()（）\[\]]+', '', (s or '')).upper()


def site_score(xls_site, db_names):
    best = 0
    xls_clean = re.sub(r'[\s\-_()]+', '', xls_site)
    for name in db_names:
        if not name:
            continue
        db_clean = re.sub(r'[\s\-_()]+', '', name)
        if len(xls_clean) >= 2 and xls_clean in db_clean:
            best = max(best, 0.9)
        elif len(db_clean) >= 2 and db_clean in xls_clean:
            best = max(best, 0.9)
        xls_words = set(re.findall(r'[가-힣]{2,}', xls_site))
        db_words = set(re.findall(r'[가-힣]{2,}', name))
        if xls_words and db_words:
            overlap = len(xls_words & db_words) / max(len(xls_words), len(db_words))
            best = max(best, overlap)
        ratio = SequenceMatcher(None, xls_clean, db_clean).ratio()
        best = max(best, ratio)
    return best


def model_match(xls_norm, db_norm):
    if not xls_norm or not db_norm:
        return False
    if xls_norm in db_norm or db_norm in xls_norm:
        return True
    xls_nums = re.findall(r'\d+', xls_norm)
    db_nums = re.findall(r'\d+', db_norm)
    if xls_nums and db_nums and xls_nums[-1] == db_nums[-1]:
        xls_alpha = re.sub(r'\d+', '', xls_norm)
        db_alpha = re.sub(r'\d+', '', db_norm)
        if xls_alpha in db_alpha or db_alpha in xls_alpha:
            return True
    return False


def extract_model_from_product(product_str):
    """등주 시트의 '스테인리스 가로등주 ( MTPS-104-8 )' 형태에서 모델명 추출."""
    m = re.search(r'[（(]\s*([A-Z][\w\-]+)\s*[）)]', product_str)
    if m:
        return m.group(1)
    # 괄호 없으면 그대로
    m = re.search(r'(MT\w[\w\-]*|MTPS[\w\-]*|MTPF[\w\-]*|MTT\w+|ARENA[\w\-]*)', product_str, re.IGNORECASE)
    if m:
        return m.group(1)
    return product_str


# ── 스타일 ──
header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
header_font = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
match_fill = PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid')
warn_fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
fail_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
thin = Side(style='thin', color='d1d5db')
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
body_font = Font(name='맑은 고딕', size=9)

wb = openpyxl.Workbook()
used_db_ids = set()
sheet_stats = {}


def write_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bdr


def write_row(ws, row_idx, values, fill):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=c, value=val)
        cell.fill = fill
        cell.font = body_font
        cell.border = bdr
        cell.alignment = Alignment(vertical='center', wrap_text=(c in (3, 9, 13)))


def auto_width(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 42)


# ════════════════════════════════════════
# 1) 등기구 시트
# ════════════════════════════════════════
sheet_src = wb_old.sheet_by_name('등기구')
ws1 = wb.active
ws1.title = '등기구_매칭'

headers1 = [
    '매칭', '유사도',
    '엑셀_현장명', '엑셀_모델',
    '엑셀_렌즈', '엑셀_이격', '엑셀_SMPS모델', '엑셀_케이블종류', '엑셀_비고',
    '엑셀_생산', '엑셀_출고',
    'DB_item_id', 'DB_현장명', 'DB_모델', 'DB_계약명', 'DB_납기', 'DB_협의상태',
    '수동매칭_item_id',
]
write_header(ws1, headers1)

rows1 = []
row_idx = 2
for r in range(2, sheet_src.nrows):
    site = str(sheet_src.cell_value(r, 6)).strip()
    product = str(sheet_src.cell_value(r, 7)).strip()
    if not product:
        continue
    # 숨김 행 + 출고 완료건 제외
    row_info = sheet_src.rowinfo_map.get(r)
    if row_info and row_info.hidden:
        continue
    ship_st = str(sheet_src.cell_value(r, 14)).strip()
    if ship_st == '출고':
        continue

    xls_norm = normalize(product)
    best_m, best_s = None, 0
    for db_item in active:
        if db_item.id in used_db_ids:
            continue
        if not model_match(xls_norm, normalize(db_item.model_name)):
            continue
        s = site_score(site, [db_item.temp_name, db_item.short_name, db_item.contract_name])
        if s > best_s:
            best_s = s
            best_m = db_item

    if best_m and best_s >= 0.3:
        used_db_ids.add(best_m.id)

    has = best_m and best_s >= 0.3
    if best_s >= 0.6:
        status, fill = '✅', match_fill
    elif best_s >= 0.3:
        status, fill = '⚠️', warn_fill
    else:
        status, fill = '❌', fail_fill

    smps_val = str(sheet_src.cell_value(r, 19)).strip()
    if smps_val in ('有', ''):
        smps_val = ''

    vals = [
        status, f"{best_s:.0%}" if best_s > 0 else '',
        site, product,
        str(sheet_src.cell_value(r, 10)).strip(),
        str(sheet_src.cell_value(r, 11)).strip(),
        smps_val,
        str(sheet_src.cell_value(r, 21)).strip(),
        str(sheet_src.cell_value(r, 12)).strip(),
        str(sheet_src.cell_value(r, 13)).strip(),
        ship_st,
        best_m.id if has else '',
        (best_m.short_name or best_m.temp_name or '') if has else '',
        best_m.model_name if has else '',
        (best_m.contract_name or '')[:50] if has else '',
        str(best_m.delivery_due_date) if has and best_m.delivery_due_date else '',
        (best_m.status_sales or '계약확인') if has else '',
        '',
    ]
    write_row(ws1, row_idx, vals, fill)
    rows1.append({'score': best_s})
    row_idx += 1

ws1.auto_filter.ref = f'A1:R{row_idx - 1}'
ws1.freeze_panes = 'A2'
auto_width(ws1)

s1_ok = sum(1 for x in rows1 if x['score'] >= 0.6)
s1_warn = sum(1 for x in rows1 if 0.3 <= x['score'] < 0.6)
s1_fail = sum(1 for x in rows1 if x['score'] < 0.3)

# ════════════════════════════════════════
# 2) 등주 시트
# ════════════════════════════════════════
sheet_src2 = wb_old.sheet_by_name('등주')
ws2 = wb.create_sheet('등주_매칭')

headers2 = [
    '매칭', '유사도',
    '엑셀_현장명', '엑셀_제품(원본)', '엑셀_모델(추출)',
    '엑셀_기초', '엑셀_비고',
    '엑셀_생산', '엑셀_출고',
    'DB_item_id', 'DB_현장명', 'DB_모델', 'DB_계약명', 'DB_납기', 'DB_협의상태',
    '수동매칭_item_id',
]
write_header(ws2, headers2)

rows2 = []
row_idx = 2
for r in range(2, sheet_src2.nrows):
    site = str(sheet_src2.cell_value(r, 6)).strip()
    product_raw = str(sheet_src2.cell_value(r, 7)).strip()
    if not product_raw:
        continue

    # 숨김 행 + 출고 완료건 제외
    row_info2 = sheet_src2.rowinfo_map.get(r)
    if row_info2 and row_info2.hidden:
        continue
    ship_st = str(sheet_src2.cell_value(r, 14)).strip()
    if ship_st == '출고':
        continue

    product_model = extract_model_from_product(product_raw)
    xls_norm = normalize(product_model)

    best_m, best_s = None, 0
    for db_item in active:
        if db_item.id in used_db_ids:
            continue
        if not model_match(xls_norm, normalize(db_item.model_name)):
            continue
        s = site_score(site, [db_item.temp_name, db_item.short_name, db_item.contract_name])
        if s > best_s:
            best_s = s
            best_m = db_item

    if best_m and best_s >= 0.3:
        used_db_ids.add(best_m.id)

    has = best_m and best_s >= 0.3
    if best_s >= 0.6:
        status, fill = '✅', match_fill
    elif best_s >= 0.3:
        status, fill = '⚠️', warn_fill
    else:
        status, fill = '❌', fail_fill

    vals = [
        status, f"{best_s:.0%}" if best_s > 0 else '',
        site, product_raw, product_model,
        str(sheet_src2.cell_value(r, 10)).strip(),
        str(sheet_src2.cell_value(r, 11)).strip(),
        str(sheet_src2.cell_value(r, 13)).strip(),
        str(sheet_src2.cell_value(r, 14)).strip(),
        best_m.id if has else '',
        (best_m.short_name or best_m.temp_name or '') if has else '',
        best_m.model_name if has else '',
        (best_m.contract_name or '')[:50] if has else '',
        str(best_m.delivery_due_date) if has and best_m.delivery_due_date else '',
        (best_m.status_sales or '계약확인') if has else '',
        '',
    ]
    write_row(ws2, row_idx, vals, fill)
    rows2.append({'score': best_s})
    row_idx += 1

ws2.auto_filter.ref = f'A1:P{row_idx - 1}'
ws2.freeze_panes = 'A2'
auto_width(ws2)

s2_ok = sum(1 for x in rows2 if x['score'] >= 0.6)
s2_warn = sum(1 for x in rows2 if 0.3 <= x['score'] < 0.6)
s2_fail = sum(1 for x in rows2 if x['score'] < 0.3)

# ════════════════════════════════════════
# 3) 미매칭 DB 품목 시트
# ════════════════════════════════════════
ws3 = wb.create_sheet('미매칭_DB품목')
headers3 = ['item_id', '현장명(short)', '현장명(full)', '모델', '카테고리', '계약명', '납기', '협의상태', 'payment']
write_header(ws3, headers3)

unmatched_db = [a for a in active if a.id not in used_db_ids]
for i, db_item in enumerate(unmatched_db, 2):
    vals = [
        db_item.id,
        db_item.short_name or '',
        db_item.temp_name or '',
        db_item.model_name,
        db_item.category or '',
        (db_item.contract_name or '')[:50],
        str(db_item.delivery_due_date) if db_item.delivery_due_date else '',
        db_item.status_sales or '계약확인',
        db_item.payment_status,
    ]
    for c, val in enumerate(vals, 1):
        cell = ws3.cell(row=i, column=c, value=val)
        cell.font = body_font
        cell.border = bdr

ws3.auto_filter.ref = f'A1:I{len(unmatched_db) + 1}'
ws3.freeze_panes = 'A2'
auto_width(ws3)

out_path = '현장별_매칭결과_v2.xlsx'
wb.save(out_path)

print(f'{out_path} 생성 완료')
print(f'  등기구_매칭: {len(rows1)}행 (✅{s1_ok} / ⚠️{s1_warn} / ❌{s1_fail})')
print(f'  등주_매칭: {len(rows2)}행 (✅{s2_ok} / ⚠️{s2_warn} / ❌{s2_fail})')
print(f'  미매칭_DB품목: {len(unmatched_db)}행')
