"""
2023년 ~ 현재 분기별 원자재(매입 전체 품목) 단가 추이 분석 → 엑셀 생성 → 이메일 발송.

데이터 소스: light_sync.receiving_history (iCUBE 입고이력 마이그레이션)
  - items_json 내 각 품목의 unit_price(단가)를 분기별로 집계
  - 분기 단가 = SUM(amount) / SUM(qty)  (수량가중평균 단가)
  - 신규 ERP purchase_orders는 2026년 일부(51건)뿐이라 다년도 일관성 위해 제외
  - 입고이력이 2020~2026 전 기간 단가를 담은 유일한 다년도 매입 데이터

집계 단위: 품목(item_cd) × 분기(YYYY-Qn)
  - 분기단가: 해당 분기 입고 수량가중평균 단가
  - 최초/최근단가: 데이터가 있는 첫/마지막 분기 단가
  - 변동액 / 변동률(%): 최근 - 최초 기준
  - 총매입액: 전 기간 입고금액 합계 (정렬 기준)
"""

import io
import json
import datetime
from collections import defaultdict

from dotenv import load_dotenv
load_dotenv()

from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from modules.db_context import get_db
from modules.services.email_sender import send_email_with_attachments

RECIPIENT = "ljh383177@mgnt.kr"
START_DATE = "2023-01-01"


# ---------------------------------------------------------------------------
# 1) 데이터 집계
# ---------------------------------------------------------------------------
def quarter_of(d):
    """date -> 'YYYY-Qn'"""
    return f"{d.year}-Q{(d.month - 1) // 3 + 1}"


def aggregate():
    """item_cd -> {name, spec, unit, total_amt, q: {quarter: {amt, qty, cnt}}}"""
    items = {}
    quarters = set()
    sql = text("""
        SELECT receive_date, items_json
        FROM light_sync.receiving_history
        WHERE receive_date >= :start AND items_json IS NOT NULL
    """)
    with get_db() as db:
        for receive_date, items_json in db.execute(sql, {"start": START_DATE}):
            if not receive_date:
                continue
            q = quarter_of(receive_date)
            quarters.add(q)
            try:
                parsed = json.loads(items_json)
            except (ValueError, TypeError):
                continue
            for it in parsed:
                code = (it.get('item_cd') or '').strip()
                name = (it.get('item_name') or '').strip()
                key = code or name or '(미상)'
                qty = float(it.get('qty') or 0)
                up = float(it.get('unit_price') or 0)
                amt = float(it.get('amount') or 0)
                # unit_price 누락 시 금액/수량으로 역산
                if amt == 0 and up and qty:
                    amt = up * qty
                if amt == 0 and up == 0:
                    continue
                rec = items.get(key)
                if rec is None:
                    rec = {'code': code, 'name': name,
                           'spec': (it.get('spec') or '').strip(),
                           'unit': (it.get('unit') or '').strip(),
                           'total_amt': 0.0,
                           'q': defaultdict(lambda: {'amt': 0.0, 'qty': 0.0, 'cnt': 0})}
                    items[key] = rec
                # 최신 명칭/규격/단위로 갱신(빈 값이면 유지)
                if name:
                    rec['name'] = name
                if it.get('spec'):
                    rec['spec'] = it.get('spec').strip()
                if it.get('unit'):
                    rec['unit'] = it.get('unit').strip()
                b = rec['q'][q]
                b['amt'] += amt
                b['qty'] += qty
                b['cnt'] += 1
                rec['total_amt'] += amt
    return items, sorted(quarters)


def qprice(bucket):
    """분기 수량가중평균 단가"""
    if bucket and bucket['qty']:
        return bucket['amt'] / bucket['qty']
    return None


# ---------------------------------------------------------------------------
# 2) 엑셀 생성
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
Q_FILL = PatternFill("solid", fgColor="2E75B6")
DELTA_FILL = PatternFill("solid", fgColor="C55A11")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT = Font(bold=True)
UP_FONT = Font(color="C00000")    # 가격 상승 = 빨강
DOWN_FONT = Font(color="2E75B6")  # 가격 하락 = 파랑
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
NUMFMT = '#,##0'
PCTFMT = '+#,##0.0"%";-#,##0.0"%";0.0"%"'


def _hdr(cell, fill=HEADER_FILL):
    cell.fill = fill
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def build_price_sheet(ws, items, quarters):
    ws.title = "품목별_분기단가추이"
    base = ["품번", "품명", "규격", "단위"]
    tail = ["최초단가", "최근단가", "변동액", "변동률", "총매입액", "매입\n분기수"]
    headers = base + quarters + tail
    for i, h in enumerate(headers, 1):
        fill = HEADER_FILL
        if i > len(base) and i <= len(base) + len(quarters):
            fill = Q_FILL
        elif i > len(base) + len(quarters):
            fill = DELTA_FILL
        _hdr(ws.cell(row=1, column=i, value=h), fill)

    # 총매입액 기준 내림차순 정렬
    rows = sorted(items.values(), key=lambda r: r['total_amt'], reverse=True)
    r = 2
    for rec in rows:
        ws.cell(row=r, column=1, value=rec['code']).alignment = LEFT
        ws.cell(row=r, column=2, value=rec['name']).alignment = LEFT
        ws.cell(row=r, column=3, value=rec['spec']).alignment = LEFT
        ws.cell(row=r, column=4, value=rec['unit']).alignment = CENTER

        col = len(base) + 1
        present = []  # (quarter_index, price)
        for qi, q in enumerate(quarters):
            p = qprice(rec['q'].get(q))
            cell = ws.cell(row=r, column=col, value=round(p) if p else None)
            cell.number_format = NUMFMT
            cell.alignment = RIGHT
            if p:
                present.append((qi, p))
            col += 1

        first_p = present[0][1] if present else None
        last_p = present[-1][1] if present else None
        delta = (last_p - first_p) if (first_p is not None and last_p is not None and len(present) > 1) else None
        pct = (delta / first_p * 100) if (delta is not None and first_p) else None

        c = ws.cell(row=r, column=col, value=round(first_p) if first_p else None)
        c.number_format = NUMFMT; c.alignment = RIGHT
        c = ws.cell(row=r, column=col + 1, value=round(last_p) if last_p else None)
        c.number_format = NUMFMT; c.alignment = RIGHT
        c = ws.cell(row=r, column=col + 2, value=round(delta) if delta is not None else None)
        c.number_format = NUMFMT; c.alignment = RIGHT
        if delta is not None and delta > 0:
            c.font = UP_FONT
        elif delta is not None and delta < 0:
            c.font = DOWN_FONT
        c = ws.cell(row=r, column=col + 3, value=round(pct, 1) if pct is not None else None)
        c.number_format = PCTFMT; c.alignment = RIGHT
        if pct is not None and pct > 0:
            c.font = UP_FONT
        elif pct is not None and pct < 0:
            c.font = DOWN_FONT
        c = ws.cell(row=r, column=col + 4, value=round(rec['total_amt']))
        c.number_format = NUMFMT; c.alignment = RIGHT
        c = ws.cell(row=r, column=col + 5, value=len(present))
        c.alignment = CENTER

        for col_i in range(1, len(headers) + 1):
            ws.cell(row=r, column=col_i).border = BORDER
        r += 1

    ws.freeze_panes = ws.cell(row=2, column=len(base) + 1).coordinate
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 7
    for i in range(len(base) + 1, len(base) + len(quarters) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11
    for j in range(len(tail)):
        ws.column_dimensions[get_column_letter(len(base) + len(quarters) + 1 + j)].width = 12
    ws.row_dimensions[1].height = 30


def build_quarter_summary(ws, items, quarters):
    ws.title = "분기별_매입요약"
    headers = ["분기", "입고품목수(distinct)", "총입고건수", "총매입액(원)"]
    for i, h in enumerate(headers, 1):
        _hdr(ws.cell(row=1, column=i, value=h))
    # quarter -> totals
    qstat = {q: {'items': set(), 'cnt': 0, 'amt': 0.0} for q in quarters}
    for key, rec in items.items():
        for q, b in rec['q'].items():
            qstat[q]['items'].add(key)
            qstat[q]['cnt'] += b['cnt']
            qstat[q]['amt'] += b['amt']
    r = 2
    tot_cnt = 0
    tot_amt = 0.0
    for q in quarters:
        s = qstat[q]
        ws.cell(row=r, column=1, value=q).alignment = CENTER
        ws.cell(row=r, column=2, value=len(s['items'])).alignment = RIGHT
        ws.cell(row=r, column=3, value=s['cnt']).alignment = RIGHT
        c = ws.cell(row=r, column=4, value=round(s['amt']))
        c.number_format = NUMFMT; c.alignment = RIGHT
        for col_i in range(1, 5):
            ws.cell(row=r, column=col_i).border = BORDER
        tot_cnt += s['cnt']; tot_amt += s['amt']
        r += 1
    ws.cell(row=r, column=1, value="합계").font = TOTAL_FONT
    ws.cell(row=r, column=3, value=tot_cnt).alignment = RIGHT
    c = ws.cell(row=r, column=4, value=round(tot_amt)); c.number_format = NUMFMT; c.alignment = RIGHT
    for col_i in range(1, 5):
        ws.cell(row=r, column=col_i).fill = TOTAL_FILL
        ws.cell(row=r, column=col_i).font = TOTAL_FONT
        ws.cell(row=r, column=col_i).border = BORDER
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    return tot_cnt, tot_amt


def build_excel(items, quarters):
    wb = Workbook()
    build_price_sheet(wb.active, items, quarters)
    build_quarter_summary(wb.create_sheet(), items, quarters)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 3) 메인
# ---------------------------------------------------------------------------
def main(send=True):
    items, quarters = aggregate()
    tot_cnt = sum(b['cnt'] for r in items.values() for b in r['q'].values())
    tot_amt = sum(r['total_amt'] for r in items.values())
    xlsx_bytes = build_excel(items, quarters)
    today = datetime.date.today().strftime("%Y%m%d")
    fname = f"분기별_원자재_단가추이_{quarters[0]}~{quarters[-1]}_{today}.xlsx"

    print(f"[집계] 품목 {len(items)}개 / 분기 {len(quarters)}개 ({quarters[0]}~{quarters[-1]})")
    print(f"       총 입고 {tot_cnt:,}건 / 총 매입액 {round(tot_amt):,}원")
    print(f"[엑셀] {fname} ({len(xlsx_bytes):,} bytes)")

    # 로컬 사본(검토용) — static 아님, logs 디렉토리에 저장
    import os
    os.makedirs("logs", exist_ok=True)
    local_path = os.path.join("logs", fname)
    with open(local_path, "wb") as f:
        f.write(xlsx_bytes)
    print(f"[사본] {local_path}")

    if not send:
        print("[발송] 생략(send=False)")
        return

    body = (
        "안녕하세요, 이지훈 부장님.\n\n"
        f"2023년 1분기부터 {quarters[-1].replace('-', ' ')}까지 분기별 원자재(매입 전체 품목) 단가 추이 분석 자료를 첨부드립니다.\n\n"
        "[집계 개요]\n"
        f" - 대상 기간: {quarters[0]} ~ {quarters[-1]} (분기 {len(quarters)}개)\n"
        f" - 분석 품목수: {len(items):,}개\n"
        f" - 총 입고건수: {tot_cnt:,}건 / 총 매입액: {round(tot_amt):,}원\n\n"
        "[데이터 기준]\n"
        " - 소스: iCUBE 입고이력(receiving_history)의 품목별 단가\n"
        " - 분기단가: 해당 분기 입고 수량가중평균 단가 (총금액 / 총수량)\n"
        " - 변동액/변동률: 데이터가 있는 첫 분기 대비 마지막 분기 단가 기준\n"
        " - 정렬: 총매입액 내림차순\n\n"
        "[시트 구성]\n"
        " 1) 품목별_분기단가추이: 품목 × 분기 단가표 + 변동액/변동률\n"
        " 2) 분기별_매입요약: 분기별 품목수/건수/매입액\n\n"
        "감사합니다.\n"
    )

    result = send_email_with_attachments(
        to_email=RECIPIENT,
        subject=f"[매그나텍] 분기별 원자재 단가 추이 분석 ({quarters[0]}~{quarters[-1]})",
        body_text=body,
        attachments=[(fname, xlsx_bytes)],
        from_name="매그나텍 구매팀",
    )
    print(f"[발송] -> {RECIPIENT} / {result}")


if __name__ == "__main__":
    import sys
    main(send="--no-send" not in sys.argv)
