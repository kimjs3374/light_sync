"""삭제된 품번없는 BomItem 복구 (MTT/MTPS/MTPF/MTSOLAR/MT-SV만)

기존 import_bom_from_excel.py의 파싱 로직을 재사용하여
품번 없는(variant) 시트의 아이템을 복구합니다.
"""
import json, os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
from app import app
from modules.db_context import get_db
from modules.models import BomHeader, BomItem, Item

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'R1 제품 및 자재LIST(BOM+인건비).xlsx')

# 변형 시트 컬럼 (품번 없음: 품명+규격)
# col: 0=No, 1=cert, 2=product_code, 3=item_name, 4=item_spec, 5=qty, 6=supplier, 7=need_qty, 8=price, 9=amount
VARIANT_SHEETS = {10, 12, 13, 14, 15}

def safe_str(v):
    return str(v).strip() if v is not None else ''

def safe_float(v):
    try: return float(v)
    except: return None


def parse_variant_sheet(ws):
    """변형 시트 파싱: product_code -> [items]."""
    # 헤더 행 찾기
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=False):
        if row[0].value == 'No':
            header_row = row[0].row
            break
    if not header_row:
        return {}

    products = {}
    current = None
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        cells = [c.value for c in row[:12]]
        if all(v is None for v in cells):
            continue

        pc = safe_str(cells[2])
        if pc.upper() == 'TTL' or safe_str(cells[0]).upper() == 'TTL':
            continue
        if pc:
            current = pc

        if not current:
            continue

        item_name = safe_str(cells[3])
        if not item_name:
            continue

        item_spec = safe_str(cells[4])
        supplier = safe_str(cells[6])
        need_qty = safe_float(cells[7]) or 1
        unit_price = safe_float(cells[8])
        amount = safe_float(cells[9])

        products.setdefault(current, []).append({
            'item_name': item_name,
            'item_spec': item_spec,
            'supplier': supplier,
            'quantity': need_qty,
            'unit_price': unit_price,
            'amount': amount,
        })

    return products


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # 복구 대상 시트만 파싱
    all_products = {}
    for idx in VARIANT_SHEETS:
        ws = wb[wb.sheetnames[idx]]
        parsed = parse_variant_sheet(ws)
        print(f'Sheet {idx} ({wb.sheetnames[idx]}): {len(parsed)} products')
        all_products.update(parsed)

    with app.app_context():
        with get_db() as db:
            total_restored = 0

            for product_code, xl_items in sorted(all_products.items()):
                # DB에서 BOM 찾기 (product_code 정규화: 공백 제거)
                normalized = product_code.replace(' ', '')
                bom = db.query(BomHeader).filter(
                    BomHeader.product_code == normalized
                ).first()

                if not bom:
                    # 공백 포함 버전도 시도
                    bom = db.query(BomHeader).filter(
                        BomHeader.product_code == product_code
                    ).first()

                if not bom:
                    continue

                if not bom.is_active:
                    continue

                # 현재 BOM 아이템 이름 목록
                existing_names = {bi.item_name.strip() for bi in bom.bom_items if bi.item_name}

                added = 0
                for it in xl_items:
                    if it['item_name'] in existing_names:
                        continue

                    db.add(BomItem(
                        bom_id=bom.id,
                        item_code=None,
                        item_name=it['item_name'],
                        item_spec=it['item_spec'] or None,
                        quantity=it['quantity'],
                        unit_price=it['unit_price'],
                        amount=it['amount'],
                        supplier=it['supplier'] or None,
                    ))
                    existing_names.add(it['item_name'])
                    added += 1

                if added:
                    total_restored += added
                    total = len(bom.bom_items) + added
                    print(f'  {normalized}: +{added} restored (total ~{total})')

            db.commit()
            print(f'\nTotal restored: {total_restored}')


if __name__ == '__main__':
    main()
